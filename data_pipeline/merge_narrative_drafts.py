"""Merge narrative draft JSONL files into the panel datasets.

Reads every drafts/narrative_*.jsonl file, flattens each CoC's codes into
columns, and appends them to:

  * coc_analysis_ready.xlsx  — new sheet `narrative_codes`
  * coc_panel_long.csv       — additional rows with field_id = "{field}__{code}"

No reviewer-approval gate here; this is a *draft* integration for preview.
Promotion to the canonical dataset still requires a manual pass.
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from pipeline_utils import PIPELINE_DIR

DRAFTS = PIPELINE_DIR / "drafts"
ANALYSIS_XLSX = PIPELINE_DIR / "coc_analysis_ready.xlsx"
OUT_CSV = PIPELINE_DIR / "narrative_codes_flat.csv"


def iter_drafts():
    for p in sorted(DRAFTS.glob("narrative_*.jsonl")):
        with p.open() as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    yield json.loads(line)
                except json.JSONDecodeError:
                    continue


def main():
    # Collect: {(coc, year): {f"{field}__{code}": value}}
    rows = defaultdict(dict)
    code_names = set()
    extractors = defaultdict(set)
    for rec in iter_drafts():
        if rec.get("status") != "coded":
            continue
        key = (rec["coc_id"], rec["year"])
        field = rec["field_id"]
        for code, val in rec["codes"].items():
            if code == "evidence":
                continue
            if code == "summary":
                col = f"{field}__summary"
            else:
                col = f"{field}__{code}"
            # Flatten lists to semicolon-joined strings
            if isinstance(val, list):
                val = "; ".join(str(v) for v in val) if val else ""
            rows[key][col] = val
            code_names.add(col)
        extractors[rec.get("extractor", "unknown")].add(key)

    code_names = sorted(code_names)
    print(f"Loaded {len(rows)} (CoC, year) records across {len(code_names)} code columns")
    for ext, keys in extractors.items():
        print(f"  extractor {ext}: {len(keys)} records")

    # Flat CSV
    with OUT_CSV.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["coc_id", "year"] + code_names)
        for (coc, year), cols in sorted(rows.items()):
            w.writerow([coc, year] + [cols.get(c, "") for c in code_names])
    print(f"wrote {OUT_CSV}")

    # Append to analysis-ready xlsx
    if ANALYSIS_XLSX.exists():
        wb = openpyxl.load_workbook(ANALYSIS_XLSX)
        if "narrative_codes" in wb.sheetnames:
            del wb["narrative_codes"]
        ws = wb.create_sheet("narrative_codes")
        ws.append(["coc_id", "year"] + code_names)
        for (coc, year), cols in sorted(rows.items()):
            ws.append([coc, year] + [cols.get(c, "") for c in code_names])
        wb.save(ANALYSIS_XLSX)
        print(f"appended narrative_codes sheet to {ANALYSIS_XLSX.name}")


if __name__ == "__main__":
    main()
