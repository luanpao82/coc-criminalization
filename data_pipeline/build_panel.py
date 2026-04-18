"""Build 3-year panel + per-year datasets from extracted JSON records.

Strategy
--------
1. Each year gets its OWN wide dataset (only fields with reasonable coverage
   in that year are kept; rare-only fields dropped).
2. A field-map categorizes each canonical field into one of four buckets:
     - panel_safe   : ≥ 50% coverage in all 3 years  (good for longitudinal use)
     - mostly_panel : ≥ 50% in 2 of 3 years
     - year_specific: ≥ 50% coverage in only 1 year
     - sparse       : never reaches 50% coverage in any year
3. Panel (long-form) contains every (coc, year, field) triple where value is
   present. Wide panel uses `field` for panel_safe/mostly_panel columns and
   `field__yYY` suffix for year_specific columns.

Outputs
-------
coc_fy2022.{csv,xlsx}        — year-specific, raw extracted values
coc_fy2023.{csv,xlsx}
coc_fy2024.{csv,xlsx}
coc_panel_long.csv           — long form (coc_id, year, field_id, value)
coc_panel_wide.xlsx          — wide panel; columns split by category
panel_field_map.csv          — field category by year
panel_coverage.md            — coverage report
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from pipeline_utils import DATA_DIR, EXTRACTED_DIR, PIPELINE_DIR

XLSX_MANUAL = DATA_DIR / "coc_apps_all_info.xlsx"
YEARS = ("2022", "2023", "2024")

OUT_PER_YEAR = {y: PIPELINE_DIR / f"coc_fy{y}.xlsx" for y in YEARS}
OUT_LONG = PIPELINE_DIR / "coc_panel_long.csv"
OUT_WIDE = PIPELINE_DIR / "coc_panel_wide.xlsx"
OUT_FIELD_MAP = PIPELINE_DIR / "panel_field_map.csv"
OUT_COV = PIPELINE_DIR / "panel_coverage.md"


def load_canonical_fields() -> list[str]:
    wb = openpyxl.load_workbook(XLSX_MANUAL, data_only=True)
    ws = wb["2024"]
    fields = []
    for c in range(1, ws.max_column + 1):
        fid = ws.cell(row=1, column=c).value
        if fid:
            fields.append(str(fid))
    return fields


def load_extracted():
    """Return {(coc_id, year): {field_id: value}}."""
    out = {}
    for path in sorted(EXTRACTED_DIR.glob("*.json")):
        stem = path.stem
        parts = stem.rsplit("_", 1)
        if len(parts) != 2:
            continue
        coc_id, year = parts
        if year not in YEARS:
            continue
        try:
            recs = json.loads(path.read_text())
        except Exception:
            continue
        vals = {
            r["field_id"]: r["value"]
            for r in recs
            if r.get("value") is not None and not r.get("needs_review")
        }
        out[(coc_id, year)] = vals
    return out


def categorize(fields, data):
    """Return dict field_id -> category."""
    per_year_field = defaultdict(lambda: defaultdict(int))
    year_totals = defaultdict(int)
    for (coc, year), vals in data.items():
        year_totals[year] += 1
        for fid in fields:
            if fid in vals:
                per_year_field[year][fid] += 1

    cat = {}
    for fid in fields:
        rates = {y: (per_year_field[y][fid] / year_totals[y] if year_totals[y] else 0)
                 for y in YEARS}
        good = [y for y in YEARS if rates[y] >= 0.5]
        if len(good) == 3:
            cat[fid] = ("panel_safe", rates)
        elif len(good) == 2:
            cat[fid] = ("mostly_panel", rates)
        elif len(good) == 1:
            cat[fid] = ("year_specific", rates)
        else:
            cat[fid] = ("sparse", rates)
    return cat, year_totals


def write_year_xlsx(year, data, fields, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"fy{year}"
    # Keep only fields with at least 1 non-null this year
    year_rows = sorted([(c, vals) for (c, y), vals in data.items() if y == year])
    active_fields = [f for f in fields if any(f in vals for _, vals in year_rows)]
    ws.append(["coc_id"] + active_fields)
    for coc, vals in year_rows:
        ws.append([coc] + [vals.get(f, "") for f in active_fields])
    wb.save(path)
    return len(year_rows), len(active_fields)


def main():
    fields = load_canonical_fields()
    data = load_extracted()
    cat, year_totals = categorize(fields, data)

    print("CoCs extracted per year:")
    for y in YEARS:
        print(f"  FY{y}: {year_totals[y]} CoCs")

    # Category counts
    from collections import Counter
    ccount = Counter(v[0] for v in cat.values())
    print("Field categorization:")
    for k, v in ccount.most_common():
        print(f"  {k}: {v}")

    # Year-specific datasets
    for y in YEARS:
        rows, ncols = write_year_xlsx(y, data, fields, OUT_PER_YEAR[y])
        print(f"wrote {OUT_PER_YEAR[y].name} — {rows} rows × {ncols} active fields")

    # Long form
    with OUT_LONG.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["coc_id", "year", "field_id", "value", "field_category"])
        for (coc, y), vals in sorted(data.items()):
            for fid in fields:
                if fid in vals:
                    w.writerow([coc, y, fid, vals[fid], cat[fid][0]])
    long_rows = sum(1 for _ in open(OUT_LONG)) - 1
    print(f"wrote {OUT_LONG.name} — {long_rows:,} rows")

    # Wide panel
    # Columns: for panel_safe / mostly_panel → single column
    # For year_specific → field__yYY suffixed columns (only that year)
    # Sparse → drop
    wb = openpyxl.Workbook()

    # Sheet 1: panel_safe only (best for longitudinal modeling)
    ws = wb.active
    ws.title = "panel_safe"
    panel_safe_fields = [f for f in fields if cat[f][0] == "panel_safe"]
    ws.append(["coc_id", "year"] + panel_safe_fields)
    for (coc, y), vals in sorted(data.items()):
        ws.append([coc, y] + [vals.get(f, "") for f in panel_safe_fields])

    # Sheet 2: full wide (union of all fields; year-specific gets suffix)
    ws2 = wb.create_sheet("full_wide")
    header = ["coc_id", "year"]
    for f in fields:
        c, rates = cat[f]
        if c in ("panel_safe", "mostly_panel"):
            header.append(f)
        elif c == "year_specific":
            # Only the year where it exists gets a column
            years_with_field = [y for y in YEARS if rates[y] >= 0.5]
            for y in years_with_field:
                header.append(f"{f}__y{y}")
        # sparse: drop
    ws2.append(header)
    for (coc, y), vals in sorted(data.items()):
        row = [coc, y]
        for f in fields:
            c, rates = cat[f]
            if c in ("panel_safe", "mostly_panel"):
                row.append(vals.get(f, ""))
            elif c == "year_specific":
                for yy in YEARS:
                    if rates[yy] >= 0.5:
                        row.append(vals.get(f, "") if yy == y else "")
            # sparse: drop
        ws2.append(row)

    # Sheet 3: one sheet per year with full coverage for that year
    for y in YEARS:
        ws_y = wb.create_sheet(f"fy{y}")
        year_rows = sorted([(c, vals) for (c, yy), vals in data.items() if yy == y])
        active = [f for f in fields if any(f in vals for _, vals in year_rows)]
        ws_y.append(["coc_id"] + active)
        for coc, vals in year_rows:
            ws_y.append([coc] + [vals.get(f, "") for f in active])

    wb.save(OUT_WIDE)
    print(f"wrote {OUT_WIDE.name} (4 sheets: panel_safe / full_wide / fy2022 / fy2023 / fy2024)")

    # Field map CSV
    with OUT_FIELD_MAP.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["field_id", "category", "coverage_2022", "coverage_2023", "coverage_2024"])
        for fid in fields:
            c, rates = cat[fid]
            w.writerow([fid, c,
                        f"{rates['2022']:.3f}",
                        f"{rates['2023']:.3f}",
                        f"{rates['2024']:.3f}"])
    print(f"wrote {OUT_FIELD_MAP.name}")

    # Coverage report
    lines = [
        "# 3-Year Panel Coverage",
        "",
        f"- FY2022: **{year_totals['2022']}** CoCs extracted",
        f"- FY2023: **{year_totals['2023']}** CoCs extracted",
        f"- FY2024: **{year_totals['2024']}** CoCs extracted",
        f"- Total (coc, year) records: **{sum(year_totals.values())}**",
        "",
        "## Field categories",
        "",
        "| Category | Definition | Count |",
        "|---|---|---|",
        f"| `panel_safe` | ≥ 50% coverage in all 3 years (use for longitudinal) | {ccount['panel_safe']} |",
        f"| `mostly_panel` | ≥ 50% in 2 of 3 years | {ccount['mostly_panel']} |",
        f"| `year_specific` | ≥ 50% in only 1 year (use for cross-section) | {ccount['year_specific']} |",
        f"| `sparse` | < 50% in every year (extractor gap or form-specific rare field) | {ccount['sparse']} |",
        "",
        "## Criminalization variables (1d_4_*) coverage",
        "",
        "| field | 2022 | 2023 | 2024 | category |",
        "|---|---|---|---|---|",
    ]
    for fid in fields:
        if fid.startswith("1d_4_"):
            c, r = cat[fid]
            lines.append(
                f"| `{fid}` | {r['2022']:.0%} | {r['2023']:.0%} | {r['2024']:.0%} | {c} |"
            )

    lines += [
        "",
        "## Panel-safe fields (usable for 3-year longitudinal analysis)",
        "",
        "These fields have ≥50% coverage in all three years — panel models can use them directly.",
        "",
    ]
    panel_safe = [fid for fid in fields if cat[fid][0] == "panel_safe"]
    for fid in panel_safe:
        r = cat[fid][1]
        lines.append(f"- `{fid}` ({r['2022']:.0%} / {r['2023']:.0%} / {r['2024']:.0%})")

    lines += [
        "",
        "## Year-specific fields (cross-sectional use only)",
        "",
        "These fields are well-covered in one year but not others — typically because HUD changed the question.",
        "",
        "| field | 2022 | 2023 | 2024 |",
        "|---|---|---|---|",
    ]
    ys_fields = [(fid, cat[fid][1]) for fid in fields if cat[fid][0] == "year_specific"]
    for fid, r in ys_fields[:30]:
        lines.append(f"| `{fid}` | {r['2022']:.0%} | {r['2023']:.0%} | {r['2024']:.0%} |")
    if len(ys_fields) > 30:
        lines.append(f"| *...{len(ys_fields) - 30} more* | | | |")

    OUT_COV.write_text("\n".join(lines))
    print(f"wrote {OUT_COV.name}")


if __name__ == "__main__":
    main()
