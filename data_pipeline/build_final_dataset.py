"""Consolidate per-CoC extracted JSON records into one analysis-ready dataset.

Outputs three files mirroring the canonical FY2024 schema
(coc_apps_all_info.xlsx 2024 sheet):

  1. coc_auto_2024.csv          — extractor's proposed values only
  2. coc_auto_2024.xlsx         — same, xlsx format
  3. coc_merged_2024.xlsx       — manual ∪ auto, with per-cell provenance flag

Each row is one CoC; columns are the 331 canonical field IDs. The merged
file adds companion columns `<field>__source` ∈ {manual, auto, both_agree,
disagree} so coauthors can see where each cell came from.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from pipeline_utils import DATA_DIR, EXTRACTED_DIR, PIPELINE_DIR
from pilot_run import classify_field, compare

XLSX = DATA_DIR / "coc_apps_all_info.xlsx"
YEAR = "2024"
OUT_AUTO_XLSX = PIPELINE_DIR / "coc_auto_2024.xlsx"
OUT_AUTO_CSV = PIPELINE_DIR / "coc_auto_2024.csv"
OUT_MERGED_XLSX = PIPELINE_DIR / "coc_merged_2024.xlsx"


def load_schema():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2024"]
    header_rows = [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, 5)
    ]
    field_ids = header_rows[0]
    manual = {}
    for r in range(5, ws.max_row + 1):
        coc = ws.cell(row=r, column=1).value
        if not coc:
            continue
        manual[coc] = {
            fid: ws.cell(row=r, column=c + 1).value
            for c, fid in enumerate(field_ids)
            if fid
        }
    return header_rows, field_ids, manual


def load_auto():
    auto: dict[str, dict] = {}
    for path in sorted(EXTRACTED_DIR.glob(f"*_{YEAR}.json")):
        coc_id = path.stem.rsplit("_", 1)[0]
        recs = json.loads(path.read_text())
        auto[coc_id] = {
            r["field_id"]: r["value"]
            for r in recs
            if r.get("value") is not None
        }
    return auto


def main():
    header_rows, field_ids, manual = load_schema()
    auto = load_auto()

    coc_ids = sorted(set(manual) | set(auto))
    print(f"manual rows: {len(manual)}; auto rows: {len(auto)}; union: {len(coc_ids)}")

    # ---- File 1: auto-only dataset ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024_auto"
    # Preserve 4-row header block like the manual xlsx
    for ri, row in enumerate(header_rows, start=1):
        for ci, v in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=v)
    # Replace first header cell with something descriptive
    ws.cell(row=1, column=1, value="coc_id")
    for ri, coc in enumerate(coc_ids, start=5):
        ws.cell(row=ri, column=1, value=coc)
        vals = auto.get(coc, {})
        for ci, fid in enumerate(field_ids, start=1):
            if fid and fid in vals:
                ws.cell(row=ri, column=ci, value=vals[fid])
    wb.save(OUT_AUTO_XLSX)
    print(f"wrote {OUT_AUTO_XLSX}")

    # CSV variant
    import csv
    with OUT_AUTO_CSV.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["coc_id"] + [fid for fid in field_ids if fid and fid != field_ids[0]])
        for coc in coc_ids:
            vals = auto.get(coc, {})
            w.writerow([coc] + [vals.get(fid, "") for fid in field_ids if fid and fid != field_ids[0]])
    print(f"wrote {OUT_AUTO_CSV}")

    # ---- File 2: merged dataset with provenance ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024_merged"
    # Fills for provenance column colouring
    fill_manual = PatternFill("solid", fgColor="FFF4CC")   # yellow
    fill_auto = PatternFill("solid", fgColor="CCF4CC")     # green
    fill_agree = PatternFill("solid", fgColor="D6EBFF")    # light blue
    fill_disagree = PatternFill("solid", fgColor="FFCCCC") # red
    fill_missing = PatternFill("solid", fgColor="EEEEEE")  # grey

    # Double-width columns: for each field, value + source
    merged_cols = ["coc_id"]
    for fid in field_ids:
        if fid and fid != field_ids[0]:
            merged_cols.append(fid)
            merged_cols.append(f"{fid}__source")
    for ci, name in enumerate(merged_cols, start=1):
        ws.cell(row=1, column=ci, value=name)

    counts = {"manual_only": 0, "auto_only": 0, "both_agree": 0, "disagree": 0, "missing": 0}
    for ri, coc in enumerate(coc_ids, start=2):
        ws.cell(row=ri, column=1, value=coc)
        mrow = manual.get(coc, {})
        arow = auto.get(coc, {})
        col = 2
        for fid in field_ids:
            if not fid or fid == field_ids[0]:
                continue
            mval = mrow.get(fid)
            aval = arow.get(fid)
            cell = ws.cell(row=ri, column=col)
            src_cell = ws.cell(row=ri, column=col + 1)
            if mval not in (None, "") and aval not in (None, ""):
                klass = classify_field(fid)
                if klass == "unknown":
                    # No class-specific comparator; compare as strings
                    ok = str(mval).strip() == str(aval).strip()
                    reason = ""
                else:
                    ok, reason = compare(mval, aval, klass)
                if ok:
                    cell.value = aval
                    src_cell.value = "both_agree"
                    cell.fill = fill_agree
                    counts["both_agree"] += 1
                else:
                    cell.value = mval  # prefer manual when they disagree; auto is in diff file
                    src_cell.value = f"disagree(auto={aval})"
                    cell.fill = fill_disagree
                    counts["disagree"] += 1
            elif mval not in (None, ""):
                cell.value = mval
                src_cell.value = "manual"
                cell.fill = fill_manual
                counts["manual_only"] += 1
            elif aval not in (None, ""):
                cell.value = aval
                src_cell.value = "auto"
                cell.fill = fill_auto
                counts["auto_only"] += 1
            else:
                src_cell.value = "missing"
                cell.fill = fill_missing
                counts["missing"] += 1
            col += 2
    wb.save(OUT_MERGED_XLSX)
    print(f"wrote {OUT_MERGED_XLSX}")
    print()
    print("Per-cell provenance counts (merged dataset):")
    for k, v in counts.items():
        print(f"  {k}: {v:,}")
    total = sum(counts.values())
    print(f"  total: {total:,}")


if __name__ == "__main__":
    main()
