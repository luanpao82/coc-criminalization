"""Build codebook.md from the manual xlsx and a sample 2024 PDF.

Inputs
------
- coc_apps_all_info.xlsx
  - row 1: field_id (e.g., "1b_1_1_meetings")
  - rows 2-4: section / subsection / label hierarchy
  - rows 5+: observed data (used to infer value domains)
- AL_500_2024.pdf as representative FY2024 source (for question text).

Output
------
codebook.md — per-section tables of variables with:
  field_id | section | label | observed values | inferred type | question id (2024)

This is a *first-pass* codebook. It infers the value domain from what coders
already entered. Human review is required before it becomes canonical.
"""
from __future__ import annotations

import re
import subprocess
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

DATA_DIR = Path(
    "/Users/seongho/Library/CloudStorage/OneDrive-UniversityofCentralFlorida/"
    "Desktop/Prof. An/01 Research/2026 CoC Anti-Criminalization on Homelessness/Data"
)
XLSX = DATA_DIR / "coc_apps_all_info.xlsx"
SAMPLE_PDF = DATA_DIR / "AL_500_2024.pdf"
OUT = Path(
    "/Users/seongho/Desktop/Obsidian/10-Research/coc_criminalization/codebook.md"
)

CATEGORICAL_VOCAB = {"Yes", "No", "Nonexistent", "N/A", "NA"}
DESIGNATION_VOCAB = {"CA", "UFA", "CoC"}


def load_schema():
    """Return list of dicts for every column in sheet '2024'."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2024"]
    cols = []
    for c in range(1, ws.max_column + 1):
        field_id = ws.cell(row=1, column=c).value
        section = ws.cell(row=2, column=c).value
        sub = ws.cell(row=3, column=c).value
        label = ws.cell(row=4, column=c).value
        if not field_id:
            continue
        # Sample observed values
        vals = []
        for r in range(5, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            vals.append(v)
        cols.append(
            {
                "field_id": field_id,
                "section": section,
                "sub": sub,
                "label": label,
                "values": vals,
            }
        )
    return cols


def classify(col):
    vals = col["values"]
    if not vals:
        return "UNOBSERVED", ""
    # string vs numeric
    as_str = [str(v).strip() for v in vals]
    uniq = set(as_str)

    # Booleans / controlled vocab
    if uniq.issubset(CATEGORICAL_VOCAB | {""}):
        return "categorical", " | ".join(sorted(uniq - {""}))
    if uniq.issubset(DESIGNATION_VOCAB | {""}):
        return "categorical_designation", " | ".join(sorted(uniq - {""}))

    # Numeric
    try:
        nums = [float(v) for v in vals if str(v).strip() != ""]
        if nums:
            # integer-like?
            if all(abs(n - round(n)) < 1e-9 for n in nums):
                kind = "integer"
            else:
                kind = "numeric"
            return kind, f"min={min(nums)}; max={max(nums)}; n={len(nums)}"
    except (ValueError, TypeError):
        pass

    # Percent? "%" in most values
    pct_like = [v for v in as_str if "%" in v]
    if pct_like and len(pct_like) / len(as_str) > 0.5:
        return "percent", f"examples: {pct_like[:3]}"

    # Long free text -> narrative
    avg_len = sum(len(s) for s in as_str) / len(as_str)
    if avg_len > 120:
        return "narrative", f"avg_chars~{int(avg_len)}; examples trimmed"

    # Short free text — identifier / label
    top = Counter(as_str).most_common(5)
    return "label", "; ".join(f"{k!r}×{n}" for k, n in top)


def qnum_to_field(field_id: str) -> str:
    """Heuristic mapping from field_id pattern to question number.

    '1b_1_1_meetings' -> '1B-1 row 1 (meetings)'
    '1d_4_1_policymakers' -> '1D-4 row 1 (policymakers)'
    '2a_5_3_coverage' -> '2A-5 row 3 (coverage)'
    """
    parts = field_id.lower().split("_")
    if not parts:
        return ""
    head = parts[0]
    m = re.match(r"^(\d)([a-z])$", head)
    if not m:
        return ""
    section = f"{m.group(1)}{m.group(2).upper()}"
    rest = parts[1:]
    # If rest starts with an integer-like token, assume primary question
    primary = rest[0] if rest else ""
    qid = f"{section}-{primary}" if primary else section
    # Check for letter suffix (e.g., 1c_4a_, 1c_4c_1_mou -> 1C-4c row 1 (mou))
    if len(rest) > 1:
        sub = "_".join(rest[1:])
        return f"{qid} · {sub}"
    return qid


def extract_pdf_anchors(pdf_path: Path) -> dict:
    """Collect question number -> question title from a sample FY2024 PDF."""
    txt = subprocess.run(
        ["pdftotext", "-layout", str(pdf_path), "-"],
        capture_output=True,
        text=True,
        timeout=120,
    ).stdout
    # Capture patterns like "1B-1." or "1B-1a." followed by title text until \n
    anchors = {}
    for m in re.finditer(
        r"\b(\d[A-E]-\d+[a-z]?)(?:\.\d+)?\.\s+([^\n\r]{3,200})", txt
    ):
        qid = m.group(1)
        title = m.group(2).strip()
        # Take the first clean occurrence
        if qid not in anchors:
            anchors[qid] = title
    return anchors


def main():
    cols = load_schema()
    anchors = extract_pdf_anchors(SAMPLE_PDF)

    # Group by section label (row 2)
    by_section = defaultdict(list)
    for c in cols:
        by_section[c["section"] or "(uncategorized)"].append(c)

    out = []
    out.append("---")
    out.append("title: Codebook — HUD CoC Application Variables (v0.1 draft)")
    out.append("project: CoC Criminalization / PLE Engagement (Lee & Kim, UCF)")
    out.append("status: working draft — PI review required before adoption")
    out.append("last_updated: 2026-04-17")
    out.append("schema_version: FY2024")
    out.append("related:")
    out.append('  - "[[data_construction_methodology]]"')
    out.append('  - "[[KL_From_Inclusion_to_Influence]]"')
    out.append("---")
    out.append("")
    out.append("# Codebook")
    out.append("")
    out.append(
        "This codebook catalogues every variable in the canonical FY2024 schema "
        "(~331 columns), organized by HUD application section. Value domains are "
        "inferred from the manually coded `coc_apps_all_info.xlsx` and should be "
        "treated as a **first-pass draft**: observed values reflect what coders "
        "entered, not necessarily what is permitted. PIs must review and lock the "
        "controlled vocabulary before the iteration loop begins."
    )
    out.append("")
    out.append("## Missing-value conventions")
    out.append("")
    out.append(
        "| Convention | Meaning |\n"
        "|---|---|\n"
        "| `Nonexistent` | HUD option: the entity/process does not exist in the CoC's geographic area |\n"
        "| `N/A` | Question is skipped because a prior gate answer excluded it |\n"
        "| `NA_not_asked_this_year` | Variable did not exist in this year's form (crosswalk-derived) |\n"
        "| blank | Source PDF is genuinely missing a value — needs human review |\n"
    )
    out.append("")
    out.append("## Variable type taxonomy")
    out.append("")
    out.append(
        "| Code | Class | Example | Validation |\n"
        "|---|---|---|---|\n"
        "| `categorical` | A | Yes/No/Nonexistent | Exact match against locked vocab |\n"
        "| `categorical_designation` | A | CA / UFA | Exact match |\n"
        "| `integer` | B | PIT count, HIC beds | Exact match after unit normalization |\n"
        "| `numeric` | B | Percentages, dollars | Exact within rounding tolerance |\n"
        "| `percent` | B | Bed coverage % | Normalized to decimal |\n"
        "| `label` | C | CoC name, PHA name | Case/whitespace-normalized string match |\n"
        "| `narrative` | D | 1B-1a racial-equity response | Reviewer acceptance + verbatim evidence |\n"
        "| `UNOBSERVED` | — | No value seen in manual xlsx | PI must specify expected type |\n"
    )
    out.append("")

    for section, items in by_section.items():
        out.append(f"## {section}")
        out.append("")
        out.append("| field_id | sub-group | label | type | observed domain | likely Q (FY2024) |")
        out.append("|---|---|---|---|---|---|")
        for c in items:
            kind, domain = classify(c)
            # Map heuristic to probable question number; verify with anchors
            probable = qnum_to_field(c["field_id"])
            # Extract just the leading "1B-1" part for anchor lookup
            head = probable.split(" ")[0] if probable else ""
            anchor_title = anchors.get(head.split("·")[0].strip(), "")
            label = (c["label"] or "").replace("|", "\\|").replace("\n", " ")
            sub = (c["sub"] or "").replace("|", "\\|").replace("\n", " ")
            domain = domain.replace("|", "\\|").replace("\n", " ")
            q_disp = probable
            if anchor_title and head in anchors:
                q_disp = f"{probable} — *{anchor_title[:60]}*"
            out.append(
                f"| `{c['field_id']}` | {sub} | {label} | {kind} | {domain} | {q_disp} |"
            )
        out.append("")

    out.append("## Variables flagged for PI adjudication")
    out.append("")
    flags = []
    for c in cols:
        kind, domain = classify(c)
        if kind == "UNOBSERVED":
            flags.append(f"- `{c['field_id']}` ({c['label']}) — no manual data observed")
        elif kind == "label" and c["label"] and "name" not in (c["label"] or "").lower():
            # Short label-type variables that aren't clearly name-ish are ambiguous
            pass
    if flags:
        out.extend(flags)
    else:
        out.append("*(none at first pass)*")
    out.append("")

    out.append("## Narrative variables — require Stage-2 LLM extraction + reviewer approval")
    out.append("")
    for c in cols:
        kind, _ = classify(c)
        if kind == "narrative":
            out.append(f"- `{c['field_id']}` — {c['label']}")
    out.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(out))
    print(f"wrote codebook -> {OUT}")
    print(f"variables: {len(cols)}; sections: {len(by_section)}")


if __name__ == "__main__":
    main()
