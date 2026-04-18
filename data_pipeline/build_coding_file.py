"""Build the research-specific coding file WITH inline coding columns.

This is both an extract of `coc_raw_data.xlsx` and the workbook where the
LLM-derived structured codes + composite indices are recorded.

Columns per row (one per CoC × year):
  meta: coc_id, year, original_filename, is_scanned
  for each of 3 PLE fields — status, page, text, then atomic codes + summary + evidence
  aggregates (computed at end of row): ple_representation, ple_compensation,
    ple_feedback_loop, ple_institution

If `drafts/pilot_ple_codes.jsonl` exists, already-coded records are merged
into this workbook so the reviewer can see coded + uncoded rows side by side.

Output: coc_raw_for_coding.xlsx
"""
from __future__ import annotations
import json, sys
from pathlib import Path
import pandas as pd

HERE = Path(__file__).parent
RAW_PATH = HERE / "coc_raw_data.xlsx"
CODES_JSONL = HERE / "pilot_ple_codes.jsonl"
OUT_PATH = HERE / "coc_raw_for_coding.xlsx"

# Atomic code definitions per field
ATOMIC_CODES = {
    "ple_umbrella": [
        "on_board", "on_committees", "compensated",
        "hiring_advertised", "formal_policy", "decisionmaking",
    ],
    "ple_prof_dev": [
        "paid_positions", "comp_policy_formal", "training_pipeline",
        "career_advancement", "scope_beyond_tokenism",
    ],
    "ple_feedback": [
        "mechanism_formal", "acts_on_feedback",
        "addresses_barriers", "closes_the_loop",
    ],
}

# How atomic codes map to the JSON schema keys used by extract_narrative.py
CODE_KEY_MAP = {
    ("ple_umbrella", "on_board"):           "ple_on_board",
    ("ple_umbrella", "on_committees"):      "ple_on_committees",
    ("ple_umbrella", "compensated"):        "ple_compensated",
    ("ple_umbrella", "hiring_advertised"):  "ple_hiring_advertised",
    ("ple_umbrella", "formal_policy"):      "formal_policy",
    ("ple_umbrella", "decisionmaking"):     "decisionmaking_authority",
    ("ple_prof_dev", "paid_positions"):       "paid_positions_exist",
    ("ple_prof_dev", "comp_policy_formal"):   "compensation_policy_formal",
    ("ple_prof_dev", "training_pipeline"):    "training_pipeline_described",
    ("ple_prof_dev", "career_advancement"):   "career_advancement_described",
    ("ple_prof_dev", "scope_beyond_tokenism"):"scope_beyond_tokenism",
    ("ple_feedback", "mechanism_formal"):     "feedback_mechanism_formal",
    ("ple_feedback", "acts_on_feedback"):     "acts_on_feedback",
    ("ple_feedback", "addresses_barriers"):   "addresses_barriers",
    ("ple_feedback", "closes_the_loop"):      "closes_the_loop",
}

# Aggregate index definitions
INDEX_DEFS = {
    "ple_representation": [
        ("ple_umbrella", "on_board"), ("ple_umbrella", "on_committees"),
        ("ple_umbrella", "formal_policy"), ("ple_umbrella", "decisionmaking"),
    ],
    "ple_compensation": [
        ("ple_umbrella", "compensated"),
        ("ple_prof_dev", "paid_positions"),
        ("ple_prof_dev", "comp_policy_formal"),
    ],
    "ple_feedback_loop": [
        ("ple_feedback", "mechanism_formal"),
        ("ple_feedback", "acts_on_feedback"),
        ("ple_feedback", "addresses_barriers"),
        ("ple_feedback", "closes_the_loop"),
    ],
}


def bool_to_int(v):
    if isinstance(v, bool):
        return 1 if v else 0
    if v is None:
        return 0  # Krippendorff 2018 §4.3 direct inference: absent evidence → 0
    return 0


def load_existing_codes():
    """{(coc_id, year): {field_id: codes_dict}}"""
    out = {}
    if not CODES_JSONL.exists():
        return out
    for line in CODES_JSONL.open():
        r = json.loads(line)
        key = (r["coc_id"], int(r["year"]))
        out.setdefault(key, {})[r["field_id"]] = r.get("codes", {})
    return out


def main():
    raw = pd.read_excel(RAW_PATH, sheet_name="raw_data")
    print(f"Loaded raw: {len(raw)} rows × {len(raw.columns)} cols")

    existing_codes = load_existing_codes()
    print(f"Loaded {sum(len(v) for v in existing_codes.values())} existing code records "
          f"across {len(existing_codes)} CoC-years")

    # Build output columns per row
    records = []
    for _, r in raw.iterrows():
        try:
            year = int(r["year"])
        except Exception:
            continue
        if year not in (2022, 2023, 2024):
            continue

        key = (r["coc_id"], year)
        codes_for_row = existing_codes.get(key, {})

        row = {
            "coc_id": r["coc_id"],
            "year": year,
            "original_filename": r["original_filename"],
            "is_scanned": r["is_scanned"],
        }

        # For each PLE field, add status/page/text, then atomic codes + summary + evidence
        field_to_canonical = {
            "ple_umbrella": "1d_10",
            "ple_prof_dev": "1d_10b",
            "ple_feedback": "1d_10c",
        }
        for fld, canon in field_to_canonical.items():
            row[f"{fld}_status"] = r.get(f"{fld}_status", "")
            row[f"{fld}_page"] = r.get(f"{fld}_page")
            txt = r.get(canon)
            row[f"{fld}_text"] = txt if isinstance(txt, str) else ""

            codes = codes_for_row.get(fld, {})
            for atom in ATOMIC_CODES[fld]:
                key_in_json = CODE_KEY_MAP[(fld, atom)]
                v = codes.get(key_in_json)
                if v is True:
                    row[f"{fld}__{atom}"] = 1
                elif v is False:
                    row[f"{fld}__{atom}"] = 0
                elif v is None and codes:  # coded but absent → 0
                    row[f"{fld}__{atom}"] = 0
                else:
                    row[f"{fld}__{atom}"] = ""  # not yet coded
            # feedback also has a frequency string
            if fld == "ple_feedback":
                row[f"{fld}__frequency"] = codes.get("feedback_frequency") or ""
            row[f"{fld}__summary"] = codes.get("summary", "") if codes else ""
            ev = codes.get("evidence", []) if codes else []
            row[f"{fld}__evidence"] = "\n\n".join(f"[{i+1}] {q}" for i, q in enumerate(ev)) if ev else ""

        # Composite indices: only if all 3 fields have been coded
        all_coded = all(codes_for_row.get(f) for f in ATOMIC_CODES)
        if all_coded:
            for idx_name, atoms in INDEX_DEFS.items():
                bools = []
                for (fld, atom) in atoms:
                    val = codes_for_row.get(fld, {}).get(CODE_KEY_MAP[(fld, atom)])
                    bools.append(bool_to_int(val))
                row[idx_name] = round(sum(bools) / len(bools), 3) if bools else ""
            row["ple_institution"] = round(
                sum(row[i] for i in INDEX_DEFS) / len(INDEX_DEFS), 3)
        else:
            for idx_name in INDEX_DEFS:
                row[idx_name] = ""
            row["ple_institution"] = ""

        # Reviewer fields
        row["reviewer_accept"] = ""
        row["reviewer_notes"] = ""
        records.append(row)

    out = pd.DataFrame(records).sort_values(["coc_id", "year"]).reset_index(drop=True)

    # Coverage summary
    n_fully_coded = (out["ple_institution"] != "").sum()
    print(f"\nFully coded CoC-years: {n_fully_coded}/{len(out)}")

    # ---- Write xlsx with formatting ----
    from openpyxl.styles import Alignment, PatternFill, Font
    print(f"Writing {OUT_PATH.name} ...")
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name="narratives_for_coding", index=False)
        ws = xw.sheets["narratives_for_coding"]

        text_cols = {f"{f}_text" for f in ATOMIC_CODES}
        status_cols = {f"{f}_status" for f in ATOMIC_CODES}
        summary_cols = {f"{f}__summary" for f in ATOMIC_CODES}
        evidence_cols = {f"{f}__evidence" for f in ATOMIC_CODES}
        atom_cols = set()
        for f, atoms in ATOMIC_CODES.items():
            for a in atoms:
                atom_cols.add(f"{f}__{a}")
        atom_cols.add("ple_feedback__frequency")
        index_cols = set(INDEX_DEFS.keys()) | {"ple_institution"}

        for idx, col in enumerate(out.columns, start=1):
            letter = ws.cell(row=1, column=idx).column_letter
            if col in text_cols:
                ws.column_dimensions[letter].width = 55
            elif col in evidence_cols:
                ws.column_dimensions[letter].width = 50
            elif col in summary_cols:
                ws.column_dimensions[letter].width = 45
            elif col in status_cols:
                ws.column_dimensions[letter].width = 22
            elif col in atom_cols:
                ws.column_dimensions[letter].width = 10
            elif col in index_cols:
                ws.column_dimensions[letter].width = 12
            elif col == "original_filename":
                ws.column_dimensions[letter].width = 20
            elif col == "reviewer_notes":
                ws.column_dimensions[letter].width = 25
            else:
                ws.column_dimensions[letter].width = 12

        red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        green = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        pale_blue = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        pale_yellow = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        bold = Font(bold=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col_name = ws.cell(row=1, column=cell.column).value
                if col_name in text_cols | summary_cols | evidence_cols and cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_name in status_cols and cell.value:
                    if cell.value == "ok":
                        cell.fill = green
                    else:
                        cell.fill = red; cell.font = bold
                if col_name in atom_cols and cell.value != "":
                    cell.fill = pale_blue
                if col_name in index_cols and cell.value != "":
                    cell.fill = pale_yellow; cell.font = bold

        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "C2"

    print(f"\nWrote {OUT_PATH.name} ({OUT_PATH.stat().st_size / 1024:.0f} KB)")
    print(f"  {len(out)} rows × {len(out.columns)} columns")


if __name__ == "__main__":
    main()
