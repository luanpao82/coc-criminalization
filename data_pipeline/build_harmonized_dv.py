"""Compute harmonized criminalization DV variables across FY2022–FY2024.

Why this exists
---------------
The HUD 1D-4 instrument changed fundamentally between FY2023 and FY2024.
Rows and columns represent different constructs. See
`dv_harmonization_strategies.md` for the full treatment.

This script produces three derived DVs per (CoC, year) that ARE
comparable across all three years:

  crim_activity_index    — proportion of 1D-4 cells = Yes (0.0–1.0)
  engaged_policymakers   — binary: any policymaker-engagement Yes
  engaged_law_enforce    — binary: any law-enforcement-related Yes
  implemented_practice   — binary: any "implementation" column Yes

Field-level mapping (what the extractor wrote vs. what it meant per year)
------------------------------------------------------------------------
FY2024 1D-4:
    row r ∈ {1: co-responder, 2: minimize LE, 3: avoid sanctions}
    col c ∈ {policymakers: "engaged legislators", prevent_crim: "implemented laws"}

FY2022/2023 1D-4 (extractor captured rows 1–3 only, row 4 dropped):
    row r ∈ {1: engaged policymakers, 2: engaged law enforcement,
             3: engaged business leaders}
    col c ∈ {policymakers: "ensure not criminalized",
             prevent_crim: "reverse existing"}

Consequence: the *column suffix* in the extracted field IDs means
something different per year. This script reads the raw values and
applies the year-specific interpretation before building the harmonized
binaries.

Outputs
-------
harmonized_dv.csv                    — one row per (CoC, year) × 4 DVs
harmonized_dv.xlsx                   — same, xlsx
coc_panel_wide.xlsx[dv_harmonized]   — new sheet appended to the wide panel
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

from pipeline_utils import EXTRACTED_DIR, PIPELINE_DIR

OUT_CSV = PIPELINE_DIR / "harmonized_dv.csv"
OUT_XLSX = PIPELINE_DIR / "harmonized_dv.xlsx"
PANEL_WIDE = PIPELINE_DIR / "coc_panel_wide.xlsx"


def load_records():
    by_cocyear: dict[tuple[str, str], dict] = {}
    for path in sorted(EXTRACTED_DIR.glob("*.json")):
        stem = path.stem
        parts = stem.rsplit("_", 1)
        if len(parts) != 2:
            continue
        coc, yr = parts
        try:
            recs = json.loads(path.read_text())
        except Exception:
            continue
        vals = {r["field_id"]: r["value"] for r in recs
                if r.get("value") is not None and not r.get("needs_review")}
        by_cocyear[(coc, yr)] = vals
    return by_cocyear


def _yes(v) -> bool:
    return str(v).strip().lower() in {"yes", "y"}


def _no(v) -> bool:
    return str(v).strip().lower() in {"no", "n", "nonexistent"}


def compute_dv(vals: dict, year: str) -> dict:
    """Return the four harmonized DVs for a single (CoC, year) record."""
    # Cells we have (up to 6 per CoC-year): 1d_4_{1,2,3}_{policymakers,prevent_crim}
    cells = []
    for r in (1, 2, 3):
        for c in ("policymakers", "prevent_crim"):
            key = f"1d_4_{r}_{c}"
            if key in vals:
                cells.append((r, c, vals[key]))

    n_yes = sum(1 for _, _, v in cells if _yes(v))
    n_seen = sum(1 for _, _, v in cells if _yes(v) or _no(v))

    crim_activity_index = (n_yes / n_seen) if n_seen else None

    # Year-specific construct mapping:
    if year == "2024":
        # Row ∈ {1: co-responder, 2: minimize LE, 3: avoid sanctions}
        # Col: policymakers = "engaged legislators", prevent_crim = "implemented laws"
        engaged_policymakers = any(
            _yes(v) for r, c, v in cells if c == "policymakers"
        )
        # Law enforcement relevance: rows 1 (co-responder) and 2 (minimize LE)
        engaged_law_enforce = any(
            _yes(v) for r, c, v in cells if r in (1, 2)
        )
        # Implemented / enacted policy dimension
        implemented_practice = any(
            _yes(v) for r, c, v in cells if c == "prevent_crim"
        )
    else:  # FY2022 or FY2023
        # Row ∈ {1: policymakers, 2: law enforcement, 3: business leaders}
        # Col: policymakers = "ensure not criminalized",
        #      prevent_crim = "reverse existing policies"
        engaged_policymakers = any(
            _yes(v) for r, c, v in cells if r == 1
        )
        engaged_law_enforce = any(
            _yes(v) for r, c, v in cells if r == 2
        )
        # Reverse-existing column treated as the "implemented" dimension
        implemented_practice = any(
            _yes(v) for r, c, v in cells if c == "prevent_crim"
        )

    return {
        "cells_observed": n_seen,
        "cells_yes": n_yes,
        "crim_activity_index": None if crim_activity_index is None else round(crim_activity_index, 3),
        "engaged_policymakers_crim": int(engaged_policymakers) if n_seen else None,
        "engaged_law_enforce_crim": int(engaged_law_enforce) if n_seen else None,
        "implemented_anticrim_practice": int(implemented_practice) if n_seen else None,
    }


def main():
    by_cy = load_records()
    rows = []
    for (coc, yr), vals in sorted(by_cy.items()):
        dv = compute_dv(vals, yr)
        # Only record when we actually observed any 1d_4_ cells
        if dv["cells_observed"] == 0:
            continue
        rows.append({
            "coc_id": coc, "year": yr, **dv,
        })

    # CSV
    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {OUT_CSV} — {len(rows)} (CoC, year) records")

    # Standalone xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "harmonized_dv"
    ws.append(list(rows[0].keys()))
    for r in rows:
        ws.append([r[k] for k in rows[0].keys()])
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    # Append to coc_panel_wide.xlsx
    if PANEL_WIDE.exists():
        wb = openpyxl.load_workbook(PANEL_WIDE)
        if "dv_harmonized" in wb.sheetnames:
            del wb["dv_harmonized"]
        ws = wb.create_sheet("dv_harmonized")
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append([r[k] for k in rows[0].keys()])
        wb.save(PANEL_WIDE)
        print(f"updated {PANEL_WIDE} (new sheet: dv_harmonized)")

    # Summary stats
    from collections import defaultdict
    by_year = defaultdict(list)
    for r in rows:
        by_year[r["year"]].append(r)
    print()
    print("=== Distribution by year ===")
    print(f"{'year':<6} {'n':>4} {'activity_mean':>14} {'eng_policymaker_rate':>22} {'eng_le_rate':>13} {'implemented_rate':>18}")
    for yr in ("2022", "2023", "2024"):
        rs = by_year[yr]
        if not rs:
            continue
        aidx = [r["crim_activity_index"] for r in rs if r["crim_activity_index"] is not None]
        poly = [r["engaged_policymakers_crim"] for r in rs if r["engaged_policymakers_crim"] is not None]
        le = [r["engaged_law_enforce_crim"] for r in rs if r["engaged_law_enforce_crim"] is not None]
        imp = [r["implemented_anticrim_practice"] for r in rs if r["implemented_anticrim_practice"] is not None]
        print(
            f"{yr:<6} {len(rs):>4} {sum(aidx)/len(aidx):>14.3f} "
            f"{sum(poly)/len(poly):>22.3f} {sum(le)/len(le):>13.3f} "
            f"{sum(imp)/len(imp):>18.3f}"
        )


if __name__ == "__main__":
    main()
