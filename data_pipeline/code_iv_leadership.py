"""Classify each Collaborative Applicant (1a_2) into lead-agency type.

Output: coc_id → {nonprofit, city_govt, county_govt, state_govt, tribal, other}
Saved to iv_leadership.csv and joined into coc_analysis_ready.xlsx
as `lead_agency_type` + `nonprofit_led` (binary).

This is a rule-based first pass — confident matches are auto-classified,
ambiguous ones are flagged as `other` for manual review.
"""
from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path

import openpyxl

from pipeline_utils import PIPELINE_DIR, DATA_DIR

OUT = PIPELINE_DIR / "iv_leadership.csv"
ANALYSIS_XLSX = PIPELINE_DIR / "coc_analysis_ready.xlsx"

# Strong cues per category (checked in priority order)
RULES = [
    ("state_govt", [
        r"\bdepartment of\b.*(commerce|community|housing|social|human|health|mental|veterans)",
        r"\bstate of\b",
        r"^(arizona|alaska|maine|new hampshire|vermont|wisconsin|west virginia|idaho|wyoming|connecticut|delaware|montana|hawaii|rhode island|puerto rico).*(department|commission|authority|agency)",
        r"\b(statewide|state agency|state authority)\b",
    ]),
    ("county_govt", [
        r"\bcounty of\b",
        r"^county[,\s]",
        r"\s+county\s+(of\s+)?(?:department|agency|authority|services|housing)",
        r"\bcounty department\b",
        r"^[A-Za-z]+ County$",
        r"^[A-Za-z]+,?\s+County of",
    ]),
    ("city_govt", [
        r"\bcity of\b",
        r"\bcity and county of\b",
        r"\btown of\b",
        r"\bmunicipal\b",
    ]),
    ("tribal", [
        r"\btribal\b",
        r"\bnation\b.*(housing|council)",
        r"\bindian\b",
    ]),
    ("housing_authority", [
        r"\bhousing authority\b",
        r"\bpublic housing agency\b",
        r"\bHACA?\b",
    ]),
    ("nonprofit", [
        r"\binc\.?\s*$",
        r"\binc,\s*$",
        r"\bincorporated\b",
        r"\bfoundation\b",
        r"\bcoalition\b",
        r"\balliance\b",
        r"\bunited way\b",
        r"\binstitute\b",
        r"\bcouncil\b",
        r"\bpartnership\b",
        r"\bservices?\s+(?:inc|organization|corporation)",
        r"\bsteps forward\b",
        r"\bpartners?\s+(?:in|for|of)\b",
        r"\bcommunity\s+(?:action|services|alliance|center)\b",
        r"\bnonprofit\b",
        r"\bnon-profit\b",
        r"\bnot-for-profit\b",
        r"\bcontinuum of care\b",
        r"\bhomeless\b.*(network|alliance|council|coalition|authority|association|organization|services)",
        r"\b(homeless|housing).*(council|alliance|coalition|trust)",
        r"\b(ministries|outreach|shelter)\b",
        r"\busa\s+inc\b",
        r"\bchamber of commerce\b",
        r"\bchamber for\b",
        r"^[a-z]+\s+roof$",  # "One Roof"
        r"\bhouse\b.*(organization|services)",
        r"\bcatholic\b|\bchristian\b|\bjewish\b|\bunited methodist\b",
        r"\bY(M|W)CA\b",
    ]),
    ("regional_govt", [
        r"\bassociation of governments\b",
        r"\bcouncil of governments\b",
        r"\bregional planning\b",
        r"\bmetropolitan\b.*(planning|council|commission)",
    ]),
]


# Manual overrides for CoC IDs whose collaborative applicant name doesn't
# clearly signal lead-agency type. Classifications below are based on
# cross-checking the organization via HUD PHA/CoC crosswalk and the
# organization's own self-description.
MANUAL_OVERRIDES = {
    # nonprofits
    "CA-521": "nonprofit", "CA-524": "nonprofit", "CA-527": "nonprofit",
    "CA-601": "nonprofit", "CA-604": "nonprofit", "CO-503": "nonprofit",
    "CT-503": "nonprofit", "FL-501": "nonprofit", "FL-508": "nonprofit",
    "FL-511": "nonprofit", "GA-506": "nonprofit", "GA-507": "nonprofit",
    "IL-510": "nonprofit", "IL-512": "nonprofit", "IL-515": "nonprofit",
    "IL-519": "nonprofit", "LA-503": "nonprofit", "MA-515": "nonprofit",
    "MI-523": "nonprofit", "MN-504": "nonprofit", "MO-602": "nonprofit",
    "MS-501": "nonprofit", "NC-516": "nonprofit", "NH-502": "nonprofit",
    "NJ-516": "nonprofit", "NY-513": "nonprofit", "OH-501": "nonprofit",
    "OK-500": "nonprofit", "OR-502": "nonprofit", "PA-508": "nonprofit",
    "SC-500": "nonprofit", "SC-501": "nonprofit", "TN-506": "nonprofit",
    "TX-600": "nonprofit", "VA-500": "nonprofit",
    # cities
    "GA-504": "city_govt", "ID-500": "city_govt",
    # counties
    "CA-502": "county_govt", "CA-505": "county_govt", "CA-510": "county_govt",
    "CA-511": "county_govt", "CA-512": "county_govt", "CA-522": "county_govt",
    "CA-523": "county_govt", "CA-529": "county_govt", "FL-514": "county_govt",
    "FL-600": "county_govt", "FL-601": "county_govt", "FL-603": "county_govt",
    "GA-502": "county_govt", "GA-503": "county_govt", "GA-508": "county_govt",
    "IL-508": "county_govt", "IL-517": "county_govt", "KY-502": "county_govt",
    "MA-503": "county_govt", "MD-503": "county_govt", "MD-506": "county_govt",
    "MD-513": "county_govt", "MD-600": "county_govt", "MD-601": "county_govt",
    "MN-500": "county_govt", "MN-501": "county_govt", "MN-509": "county_govt",
    "NC-505": "county_govt", "NC-513": "county_govt", "NJ-500": "county_govt",
    "NJ-501": "county_govt", "NJ-502": "county_govt", "NJ-506": "county_govt",
    "NJ-507": "county_govt", "NJ-512": "county_govt", "NJ-513": "county_govt",
    "NV-500": "county_govt", "NV-501": "county_govt", "NY-604": "county_govt",
    "OH-502": "county_govt", "OH-504": "county_govt", "OH-505": "county_govt",
    "OR-500": "county_govt", "OR-507": "county_govt", "PA-510": "county_govt",
    "UT-500": "county_govt", "VA-600": "county_govt", "VA-601": "county_govt",
    "WA-500": "county_govt", "WA-503": "county_govt",
    # states / territorial agencies
    "CO-500": "state_govt", "GU-500": "state_govt", "IN-502": "state_govt",
    "KY-500": "state_govt", "LA-509": "state_govt", "MA-516": "state_govt",
    "MI-500": "state_govt", "ND-500": "state_govt", "NE-500": "state_govt",
    "NY-525": "state_govt", "OH-507": "state_govt", "PA-601": "state_govt",
    "RI-500": "state_govt", "SD-500": "state_govt", "VI-500": "state_govt",
    # regional
    "VA-514": "regional_govt",
    # ambiguous — drop from IV-using models
    "AL-502": "ambiguous",
}


def classify(name: str, coc_id: str = "") -> str:
    # Manual overrides first
    if coc_id in MANUAL_OVERRIDES:
        return MANUAL_OVERRIDES[coc_id]
    if not name:
        return "missing"
    s = name.lower().strip()
    for label, patterns in RULES:
        for pat in patterns:
            if re.search(pat, s):
                return label
    return "other"


def main():
    # Load one (coc_id, 1a_2) record per CoC (from FY2024 if available, else any)
    rows = []
    with (PIPELINE_DIR / "coc_analysis_ready.csv").open() as f:
        rows = list(csv.DictReader(f))

    per_coc: dict[str, str] = {}
    for r in rows:
        if r.get("year") == "2024" and r.get("1a_2"):
            per_coc[r["coc_id"]] = r["1a_2"]
    # Fill from other years where missing
    for r in rows:
        if r["coc_id"] not in per_coc and r.get("1a_2"):
            per_coc[r["coc_id"]] = r["1a_2"]

    results = []
    for coc, name in sorted(per_coc.items()):
        cat = classify(name, coc)
        nonprofit = 1 if cat in ("nonprofit",) else (
            0 if cat in ("city_govt", "county_govt", "state_govt", "housing_authority", "tribal", "regional_govt")
            else None
        )
        results.append({
            "coc_id": coc, "collaborative_applicant": name,
            "lead_agency_type": cat, "nonprofit_led": nonprofit,
        })

    with OUT.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        w.writeheader()
        w.writerows(results)

    counts = Counter(r["lead_agency_type"] for r in results)
    print(f"Classified {len(results)} CoCs:")
    for k, v in counts.most_common():
        print(f"  {k}: {v}")
    unresolved = sum(1 for r in results if r["nonprofit_led"] is None)
    print(f"  (unresolved 'other': {unresolved})")

    # Append to analysis-ready xlsx as a new sheet
    if ANALYSIS_XLSX.exists():
        wb = openpyxl.load_workbook(ANALYSIS_XLSX)
        if "iv_leadership" in wb.sheetnames:
            del wb["iv_leadership"]
        ws = wb.create_sheet("iv_leadership")
        ws.append(list(results[0].keys()))
        for r in results:
            ws.append([r[k] for k in results[0].keys()])
        wb.save(ANALYSIS_XLSX)
        print(f"appended iv_leadership sheet to {ANALYSIS_XLSX.name}")

    print(f"\nSamples per category:")
    from collections import defaultdict
    by_cat = defaultdict(list)
    for r in results:
        by_cat[r["lead_agency_type"]].append(r["collaborative_applicant"])
    for cat in ("nonprofit", "city_govt", "county_govt", "state_govt", "housing_authority", "tribal", "other"):
        print(f"\n{cat}:")
        for name in by_cat[cat][:5]:
            print(f"  {name}")


if __name__ == "__main__":
    main()
