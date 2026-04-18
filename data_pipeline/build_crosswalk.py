"""Build crosswalk.csv mapping FY2022/FY2023/FY2024 question IDs to the
canonical FY2024 schema used by coc_apps_all_info.xlsx.

Strategy
--------
1. Pick a single representative CoC that has all three years (AL-500).
2. Extract (qid, title) anchors from each year's PDF.
3. For each canonical FY2024 variable group (derived from the schema column
   prefixes, not individual subfields), align the 2022/2023 anchors by
   (a) exact qid match, (b) fuzzy title match above a threshold.
4. Write crosswalk.csv; emit a review file listing ambiguous mappings.

Output columns:
    canonical_prefix, title_2024, qid_2024, qid_2023, title_2023,
    qid_2022, title_2022, match_confidence, needs_review, notes

"Canonical prefix" is a shortened field_id family, e.g. "1b_1" covers the
entire 1B-1 participation chart because in the manual xlsx every row of the
chart becomes its own column suffix but maps back to the same source
question in all three years.
"""
from __future__ import annotations

import csv
import difflib
import re
from pathlib import Path

import openpyxl

from pipeline_utils import (
    DATA_DIR,
    PIPELINE_DIR,
    extract_question_anchors,
    pdftotext_layout,
)

REPRESENTATIVE_COC = "AL-500"
SOURCE_FILES = {
    "2022": DATA_DIR / "AL_500_2022.pdf",
    "2023": DATA_DIR / "AL_500_2023.pdf",
    "2024": DATA_DIR / "AL_500_2024.pdf",
}
OUT_CSV = PIPELINE_DIR / "crosswalk.csv"
OUT_REVIEW = PIPELINE_DIR / "crosswalk_review.md"
XLSX = DATA_DIR / "coc_apps_all_info.xlsx"

FIELD_ID_RE = re.compile(r"^(?P<sec>\d)(?P<letter>[a-e])_(?P<rest>.+)$")


def canonical_prefix(field_id: str) -> str | None:
    """'1b_1_1_meetings' -> '1b_1'; '1c_4c_1_mou' -> '1c_4c'; '1a_1a' -> '1a_1a'."""
    m = FIELD_ID_RE.match(field_id.lower())
    if not m:
        return None
    rest = m.group("rest")
    # For 1A-family (1a_1a, 1a_2, 1a_3, 1a_4) keep the full id; these are scalars.
    if m.group("letter") == "a":
        return field_id.lower()
    # For everything else, take the first numeric chunk (and any trailing letter).
    parts = rest.split("_")
    head = parts[0]
    # Detect qids like "4c" (number + letter, e.g., 1c_4c_1_mou).
    if len(parts) >= 2 and re.match(r"^[a-z]$", parts[1]) and head.isdigit():
        head = f"{head}{parts[1]}"
    return f"{m.group('sec')}{m.group('letter')}_{head}"


def xlsx_canonical_families() -> list[tuple[str, str, str]]:
    """Return ordered (canonical_prefix, example_field_id, representative_label)."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2024"]
    seen: dict[str, tuple[str, str]] = {}
    for c in range(1, ws.max_column + 1):
        fid = ws.cell(row=1, column=c).value
        label = ws.cell(row=2, column=c).value or ws.cell(row=4, column=c).value or ""
        if not fid:
            continue
        pref = canonical_prefix(str(fid))
        if pref and pref not in seen:
            seen[pref] = (str(fid), str(label))
    return [(p, ex, lab) for p, (ex, lab) in seen.items()]


def prefix_to_qid(prefix: str) -> str:
    """Heuristic: '1b_1' -> '1B-1'; '1c_4c' -> '1C-4c'; '1a_1a' -> '1A-1'."""
    m = re.match(r"^(\d)([a-e])_(.+)$", prefix)
    if not m:
        return prefix
    section_letter = m.group(2).upper()
    rest = m.group(3)
    # Trim trailing 'a' sub-letter style like '1a_1a'
    if rest.endswith("a") and len(rest) == 2 and rest[0].isdigit():
        rest = rest[0]
    return f"{m.group(1)}{section_letter}-{rest}"


def fuzzy_best(title: str, candidates: dict[str, str]) -> tuple[str | None, float]:
    """Return (best_qid, ratio) of the highest-scoring candidate title."""
    if not title:
        return None, 0.0
    best_qid, best_ratio = None, 0.0
    target = title.lower()
    for qid, t in candidates.items():
        r = difflib.SequenceMatcher(None, target, t.lower()).ratio()
        if r > best_ratio:
            best_qid, best_ratio = qid, r
    return best_qid, best_ratio


def main():
    # Extract anchors per year
    anchors = {}
    for y, path in SOURCE_FILES.items():
        text = pdftotext_layout(path)
        anchors[y] = extract_question_anchors(text)
        print(f"{y}: {len(anchors[y])} anchors")

    families = xlsx_canonical_families()
    print(f"canonical families: {len(families)}")

    rows = []
    review = []

    for prefix, example_field, representative_label in families:
        likely_qid = prefix_to_qid(prefix)
        # Pull 2024 title from anchor if present
        title_2024 = anchors["2024"].get(likely_qid, "")
        if not title_2024:
            # Fall back to label row in xlsx
            title_2024 = representative_label

        # For 2023 & 2022 prefer same qid; fuzzy fallback
        def resolve(year: str):
            if likely_qid in anchors[year]:
                return likely_qid, anchors[year][likely_qid], 1.0, False
            best, ratio = fuzzy_best(title_2024, anchors[year])
            if best and ratio >= 0.60:
                needs_review = ratio < 0.85
                return best, anchors[year][best], round(ratio, 3), needs_review
            return "", "", 0.0, True

        qid_2023, t_2023, c_2023, nr_2023 = resolve("2023")
        qid_2022, t_2022, c_2022, nr_2022 = resolve("2022")

        match_confidence = round(min(c_2023 or 1.0, c_2022 or 1.0), 3)
        needs_review = nr_2023 or nr_2022 or not qid_2023 or not qid_2022
        note_parts = []
        if not qid_2023:
            note_parts.append("no_2023_match")
        if not qid_2022:
            note_parts.append("no_2022_match")
        if nr_2023 or nr_2022:
            note_parts.append("low_similarity_fuzzy")
        notes = ";".join(note_parts)

        rows.append(
            {
                "canonical_prefix": prefix,
                "example_field_id": example_field,
                "title_2024": title_2024,
                "qid_2024": likely_qid if likely_qid in anchors["2024"] else "",
                "qid_2023": qid_2023,
                "title_2023": t_2023,
                "qid_2022": qid_2022,
                "title_2022": t_2022,
                "match_confidence": match_confidence,
                "needs_review": str(needs_review).upper(),
                "notes": notes,
            }
        )
        if needs_review:
            review.append(rows[-1])

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"wrote crosswalk -> {OUT_CSV} ({len(rows)} rows)")

    # Review markdown
    lines = ["# Crosswalk — Items flagged for PI review", ""]
    lines.append(f"Generated from representative CoC `{REPRESENTATIVE_COC}`.")
    lines.append("")
    lines.append(f"**{len(review)}** of **{len(rows)}** canonical families need review.")
    lines.append("")
    lines.append("| prefix | 2024 title | 2023 match | 2022 match | confidence | notes |")
    lines.append("|---|---|---|---|---|---|")
    for r in review:
        lines.append(
            f"| `{r['canonical_prefix']}` | {r['title_2024'][:60]} | "
            f"`{r['qid_2023']}` {r['title_2023'][:50]} | "
            f"`{r['qid_2022']}` {r['title_2022'][:50]} | "
            f"{r['match_confidence']} | {r['notes']} |"
        )
    OUT_REVIEW.write_text("\n".join(lines))
    print(f"wrote review -> {OUT_REVIEW} ({len(review)} flagged)")


if __name__ == "__main__":
    main()
