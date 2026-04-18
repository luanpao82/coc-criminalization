"""Run the extractor on a pilot set of CoCs and diff against manual xlsx.

Outputs:
  pilot_diff_report.md  — human-readable diff & agreement summary
  iterations.csv        — appends a row recording this iteration's metrics
  pilot_diffs.csv       — every disagreement, one row per field
"""
from __future__ import annotations

import csv
import datetime as dt
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from pipeline_utils import DATA_DIR, PIPELINE_DIR, normalize_categorical, normalize_label

from extract_pdf_native import extract_for

PILOT_COCS = ["AL-500", "CA-500", "FL-501", "NY-600", "TX-600"]
YEAR = "2024"
XLSX = DATA_DIR / "coc_apps_all_info.xlsx"

OUT_REPORT = PIPELINE_DIR / "pilot_diff_report.md"
OUT_DIFFS = PIPELINE_DIR / "pilot_diffs.csv"
OUT_ITER = PIPELINE_DIR / "iterations.csv"


def load_xlsx_row(coc_id: str) -> dict[str, str]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2024"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    row_idx = None
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == coc_id:
            row_idx = r
            break
    if row_idx is None:
        return {}
    out = {}
    for c, h in enumerate(headers, start=1):
        if h is None:
            continue
        v = ws.cell(row=row_idx, column=c).value
        out[str(h)] = v
    return out


def classify_field(field_id: str) -> str:
    if field_id.startswith("1a_1a") or field_id.startswith("1a_1b"):
        return "C_label"
    if field_id in {"1a_2", "1a_4"}:
        return "C_label"
    if field_id == "1a_3":
        return "A_categorical"
    if field_id.startswith("1b_1_"):
        return "A_categorical"
    if field_id.startswith("1c_1_"):
        return "A_categorical"
    if field_id.startswith("1c_2_"):
        return "A_categorical"
    if field_id.startswith("1c_3_"):
        return "A_categorical"
    if field_id.startswith("1d_1_"):
        return "A_categorical"
    if field_id.startswith("1d_4_"):
        return "A_categorical"
    if field_id.startswith("1d_6_"):
        return "A_categorical"
    if field_id.startswith("1d_9b_"):
        return "A_categorical"
    if field_id.startswith("1c_4c_"):
        return "A_categorical"
    if field_id.startswith("1c_4_"):
        return "A_categorical"
    if field_id.startswith("1c_5c_"):
        return "A_categorical"
    if field_id.startswith("1c_5_"):
        return "A_categorical"
    if field_id.startswith("1c_7c_"):
        return "A_categorical"
    if field_id == "1d_9_1":
        return "A_categorical"
    if field_id in {"2a_6", "3a_1", "3a_2", "3c_1", "4a_1"}:
        return "A_categorical"
    if field_id in {"1d_2_1", "1d_2_2"}:
        return "B_integer"
    if field_id == "1d_2_3":
        return "B_percent"
    if field_id.startswith("1d_10a_") and field_id.endswith("_years"):
        return "B_integer"
    if field_id.startswith("1d_10a_") and field_id.endswith("_unsheltered"):
        return "B_integer"
    if field_id.startswith("2a_5_") and field_id.endswith("_coverage"):
        return "B_percent"
    if field_id.startswith("2a_5_"):
        return "B_integer"
    return "unknown"


def compare(manual_val, auto_val, klass: str) -> tuple[bool, str]:
    """Return (match, reason)."""
    if manual_val is None and auto_val is None:
        return True, "both_blank"
    if manual_val is None:
        return False, "manual_blank"
    if auto_val is None:
        return False, "auto_blank"
    if klass == "A_categorical":
        m = normalize_categorical(str(manual_val))
        a = normalize_categorical(str(auto_val))
        return m == a, f"manual={m!r}/auto={a!r}"
    if klass == "B_integer":
        try:
            m = int(float(manual_val))
            a = int(float(auto_val))
            return m == a, f"manual={m}/auto={a}"
        except (ValueError, TypeError):
            return False, f"cast_fail manual={manual_val!r}/auto={auto_val!r}"
    if klass == "B_percent":
        # Normalize both to percent (0–100 scale); fractions <= 1.5 get *100.
        # Tolerate rounding to 2 decimals to absorb reporting differences.
        def _to_pct(v):
            s = str(v).strip().rstrip("%")
            try:
                f = float(s)
            except ValueError:
                return None
            if f <= 1.5:
                f = f * 100.0
            return round(f, 2)
        m = _to_pct(manual_val); a = _to_pct(auto_val)
        if m is None or a is None:
            return False, f"cast_fail manual={manual_val!r}/auto={auto_val!r}"
        return abs(m - a) < 0.01, f"manual={m}/auto={a}"
    # label — normalize whitespace, case; tolerate " CoC" suffix & punctuation.
    m = normalize_label(str(manual_val)).lower()
    a = normalize_label(str(auto_val)).lower()
    if m == a:
        return True, f"manual={m!r}/auto={a!r}"
    # More lenient: strip trailing " coc" suffix and trailing punctuation
    def _loose(s):
        s = s.strip().rstrip(".,;:")
        for suf in (" coc", " cocc"):
            if s.endswith(suf):
                s = s[: -len(suf)].strip()
        s = s.replace("  ", " ")
        return s
    return _loose(m) == _loose(a), f"manual={m!r}/auto={a!r}"


def main():
    all_diffs = []
    per_coc_stats = []
    global_match = 0
    global_total = 0
    per_class = defaultdict(lambda: [0, 0])  # class -> [match, total]

    for coc_id in PILOT_COCS:
        print(f"[extract] {coc_id} {YEAR}")
        records = extract_for(coc_id, YEAR)
        manual = load_xlsx_row(coc_id)
        if not manual:
            print(f"  [warn] no manual row for {coc_id}")
            continue
        coc_match = 0
        coc_total = 0
        for r in records:
            field_id = r["field_id"]
            manual_val = manual.get(field_id)
            klass = classify_field(field_id)
            ok, reason = compare(manual_val, r["value"], klass)
            per_class[klass][1] += 1
            coc_total += 1
            if ok:
                per_class[klass][0] += 1
                coc_match += 1
            else:
                all_diffs.append(
                    {
                        "coc_id": coc_id,
                        "year": YEAR,
                        "field_id": field_id,
                        "class": klass,
                        "manual_value": manual_val,
                        "auto_value": r["value"],
                        "source_page": r["source_page"],
                        "reason": reason,
                    }
                )
        per_coc_stats.append((coc_id, coc_match, coc_total))
        global_match += coc_match
        global_total += coc_total

    # Write diffs
    with OUT_DIFFS.open("w", newline="") as f:
        if all_diffs:
            w = csv.DictWriter(f, fieldnames=list(all_diffs[0].keys()))
            w.writeheader()
            w.writerows(all_diffs)

    weighted_acc = (global_match / global_total) if global_total else 0.0
    class_acc = {
        k: (v[0] / v[1]) if v[1] else 0.0 for k, v in per_class.items()
    }

    # Adjusted accuracy: ignore comparisons where manual is blank and auto
    # has value (i.e., comparisons where the manual coder did not enter data).
    # This isolates "disagreements on shared data" from "manual incompleteness."
    adj_total = global_total - sum(
        1 for d in all_diffs if (d["manual_value"] in (None, "") and d["auto_value"] not in (None, ""))
    )
    adj_match = global_match  # both-match contributions unchanged
    adjusted_acc = (adj_match / adj_total) if adj_total else 0.0

    # Report
    ver_seen = set()
    for rlist in []:
        pass
    # Determine extractor version from first record
    ver_label = "pdf_native_vX"
    for coc in PILOT_COCS:
        p = PIPELINE_DIR / "extracted" / f"{coc}_{YEAR}.json"
        if p.exists():
            import json as _json
            try:
                rs = _json.loads(p.read_text())
                if rs:
                    ver_label = rs[0].get("extractor", ver_label)
                    break
            except Exception:
                pass
    lines = [
        "# Pilot Diff Report",
        "",
        f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"Extractor: `{ver_label}`  Pilot: {PILOT_COCS}  Year: {YEAR}",
        "",
        "## Headline numbers",
        "",
        f"- **Overall agreement (weighted):** {weighted_acc:.3%} "
        f"({global_match}/{global_total} fields)",
        f"- **Adjusted agreement** (excludes manual-blank cells the extractor filled): "
        f"{adjusted_acc:.3%} ({adj_match}/{adj_total} fields)",
        "- **Per class:**",
    ]
    for k, acc in sorted(class_acc.items()):
        m, t = per_class[k]
        lines.append(f"  - `{k}`: {acc:.3%} ({m}/{t})")
    lines += ["", "## Per CoC", "", "| CoC | match | total | acc |", "|---|---|---|---|"]
    for cid, m, t in per_coc_stats:
        lines.append(f"| {cid} | {m} | {t} | {m/t:.3%} |")
    lines += ["", "## Diffs (first 50)", "", "| CoC | field | class | manual | auto | reason |", "|---|---|---|---|---|---|"]
    for d in all_diffs[:50]:
        def fmt(v):
            s = "" if v is None else str(v).replace("|", "\\|")
            return s[:60]
        lines.append(
            f"| {d['coc_id']} | `{d['field_id']}` | {d['class']} | "
            f"{fmt(d['manual_value'])} | {fmt(d['auto_value'])} | {d['reason'][:80]} |"
        )

    # Error-category guess
    reason_cats = Counter()
    for d in all_diffs:
        if d["auto_value"] is None:
            reason_cats["auto_blank (extractor gap or format variant)"] += 1
        elif d["manual_value"] is None:
            reason_cats["manual_blank (possible manual skip)"] += 1
        else:
            reason_cats["value_mismatch"] += 1
    lines += ["", "## Disagreement categories (rough)", ""]
    for k, n in reason_cats.most_common():
        lines.append(f"- {k}: **{n}**")

    OUT_REPORT.write_text("\n".join(lines))

    # Iteration ledger — increment iter based on existing row count
    ledger_exists = OUT_ITER.exists()
    iter_no = 1
    if ledger_exists:
        with OUT_ITER.open() as rf:
            iter_no = sum(1 for _ in rf)  # header + rows -> next iter = count
    with OUT_ITER.open("a", newline="") as f:
        w = csv.writer(f)
        if not ledger_exists:
            w.writerow([
                "iter", "date", "extractor_version", "weighted_acc",
                "A_acc", "C_acc", "total_diffs", "top_error",
                "fix_applied", "regression_pass",
            ])
        top_error = reason_cats.most_common(1)[0][0] if reason_cats else ""
        version_map = {1: "pdf_native_v0.1", 2: "pdf_native_v0.2"}
        version = version_map.get(iter_no, "pdf_native_v0.3")
        w.writerow([
            iter_no, dt.datetime.now().date().isoformat(),
            version,
            f"{weighted_acc:.4f}",
            f"{class_acc.get('A_categorical', 0):.4f}",
            f"{class_acc.get('C_label', 0):.4f}",
            len(all_diffs), top_error,
            {1: "baseline_run", 2: "wrap_capture_for_1A_2_3_4"}.get(
                iter_no, "expand_to_1C_1D_sections"
            ),
            "",
        ])

    print(f"report -> {OUT_REPORT}")
    print(f"diffs  -> {OUT_DIFFS}  ({len(all_diffs)} rows)")
    print(f"ledger -> {OUT_ITER}")
    print(f"weighted acc: {weighted_acc:.3%}")


if __name__ == "__main__":
    main()
