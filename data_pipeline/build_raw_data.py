"""Build the comprehensive raw data file — 677 rows × full 331-variable schema.

Design goals (per PI direction 2026-04-18):
  1. Faithful to the original HUD CoC Consolidated Application form — every
     question in the PDF appears as a variable, even when not yet extracted.
  2. Populated where data exists (Class A/B/C structured from the pipeline,
     plus the 3 research-used PLE narratives).
  3. Empty columns present as scaffolding so future extraction passes can
     fill them without schema changes, and researchers can see at a glance
     which fields still need work.

Column order mirrors the manual reference (`coc_apps_all_info.xlsx`) which
follows the PDF's question order (1A → 1B → 1C → 1D → 1E → 2A → 2B → 2C →
3A → 3B → 4A). Cross-year crosswalk of PLE narratives uses FY24 canonical
names (1d_10 / 1d_10b / 1d_10c) — FY22/23 narrative content is populated
into the same columns since they represent the same construct.

Output: coc_raw_data.xlsx
  - Sheet `raw_data`: 677 rows × (meta + 331 canonical + completeness) cols
  - Sheet `schema`: canonical variable list with extraction status
"""
from __future__ import annotations
import sys, os
from pathlib import Path
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from pipeline_utils import DATA_DIR, pdftotext_layout
from extract_narrative import NARRATIVE_SPECS, resolve_anchors, slice_narrative

HERE = Path(__file__).parent
MANUAL_REF = Path(DATA_DIR) / "coc_apps_all_info.xlsx"
OUT_PATH = HERE / "coc_raw_data.xlsx"

META_COLS = [
    "original_filename", "coc_id", "year", "format", "num_pages",
    "text_chars", "is_scanned", "name_issue", "notes",
]
COMPLETENESS_COLS = [
    "n_fields_populated", "pct_fields_populated",
    "ple_umbrella_status_code", "ple_umbrella_status",
    "ple_prof_dev_status_code", "ple_prof_dev_status",
    "ple_feedback_status_code", "ple_feedback_status",
]

# PLE narrative fields: research-used, text populated
PLE_TEXT_FIELDS = {
    "ple_umbrella":  "1d_10",   # FY24 anchor; FY22/23 1D-11 maps here
    "ple_prof_dev":  "1d_10b",
    "ple_feedback":  "1d_10c",
}


def load_canonical_schema() -> list[str]:
    """The 331 canonical variable names (column order from manual ref)."""
    ref = pd.read_excel(MANUAL_REF, sheet_name="2024", nrows=0)
    return list(ref.columns)


def pivot_panel_long() -> pd.DataFrame:
    """Return wide dataframe of already-extracted structured values."""
    long = pd.read_csv(HERE / "coc_panel_long.csv")
    # Keep latest value per (coc_id, year, field_id)
    wide = long.pivot_table(
        index=["coc_id", "year"], columns="field_id",
        values="value", aggfunc="first",
    ).reset_index()
    wide.columns.name = None
    return wide


def slice_ple_narratives(inventory: pd.DataFrame) -> dict[tuple, dict]:
    """Return {(coc_id, year): {umbrella: text, prof_dev: text, feedback: text,
    status codes, pages}} using the configured anchor patterns."""
    out: dict[tuple, dict] = {}
    text_cache: dict[str, str] = {}

    def pdf_text(fname: str) -> str:
        if fname in text_cache:
            return text_cache[fname]
        path = Path(DATA_DIR) / fname
        try:
            text_cache[fname] = pdftotext_layout(path) if path.exists() else ""
        except Exception:
            text_cache[fname] = ""
        return text_cache[fname]

    for i, row in inventory.iterrows():
        coc_id, fname = row["coc_id"], row["original_filename"]
        try:
            year = str(int(row["year"]))
        except Exception:
            continue
        if year not in ("2022", "2023", "2024"):
            continue
        key = (coc_id, int(year))
        is_pdf = str(row.get("format", "")).lower() == "pdf"
        is_scanned = bool(row.get("is_scanned"))
        rec: dict = {}
        for ple_fid in PLE_TEXT_FIELDS:
            def _set(code, detail=None, text="", page=None):
                rec[f"{ple_fid}_status_code"] = code
                rec[f"{ple_fid}_status"] = detail or code
                rec[f"{ple_fid}_text"] = text
                rec[f"{ple_fid}_page"] = page

            if not is_pdf:
                _set("NOT_PDF", f"NOT_PDF (format={row.get('format')})")
                continue
            path = Path(DATA_DIR) / fname
            if not path.exists():
                _set("FILE_NOT_FOUND", "source file not found")
                continue
            if is_scanned:
                _set("SCANNED", "scanned PDF — pdftotext yields no usable layer (needs OCR)")
                continue
            spec = NARRATIVE_SPECS[ple_fid]
            try:
                anchor, next_anchor = resolve_anchors(spec, year)
                text, page = slice_narrative(path, anchor, next_anchor)
            except Exception as e:
                _set("ERROR", f"ERROR: {e}")
                continue
            if len(text) >= 30:
                _set("ok", "ok", text=text[:32000], page=page)
            else:
                body = pdf_text(fname)
                if "SF-424" in body and "Before Starting the Project Application" in body:
                    _set("MIS_FILED_PROJECT_APP",
                         "source is CoC Program Project Application (SF-424), not Consolidated App")
                elif "Special NOFO" in body or "Special Notice of Funding Opportunity" in body:
                    _set("SPECIAL_NOFO_INSTRUMENT",
                         "Special NOFO CoC Application (unsheltered set-aside) — different HUD instrument")
                else:
                    _set("ANCHOR_NOT_FOUND",
                         "PLE narrative anchor not found in this PDF")
        out[key] = rec
        if (i + 1) % 120 == 0:
            print(f"  ...sliced {i+1}/{len(inventory)}")
    return out


def main():
    print("=" * 70)
    print("[build_raw_data] Building comprehensive raw file")
    print("=" * 70)

    canonical = load_canonical_schema()
    print(f"Canonical schema: {len(canonical)} variables")

    inv = pd.read_csv(HERE / "file_inventory.csv")
    print(f"File inventory: {len(inv)} rows (one per source PDF/DOCX)")

    structured = pivot_panel_long()
    print(f"Structured panel (pivoted from long): "
          f"{len(structured)} CoC-years × {len(structured.columns)-2} fields")

    print("\nSlicing 3 PLE narrative fields for all PDFs ...")
    ple_data = slice_ple_narratives(inv)
    print(f"  sliced narratives for {len(ple_data)} CoC-years")

    # ---- Assemble the raw table ----
    # Start with inventory (677 rows), left-join structured, then fill PLE text
    out = inv.copy()
    out = out.merge(structured, on=["coc_id", "year"], how="left")

    # Ensure every canonical column exists (empty if not)
    for c in canonical:
        if c not in out.columns:
            out[c] = pd.NA

    # Inject PLE narrative text into canonical columns 1d_10 / 1d_10b / 1d_10c
    for ple_fid, canonical_col in PLE_TEXT_FIELDS.items():
        out[f"{ple_fid}_status"] = ""
        out[f"{ple_fid}_status_code"] = ""
        out[f"{ple_fid}_page"] = pd.NA
    for i, row in out.iterrows():
        key = (row["coc_id"], row["year"])
        if key not in ple_data:
            continue
        rec = ple_data[key]
        for ple_fid, canonical_col in PLE_TEXT_FIELDS.items():
            txt = rec.get(f"{ple_fid}_text", "")
            if txt:
                out.at[i, canonical_col] = txt
            out.at[i, f"{ple_fid}_page"] = rec.get(f"{ple_fid}_page")
            out.at[i, f"{ple_fid}_status"] = rec.get(f"{ple_fid}_status", "")
            out.at[i, f"{ple_fid}_status_code"] = rec.get(f"{ple_fid}_status_code", "")

    # ---- Row-level completeness ----
    def row_completeness(row):
        filled = sum(1 for c in canonical if pd.notna(row.get(c)) and str(row[c]).strip() != "")
        return pd.Series({
            "n_fields_populated": filled,
            "pct_fields_populated": round(100 * filled / len(canonical), 1),
        })
    print("\nComputing row-level completeness ...")
    comp = out.apply(row_completeness, axis=1)
    out = pd.concat([out, comp], axis=1)

    # ---- Final column order ----
    final_cols = (
        META_COLS
        + [c for c in COMPLETENESS_COLS if c in out.columns]
        + canonical
        + [f"{f}_page" for f in PLE_TEXT_FIELDS]
    )
    # Drop duplicates, preserve order
    seen = set(); final_cols = [c for c in final_cols if not (c in seen or seen.add(c))]
    out = out[final_cols]

    # ---- Schema inventory sheet ----
    schema_rows = []
    for c in canonical:
        non_null = out[c].notna().sum() if c in out.columns else 0
        pct = round(100 * non_null / len(out), 1)
        # Classify
        if c in PLE_TEXT_FIELDS.values():
            cls = "narrative (PLE, research-used)"
        elif c in ("1b_1a", "1b_3", "1c_4a", "1c_4b", "1c_5a", "1c_5b", "1c_5d", "1c_5e",
                   "1c_5f", "1c_6a", "1c_7a", "1d_11", "1d_2a", "1d_3", "1d_6a",
                   "1d_7", "1d_7a", "1d_8", "1d_8a", "1d_8b", "1d_9a", "1d_9b", "1d_9c", "1d_9d"):
            cls = "narrative (not yet extracted)"
        else:
            cls = "structured"
        schema_rows.append({
            "variable": c,
            "class": cls,
            "n_populated": non_null,
            "pct_populated": pct,
            "notes": "3-year stable" if non_null >= 600 else ("sparse" if non_null < 100 else ""),
        })
    schema_df = pd.DataFrame(schema_rows)

    # ---- Write xlsx with formatting ----
    from openpyxl.styles import Alignment, PatternFill, Font
    print(f"\nWriting {OUT_PATH} ...")
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name="raw_data", index=False)
        schema_df.to_excel(xw, sheet_name="schema", index=False)

        ws = xw.sheets["raw_data"]
        narr_cols = set(PLE_TEXT_FIELDS.values())
        # column widths
        for idx, col in enumerate(out.columns, start=1):
            letter = ws.cell(row=1, column=idx).column_letter
            if col in narr_cols:
                ws.column_dimensions[letter].width = 60
            elif col.endswith("_status"):
                ws.column_dimensions[letter].width = 24
            elif col == "original_filename":
                ws.column_dimensions[letter].width = 22
            elif col == "notes":
                ws.column_dimensions[letter].width = 22
            else:
                ws.column_dimensions[letter].width = 12

        red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        green = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        grey = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        bold = Font(bold=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col_name = ws.cell(row=1, column=cell.column).value
                if col_name in narr_cols and cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_name and col_name.endswith("_status") and cell.value:
                    if cell.value == "ok":
                        cell.fill = green
                    elif cell.value:
                        cell.fill = red
                        cell.font = bold
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "J2"

        # Schema sheet
        ws2 = xw.sheets["schema"]
        for idx, col in enumerate(schema_df.columns, start=1):
            letter = ws2.cell(row=1, column=idx).column_letter
            ws2.column_dimensions[letter].width = 30 if col == "class" else 18
        for cell in ws2[1]:
            cell.font = bold
        ws2.freeze_panes = "A2"
        for row in ws2.iter_rows(min_row=2):
            cls = row[1].value
            if cls and "PLE" in cls:
                for c in row:
                    c.fill = green
            elif cls and "not yet extracted" in cls:
                for c in row:
                    c.fill = red

    # ---- Write stage1_flagged.csv for the coauthor-facing site ----
    # Any row where the PLE umbrella status is not "ok" is something a
    # coauthor can act on. The status codes from classify_missing carry the
    # diagnosis; this export adds a plain-language "action_needed" column.
    ACTIONS = {
        "MIS_FILED_PROJECT_APP":
            "Replace: find correct Consolidated Application on HUD e-snaps and swap in OneDrive",
        "SCANNED":
            "OCR: run ocrmypdf on the PDF, or hand-enter key fields from the source",
        "SPECIAL_NOFO_INSTRUMENT":
            "Exclude: FY2022 unsheltered set-aside — different HUD form, no 1D PLE section; excluded by design",
        "ANCHOR_NOT_FOUND":
            "Review: PDF has text but non-standard template; manually confirm or re-source",
        "PLE_BLOCK_ABSENT":
            "Review: 1D-* anchors present but PLE block missing; likely incomplete submission",
        "LOW_TEXT":
            "Review: PDF text layer nearly empty; may be scanned or corrupt",
        "UNKNOWN_FORMAT":
            "Review: document structure unrecognized; check whether this is the correct file",
        "NOT_PDF":
            "Use PDF sibling: this is a DOCX duplicate — the parallel PDF is the authoritative source",
        "FILE_NOT_FOUND":
            "Re-source: file listed in inventory but not present in OneDrive",
        "OUT_OF_SCOPE":
            "Ignore: year outside the FY2022–2024 analysis panel",
    }
    flagged = out[out["ple_umbrella_status_code"] != "ok"].copy()
    flagged["action_needed"] = flagged["ple_umbrella_status_code"].map(ACTIONS).fillna("Review")
    flagged_export = flagged[[
        "original_filename", "coc_id", "year", "format", "num_pages", "is_scanned",
        "ple_umbrella_status_code", "ple_umbrella_status", "action_needed",
        "pct_fields_populated", "n_fields_populated",
    ]].rename(columns={"ple_umbrella_status_code": "status_code",
                       "ple_umbrella_status": "status_detail"})
    flagged_export.sort_values(["status_code", "year", "coc_id"]).to_csv(
        HERE / "stage1_flagged.csv", index=False)
    print(f"\n[stage1_flagged.csv] wrote {len(flagged_export)} flagged rows")

    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"\n[done] {OUT_PATH.name}  ({size_kb:.0f} KB)")
    print(f"  raw_data: {len(out)} rows × {len(out.columns)} cols")
    print(f"  schema: {len(schema_df)} variable entries")
    # Summary stats
    print(f"\nCompleteness distribution:")
    print(out["pct_fields_populated"].describe().round(1).to_string())
    print(f"\nPLE narrative status:")
    for ple_fid in PLE_TEXT_FIELDS:
        print(f"  {ple_fid}:")
        print(out[f"{ple_fid}_status"].value_counts().head(6).to_string().replace("\n", "\n    "))


if __name__ == "__main__":
    main()
