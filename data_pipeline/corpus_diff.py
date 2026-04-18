"""Corpus-wide diff: compare every extracted CoC to the manual xlsx.

Produces:
  corpus_agreement.md — summary by section/class/coc
  corpus_diffs.csv    — every disagreement
"""
from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from pipeline_utils import DATA_DIR, EXTRACTED_DIR, PIPELINE_DIR
from pilot_run import classify_field, compare

XLSX = DATA_DIR / "coc_apps_all_info.xlsx"
YEAR = "2024"
OUT_DIFFS = PIPELINE_DIR / "corpus_diffs.csv"
OUT_REPORT = PIPELINE_DIR / "corpus_agreement.md"


def load_manual():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["2024"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    data: dict[str, dict] = {}
    for r in range(5, ws.max_row + 1):
        coc = ws.cell(row=r, column=1).value
        if not coc:
            continue
        row = {}
        for c, h in enumerate(headers, start=1):
            if h is None:
                continue
            row[str(h)] = ws.cell(row=r, column=c).value
        data[coc] = row
    return data


def main():
    manual = load_manual()
    diffs = []
    global_match = 0
    global_total = 0
    per_class = defaultdict(lambda: [0, 0])
    per_coc = defaultdict(lambda: [0, 0])
    per_field = defaultdict(lambda: [0, 0])
    auto_fills = 0  # manual blank + auto value

    files = sorted(EXTRACTED_DIR.glob(f"*_{YEAR}.json"))
    print(f"Comparing {len(files)} extracted CoCs against manual xlsx...")

    for f in files:
        coc_id = f.stem.rsplit("_", 1)[0]
        if coc_id not in manual:
            continue
        records = json.loads(f.read_text())
        mrow = manual[coc_id]
        for r in records:
            fid = r["field_id"]
            auto = r["value"]
            man = mrow.get(fid)
            klass = classify_field(fid)
            if klass == "unknown":
                continue
            ok, reason = compare(man, auto, klass)
            per_class[klass][1] += 1
            per_coc[coc_id][1] += 1
            per_field[fid][1] += 1
            global_total += 1
            if ok:
                per_class[klass][0] += 1
                per_coc[coc_id][0] += 1
                per_field[fid][0] += 1
                global_match += 1
            else:
                if (man in (None, "")) and (auto not in (None, "")):
                    auto_fills += 1
                diffs.append({
                    "coc_id": coc_id, "year": YEAR, "field_id": fid,
                    "class": klass,
                    "manual_value": man,
                    "auto_value": auto,
                    "source_page": r.get("source_page"),
                    "reason": reason,
                })

    weighted = (global_match / global_total) if global_total else 0.0
    adj_total = global_total - auto_fills
    adj_match = global_match
    adjusted = (adj_match / adj_total) if adj_total else 0.0

    with OUT_DIFFS.open("w", newline="") as f:
        if diffs:
            w = csv.DictWriter(f, fieldnames=list(diffs[0].keys()))
            w.writeheader()
            w.writerows(diffs)

    lines = [
        "# Corpus-Wide Agreement (FY2024, all extractable CoCs)",
        "",
        f"- CoCs compared: **{len(files)}**",
        f"- Total field comparisons: **{global_total:,}**",
        f"- Matching: **{global_match:,}**",
        f"- Disagreements: **{len(diffs):,}**",
        f"  - of which manual-blank (extractor filled a gap): **{auto_fills:,}**",
        f"  - of which true value disagreements: **{len(diffs) - auto_fills:,}**",
        "",
        f"## Weighted agreement: **{weighted:.3%}**",
        f"## Adjusted agreement (exclude manual-blank fills): **{adjusted:.3%}**",
        "",
        "### By variable class",
        "",
        "| class | match | total | acc |",
        "|---|---|---|---|",
    ]
    for k, v in sorted(per_class.items()):
        lines.append(f"| {k} | {v[0]:,} | {v[1]:,} | {(v[0]/v[1] if v[1] else 0):.3%} |")
    lines += ["", "### Fields with highest disagreement rates (top 20)", ""]
    lines.append("| field | ok | total | acc |")
    lines.append("|---|---|---|---|")
    ranked = sorted(
        per_field.items(),
        key=lambda kv: (kv[1][0] / max(kv[1][1], 1)),
    )
    for fid, v in ranked[:20]:
        lines.append(f"| `{fid}` | {v[0]} | {v[1]} | {(v[0]/v[1] if v[1] else 0):.3%} |")

    lines += ["", "### CoCs with most disagreements (top 10)", "", "| CoC | ok | total | acc |", "|---|---|---|---|"]
    coc_rank = sorted(per_coc.items(), key=lambda kv: kv[1][0] / max(kv[1][1], 1))
    for coc, v in coc_rank[:10]:
        lines.append(f"| {coc} | {v[0]} | {v[1]} | {(v[0]/v[1] if v[1] else 0):.3%} |")

    OUT_REPORT.write_text("\n".join(lines))
    print(f"weighted: {weighted:.3%}  adjusted: {adjusted:.3%}")
    print(f"auto_fills: {auto_fills}  true_mismatches: {len(diffs)-auto_fills}")
    print(f"report -> {OUT_REPORT}")
    print(f"diffs  -> {OUT_DIFFS} ({len(diffs)} rows)")


if __name__ == "__main__":
    main()
