"""Build a single analysis-ready dataset.

Merges:
  - panel_safe fields (117 variables with ≥50% coverage in all 3 years)
  - harmonized DVs (crim_activity_index and 3 auxiliaries)
  - metadata (coc_id, year)

Output:
  coc_analysis_ready.xlsx  — one sheet per use case:
    * balanced_panel : only CoCs present in all 3 years (100 × 3 = 300 rows)
    * unbalanced     : every observed (CoC, year) (612 rows)
    * fy2024_only    : full 2024 cross-section (285 rows)
  coc_analysis_ready.csv   — unbalanced long form (easiest for R/Python)
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from pipeline_utils import DATA_DIR, EXTRACTED_DIR, PIPELINE_DIR

OUT_XLSX = PIPELINE_DIR / "coc_analysis_ready.xlsx"
OUT_CSV = PIPELINE_DIR / "coc_analysis_ready.csv"
HARMONIZED = PIPELINE_DIR / "harmonized_dv.csv"
FIELD_MAP = PIPELINE_DIR / "panel_field_map.csv"
XLSX_MANUAL = DATA_DIR / "coc_apps_all_info.xlsx"


def load_canonical_fields():
    wb = openpyxl.load_workbook(XLSX_MANUAL, data_only=True)
    ws = wb["2024"]
    return [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value]


def load_panel_safe_fields():
    with FIELD_MAP.open() as f:
        rows = list(csv.DictReader(f))
    return [r["field_id"] for r in rows if r["category"] == "panel_safe"]


def load_extracted():
    data = {}
    for path in sorted(EXTRACTED_DIR.glob("*.json")):
        stem = path.stem
        parts = stem.rsplit("_", 1)
        if len(parts) != 2:
            continue
        coc, yr = parts
        try:
            recs = json.loads(path.read_text())
        except Exception:
            continue
        data[(coc, yr)] = {
            r["field_id"]: r["value"]
            for r in recs
            if r.get("value") is not None and not r.get("needs_review")
        }
    return data


def load_harmonized_dv():
    with HARMONIZED.open() as f:
        rows = list(csv.DictReader(f))
    return {(r["coc_id"], r["year"]): r for r in rows}


def main():
    panel_safe = load_panel_safe_fields()
    extracted = load_extracted()
    harm = load_harmonized_dv()

    dv_cols = [
        "cells_observed", "cells_yes",
        "crim_activity_index",
        "engaged_policymakers_crim",
        "engaged_law_enforce_crim",
        "implemented_anticrim_practice",
    ]

    # Build rows: (coc_id, year) → flat dict
    all_rows = []
    for (coc, yr), vals in sorted(extracted.items()):
        row = {"coc_id": coc, "year": yr}
        for f in panel_safe:
            row[f] = vals.get(f, "")
        # Harmonized DVs
        h = harm.get((coc, yr), {})
        for c in dv_cols:
            row[c] = h.get(c, "")
        all_rows.append(row)

    header = ["coc_id", "year"] + dv_cols + panel_safe

    # CSV (long form)
    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        w.writeheader()
        w.writerows(all_rows)
    print(f"wrote {OUT_CSV} — {len(all_rows)} rows × {len(header)} columns")

    # xlsx with multiple sheets
    wb = openpyxl.Workbook()

    # Sheet 1: balanced panel
    years_per_coc = defaultdict(set)
    for r in all_rows:
        years_per_coc[r["coc_id"]].add(r["year"])
    balanced_cocs = {coc for coc, ys in years_per_coc.items()
                     if ys == {"2022", "2023", "2024"}}
    ws = wb.active
    ws.title = "balanced_panel"
    ws.append(header)
    n_bal = 0
    for r in all_rows:
        if r["coc_id"] in balanced_cocs:
            ws.append([r[k] for k in header])
            n_bal += 1
    print(f"balanced_panel sheet: {n_bal} rows ({len(balanced_cocs)} CoCs × 3 years)")

    # Sheet 2: unbalanced (all observations)
    ws2 = wb.create_sheet("unbalanced")
    ws2.append(header)
    for r in all_rows:
        ws2.append([r[k] for k in header])
    print(f"unbalanced sheet: {len(all_rows)} rows")

    # Sheet 3: FY2024 cross-section only
    ws3 = wb.create_sheet("fy2024_only")
    ws3.append(header)
    n24 = 0
    for r in all_rows:
        if r["year"] == "2024":
            ws3.append([r[k] for k in header])
            n24 += 1
    print(f"fy2024_only sheet: {n24} rows")

    # Sheet 4: variable codebook (lightweight)
    ws4 = wb.create_sheet("variables")
    ws4.append(["name", "role", "description"])
    ws4.append(["coc_id", "key", "CoC identifier (e.g., AL-500)"])
    ws4.append(["year", "key", "Fiscal year (2022, 2023, 2024)"])
    # Harmonized DVs
    dv_descriptions = {
        "cells_observed": ("diag", "# 1d_4 cells with observed Yes/No value"),
        "cells_yes": ("diag", "# 1d_4 cells = Yes"),
        "crim_activity_index": ("DV primary", "Continuous [0,1]; share of 1d_4 cells = Yes"),
        "engaged_policymakers_crim": ("DV (ceiling)", "Binary; any policymaker-engagement Yes — panel ceiling at 97%, low variance"),
        "engaged_law_enforce_crim": ("DV (ceiling)", "Binary; any law-enforcement Yes — panel ceiling at 96%"),
        "implemented_anticrim_practice": ("DV secondary", "Binary; implemented/reverse-existing column Yes — note FY2024 jump due to instrument change"),
    }
    for name, (role, desc) in dv_descriptions.items():
        ws4.append([name, role, desc])
    # Panel-safe fields — abbreviated descriptions
    for f in panel_safe:
        ws4.append([f, "IV/control/MED", f"extracted from HUD 1D-4/etc (see codebook.md)"])

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
