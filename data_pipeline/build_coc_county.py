"""Build a CoC → county partisanship mapping.

Strategy
--------
1. For each CoC, parse its name (`1a_1b`) for county hints.
2. Match parsed county names against the 2020 county election results (by state).
3. Compute CoC-level Biden vote share as the population-weighted average of
   matched counties (using total_votes as weight).
4. For CoCs with no matchable county name (e.g., "Alaska Balance of State"),
   fall back to the state's overall 2020 vote share (computed from counties).

Output: iv_county.csv with columns:
  coc_id, state, matched_counties, biden_share, source
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import pandas as pd

from pipeline_utils import PIPELINE_DIR

COUNTY_CSV = PIPELINE_DIR / "external" / "county_2020_results.csv"
OUT = PIPELINE_DIR / "iv_county.csv"
ANALYSIS_XLSX = PIPELINE_DIR / "coc_analysis_ready.xlsx"

STATE_NAME_TO_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT",
    "District of Columbia": "DC", "Delaware": "DE", "Florida": "FL",
    "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL",
    "Indiana": "IN", "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY",
    "Louisiana": "LA", "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA",
    "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO",
    "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH",
    "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY", "North Carolina": "NC",
    "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
    "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
    "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
    "Wisconsin": "WI", "Wyoming": "WY",
}


def load_county():
    df = pd.read_csv(COUNTY_CSV, dtype={"county_fips": str})
    df["state_abbr"] = df["state_name"].map(STATE_NAME_TO_ABBR)
    df["county_name_norm"] = (df["county_name"]
                              .str.replace(r"\s+County$", "", regex=True, case=False)
                              .str.replace(r"\s+Parish$", "", regex=True, case=False)
                              .str.replace(r"\s+Borough$", "", regex=True, case=False)
                              .str.replace(r"\s+city$", "", regex=True, case=False)
                              .str.strip().str.lower())
    df["biden_share"] = df["per_dem"]
    return df


def state_share(county_df, state_abbr):
    sub = county_df[county_df["state_abbr"] == state_abbr]
    if sub.empty or sub["total_votes"].sum() == 0:
        return None
    return float((sub["votes_dem"].sum()) / sub["total_votes"].sum())


def parse_counties(name: str, state_abbr: str, county_df) -> list[str]:
    if not name:
        return []
    # Common patterns: "Name, State", "Name County", "Name, Other, Third Counties"
    # Strip "CoC" suffix
    name = re.sub(r"\s*Co[Cc]\s*$", "", name).strip()
    # Capture segments like "Foo, Bar, Baz Counties" or "Foo County"
    text = name
    # Remove parenthetical
    text = re.sub(r"\([^)]*\)", "", text)
    # Replace slash with comma
    text = text.replace("/", ",")

    candidates: list[str] = []
    # Pattern A: "... X, Y, Z Counties"
    m = re.search(r"([A-Z][A-Za-z\.\-\' ,]+?)\s+(Counties|County|Parish|Parishes|Borough)\b", text)
    if m:
        block = m.group(1)
        # Split by comma and "and"
        parts = re.split(r",|\band\b", block)
        for p in parts:
            p = p.strip()
            if p and len(p) <= 40:
                candidates.append(p.lower())
    # Pattern B: "XYZ City"
    m2 = re.findall(r"\b([A-Z][A-Za-z\.\-\' ]+?)\s+City\b", text)
    for c in m2:
        c = c.strip().lower()
        if c and c not in candidates:
            candidates.append(c)
    # Clean up
    candidates = [c for c in candidates if not re.match(r"^(the|and|of|balance|city|county|counties|joint|regional)$", c)]
    # Verify against county data
    pool = county_df[county_df["state_abbr"] == state_abbr]["county_name_norm"].tolist()
    verified = []
    for c in candidates:
        if c in pool:
            verified.append(c)
        else:
            # Try partial (startswith)
            match = [p for p in pool if p == c or p.startswith(c + " ")]
            if match:
                verified.append(match[0])
    return list(dict.fromkeys(verified))  # dedupe preserving order


def main():
    county_df = load_county()
    print(f"Loaded {len(county_df)} counties across {county_df['state_abbr'].nunique()} states")

    # Pre-compute state-level shares
    state_shares = {s: state_share(county_df, s)
                    for s in county_df["state_abbr"].unique() if pd.notna(s)}

    # Load CoC names from analysis_ready
    dfx = pd.read_excel(ANALYSIS_XLSX, sheet_name="unbalanced")
    coc_names = (dfx.dropna(subset=["1a_1b"])
                 .groupby("coc_id")["1a_1b"].agg(lambda s: s.iloc[-1])
                 .to_dict())

    rows = []
    matched_count = 0
    fallback_count = 0
    unknown_count = 0
    for coc_id, name in sorted(coc_names.items()):
        state = coc_id.split("-")[0].upper()
        matched = parse_counties(str(name), state, county_df)
        if matched:
            sub = county_df[(county_df["state_abbr"] == state)
                            & (county_df["county_name_norm"].isin(matched))]
            if sub["total_votes"].sum() > 0:
                share = float(sub["votes_dem"].sum() / sub["total_votes"].sum())
                rows.append({
                    "coc_id": coc_id, "state": state,
                    "matched_counties": ";".join(matched),
                    "n_matched": len(matched),
                    "biden_share": round(share, 4),
                    "source": "county_matched",
                })
                matched_count += 1
                continue
        # Fallback: state-level
        share = state_shares.get(state)
        if share is not None:
            rows.append({
                "coc_id": coc_id, "state": state,
                "matched_counties": "",
                "n_matched": 0,
                "biden_share": round(share, 4),
                "source": "state_fallback",
            })
            fallback_count += 1
        else:
            rows.append({
                "coc_id": coc_id, "state": state,
                "matched_counties": "",
                "n_matched": 0,
                "biden_share": "",
                "source": "unknown",
            })
            unknown_count += 1

    with OUT.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {OUT}")
    print(f"  county_matched: {matched_count}")
    print(f"  state_fallback: {fallback_count}")
    print(f"  unknown:        {unknown_count}")

    # Append to analysis-ready xlsx as a new sheet
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ANALYSIS_XLSX)
        if "iv_county" in wb.sheetnames:
            del wb["iv_county"]
        ws = wb.create_sheet("iv_county")
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append([r[k] for k in rows[0].keys()])
        wb.save(ANALYSIS_XLSX)
        print(f"appended iv_county sheet to {ANALYSIS_XLSX.name}")
    except Exception as e:
        print(f"Warning: couldn't append sheet ({e})")

    # Quick stats
    shares = [r["biden_share"] for r in rows if r["biden_share"] != ""]
    print(f"\nBiden share distribution ({len(shares)} CoCs):")
    print(f"  mean={sum(shares)/len(shares):.3f}")
    print(f"  min={min(shares):.3f}, max={max(shares):.3f}")
    print(f"  >0.5: {sum(1 for s in shares if s > 0.5)}")
    print(f"  <=0.5: {sum(1 for s in shares if s <= 0.5)}")


if __name__ == "__main__":
    main()
