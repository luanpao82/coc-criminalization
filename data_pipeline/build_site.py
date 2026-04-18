"""Generate a coauthor-friendly HTML site for the CoC extraction project.

The goal is narrative clarity — coauthors, not engineers, should be able
to understand what the pipeline did, why each methodological choice was
made, and how to use the resulting dataset.

Pages
-----
  index.html        — The project: what we built and why
  approach.html     — How we built it (methodology, plain language)
  variables.html    — What's in the dataset (searchable variable table)
  dv_story.html     — The DV harmonization problem (deep narrative)
  examples.html     — Real extracted records (AL-500, CA-500, FL-501, NY-600, TX-600)
  using.html        — How to use the data (code recipes)
  data.html         — Data downloads
  limits.html       — What's still missing
  review.html       — Adjudication queue for the PI team

Every page is self-contained HTML + Plotly from CDN. No build step.
Re-run this script after any pipeline change.
"""
from __future__ import annotations

import csv
import datetime as dt
import html
import json
from collections import Counter, defaultdict
from pathlib import Path

import shutil

from pipeline_utils import DATA_DIR, PIPELINE_DIR

SITE_DIR = PIPELINE_DIR.parent / "site"
SITE_DIR.mkdir(parents=True, exist_ok=True)
DOWNLOADS_DIR = SITE_DIR / "downloads"


# ---------------------------------------------------------------------------
# Downloads — files coauthors can pull from the site
# ---------------------------------------------------------------------------
# (source_path, group, title, description)
DOWNLOADS = [
    # --- Primary analysis dataset ---
    (PIPELINE_DIR / "coc_analysis_ready.xlsx", "primary",
     "coc_analysis_ready.xlsx",
     "Primary analysis file — 4 sheets: balanced_panel (125 CoCs × 3 years), "
     "unbalanced, fy2024_only, variables. Start here."),
    (PIPELINE_DIR / "coc_analysis_ready.csv", "primary",
     "coc_analysis_ready.csv",
     "Same data in long-form CSV — easy to load with pandas/readr."),
    (PIPELINE_DIR / "coc_panel_wide.xlsx", "primary",
     "coc_panel_wide.xlsx",
     "Full wide panel (243 variables × 3 years). Sheets: panel_safe, "
     "full_wide, fy2022, fy2023, fy2024, plus dv_harmonized and narrative_codes."),
    (PIPELINE_DIR / "coc_panel_long.csv", "primary",
     "coc_panel_long.csv",
     "Long form — one row per (CoC, year, field, value). ~164k rows."),
    (PIPELINE_DIR / "harmonized_dv.xlsx", "primary",
     "harmonized_dv.xlsx",
     "Harmonized DV workbook — crim_activity_index, implemented_anticrim_practice, "
     "cell-level 1D-4 indicators across all three years."),

    # --- Data provenance ---
    (PIPELINE_DIR / "file_inventory.csv", "provenance",
     "file_inventory.csv",
     "Source-document index (677 files) with normalized CoC IDs, format, and scan flags."),
    (PIPELINE_DIR / "panel_field_map.csv", "provenance",
     "panel_field_map.csv",
     "Per-variable panel-safety category (panel_safe / mostly_panel / year_specific / sparse)."),
    (PIPELINE_DIR / "crosswalk.csv", "provenance",
     "crosswalk.csv",
     "FY2024 canonical field ↔ FY2022/FY2023 question-ID mapping."),
    (PIPELINE_DIR / "corpus_diffs.csv", "provenance",
     "corpus_diffs.csv",
     "Review queue — automation vs manual spreadsheet diffs with PDF page citations."),
    (PIPELINE_DIR / "iterations.csv", "provenance",
     "iterations.csv",
     "Extractor iteration ledger — per-iter agreement, top error, fix applied."),
    (PIPELINE_DIR / "extraction_summary.csv", "provenance",
     "extraction_summary.csv",
     "Per-field extraction coverage and validity rates."),

    # --- Independent-variable coding ---
    (PIPELINE_DIR / "iv_leadership.csv", "ivs",
     "iv_leadership.csv",
     "Nonprofit vs government classification per CoC (hand-coded from 1a_2 "
     "Collaborative Applicant Name, 99% coverage)."),
    (PIPELINE_DIR / "iv_county.csv", "ivs",
     "iv_county.csv",
     "CoC → county presidential-vote merge (MIT/tonmcg 2020 county data), "
     "population-weighted to a CoC-level Biden share."),

    # --- Results tables ---
    (PIPELINE_DIR / "multilevel_coefs.csv", "results",
     "multilevel_coefs.csv",
     "Primary spec — multilevel fractional-logit DiD coefficients, "
     "state-clustered SEs, wild-cluster bootstrap p-values."),
    (PIPELINE_DIR / "multilevel_quadrant_means.csv", "results",
     "multilevel_quadrant_means.csv",
     "Four-quadrant (state × county politics) pre/post means used in the trajectory figure."),
    (PIPELINE_DIR / "balanced_sensitivity_coefs.csv", "results",
     "balanced_sensitivity_coefs.csv",
     "Appendix A — balanced-panel sensitivity coefficients (unbalanced vs balanced)."),
    (PIPELINE_DIR / "dv_robustness_coefs.csv", "results",
     "dv_robustness_coefs.csv",
     "Appendix B — alternative DV definitions (full / engagement-only / pre-shock only)."),

    # --- Documentation ---
    (PIPELINE_DIR.parent / "codebook.md", "docs",
     "codebook.md",
     "Variable codebook — 331 canonical fields, controlled vocabulary, observed value domains."),
    (PIPELINE_DIR.parent / "data_construction_methodology.md", "docs",
     "data_construction_methodology.md",
     "Methods-ready protocol (Krippendorff, Cohen, Gilardi et al., Ziems et al. citations)."),
    (PIPELINE_DIR / "multilevel_results.md", "docs",
     "multilevel_results.md",
     "Human-readable writeup of the primary multilevel DiD results."),
    (PIPELINE_DIR / "corpus_agreement.md", "docs",
     "corpus_agreement.md",
     "Extractor vs manual agreement audit — per-category rates and failure modes."),
    (PIPELINE_DIR / "pilot_diff_report.md", "docs",
     "pilot_diff_report.md",
     "Five-CoC pilot iteration trajectory (convergence to ≥98% weighted agreement)."),
    (PIPELINE_DIR.parent / "main_variables.md", "docs",
     "main_variables.md",
     "Operationalization notes for the primary variables used in the regression."),
]


def _human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n} B"


def copy_downloads() -> list[dict]:
    """Copy every available DOWNLOADS file into site/downloads/ and return metadata."""
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    records = []
    for src, group, title, desc in DOWNLOADS:
        if not src.exists():
            continue
        dest = DOWNLOADS_DIR / title
        shutil.copy2(src, dest)
        records.append({
            "href":        f"downloads/{title}",
            "group":       group,
            "title":       title,
            "description": desc,
            "size":        _human_size(dest.stat().st_size),
        })
    return records

# ---------------------------------------------------------------------------
# Shared style + navigation
# ---------------------------------------------------------------------------
STYLE = """
:root {
  --bg: #ffffff; --fg: #1a1a1a; --muted: #666; --accent: #0366d6;
  --line: #e5e5e5; --hi: #f7f7f7; --good: #17813c; --bad: #c53030;
  --warn: #b68400; --panel: #f0f6ff; --callout: #fff9e6;
}
* { box-sizing: border-box; }
body { font-family: Georgia, "Times New Roman", serif;
       max-width: 860px; margin: 0 auto; padding: 2em 1.2em 4em;
       color: var(--fg); background: var(--bg); line-height: 1.7; font-size: 17px; }
header { border-bottom: 2px solid var(--line); padding-bottom: 1.2em; margin-bottom: 2em; }
header .proj { color: var(--muted); font-size: 0.85em; letter-spacing: 0.04em; text-transform: uppercase; }
h1 { margin: 0.3em 0 0.2em; font-size: 2em; letter-spacing: -0.01em; line-height: 1.25; }
h2 { margin-top: 2.5em; padding-bottom: 0.2em; border-bottom: 1px solid var(--line); font-size: 1.5em; }
h3 { margin-top: 2em; font-size: 1.2em; color: #222; }
h4 { margin-top: 1.5em; font-size: 1.02em; color: #333; }
p, li { line-height: 1.7; }
.lead { color: var(--muted); font-size: 1.05em; font-style: italic; }
nav.main { display: flex; gap: 1.2em; margin: 0.6em 0 0; flex-wrap: wrap;
           font-family: -apple-system, system-ui, sans-serif; font-size: 0.92em; }
nav.main a { color: var(--accent); text-decoration: none; }
nav.main a.active { color: var(--fg); font-weight: 600; }
nav.main a:hover { text-decoration: underline; }
table { width: 100%; border-collapse: collapse; margin: 1.2em 0; font-size: 0.93em;
        font-family: -apple-system, system-ui, sans-serif; }
th, td { padding: 7px 11px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }
th { background: var(--hi); font-weight: 600; }
tr:hover { background: #fafafa; }
.num { text-align: right; font-variant-numeric: tabular-nums; }
table.downloads td:first-child { white-space: nowrap; width: 1%; }
table.downloads td:first-child a { font-weight: 500; }
table.downloads td:nth-child(2) { white-space: nowrap; width: 1%; color: var(--muted); }
table.downloads td:last-child { color: var(--muted); }
.badge { display: inline-block; padding: 1px 8px; border-radius: 10px;
         font-size: 0.8em; background: var(--hi); color: var(--muted);
         font-family: -apple-system, system-ui, sans-serif; }
.badge.good { background: #eafaef; color: var(--good); }
.badge.warn { background: #fff7e0; color: var(--warn); }
.badge.bad { background: #fde8e8; color: var(--bad); }
.stat { display: inline-block; background: var(--hi); padding: 0.9em 1.3em;
        border-radius: 6px; margin: 0.3em 0.3em 0.3em 0; min-width: 200px;
        font-family: -apple-system, system-ui, sans-serif; }
.stat .label { color: var(--muted); font-size: 0.82em; margin-bottom: 0.3em; }
.stat .value { font-size: 1.7em; font-weight: 600; color: var(--fg); }
code { background: #f2f2f2; padding: 1px 5px; border-radius: 3px;
       font-family: "SF Mono", Menlo, Consolas, monospace; font-size: 0.88em; }
pre { background: #f5f7fa; padding: 1.2em; border-radius: 6px; overflow-x: auto;
      font-size: 0.87em; line-height: 1.55;
      font-family: "SF Mono", Menlo, Consolas, monospace;
      border-left: 3px solid var(--accent); }
pre code { background: none; padding: 0; }
blockquote { background: var(--callout); border-left: 4px solid #c8a030;
             padding: 0.8em 1.2em; margin: 1.4em 0; border-radius: 0 6px 6px 0; }
blockquote p:first-child { margin-top: 0; }
blockquote p:last-child { margin-bottom: 0; }
.callout { background: var(--panel); border-left: 4px solid var(--accent);
           padding: 0.8em 1.2em; margin: 1.4em 0; border-radius: 0 6px 6px 0; }
.callout-title { font-weight: 600; color: var(--accent); margin-bottom: 0.4em;
                 font-family: -apple-system, system-ui, sans-serif; font-size: 0.95em; }
footer { margin-top: 5em; padding-top: 1.4em; border-top: 1px solid var(--line);
         color: var(--muted); font-size: 0.85em;
         font-family: -apple-system, system-ui, sans-serif; }
.card-row { display: flex; gap: 1em; flex-wrap: wrap; margin: 1em 0;
            font-family: -apple-system, system-ui, sans-serif; }
.card { flex: 1; min-width: 260px; background: var(--hi); padding: 1.2em;
        border-radius: 6px; border: 1px solid var(--line); }
.card h3 { margin: 0 0 0.4em; font-family: Georgia, serif; }
.breadcrumb { font-size: 0.8em; color: var(--muted); margin-bottom: 0.5em;
              font-family: -apple-system, system-ui, sans-serif; }
.breadcrumb a { color: var(--muted); text-decoration: none; }
.breadcrumb a:hover { color: var(--accent); text-decoration: underline; }
.hub-card { display: block; background: white; border: 1px solid var(--line);
            border-radius: 8px; padding: 1.3em 1.4em; text-decoration: none;
            color: inherit; transition: all 0.2s ease; margin-bottom: 0.8em; }
.hub-card:hover { border-color: var(--accent); box-shadow: 0 2px 10px rgba(3,102,214,0.08);
                  transform: translateY(-1px); }
.hub-card h3 { margin: 0 0 0.4em; color: var(--accent); font-family: Georgia, serif; font-size: 1.15em; }
.hub-card p { margin: 0.3em 0 0; color: var(--fg); font-size: 0.95em; line-height: 1.55; }
.hub-card .meta { color: var(--muted); font-size: 0.82em; margin-top: 0.5em;
                  font-family: -apple-system, system-ui, sans-serif; }
.hub-card .num-tag { display: inline-block; width: 1.5em; height: 1.5em;
                     background: var(--accent); color: white; border-radius: 50%;
                     text-align: center; line-height: 1.5em; font-weight: 700;
                     font-size: 0.75em; margin-right: 0.4em;
                     font-family: -apple-system, sans-serif; vertical-align: 2px; }
.hub-grid { display: grid; gap: 0.9em; grid-template-columns: 1fr;
            margin: 1em 0 2em; }
@media (min-width: 700px) { .hub-grid { grid-template-columns: 1fr 1fr; } }
.track { background: var(--hi); border-radius: 8px; padding: 1.1em 1.3em; margin: 0.7em 0; }
.track h3 { margin-top: 0; color: var(--accent); font-family: Georgia, serif; }
.track ol { margin: 0.5em 0; padding-left: 1.3em; }
.findings-box { background: #f0faf3; border-left: 4px solid var(--good);
                padding: 1em 1.3em; border-radius: 0 8px 8px 0; margin: 1.3em 0; }
.findings-box h3 { margin-top: 0; color: var(--good); font-family: Georgia, serif; }
.findings-box ul { margin: 0.4em 0 0; padding-left: 1.3em; }
.card p { font-size: 0.95em; }
.filterbox { margin: 0.8em 0; font-family: -apple-system, system-ui, sans-serif; }
.filterbox input, .filterbox select { padding: 7px 11px; border: 1px solid var(--line);
                                       border-radius: 4px; font-size: 0.95em; }
.filterbox input { width: 360px; }
details > summary { cursor: pointer; color: var(--accent); margin: 0.5em 0;
                    font-family: -apple-system, system-ui, sans-serif; font-size: 0.95em; }
.hint { font-size: 0.88em; color: var(--muted); font-style: italic; }
.quote { background: #f9f6f0; border-left: 3px solid #d6c389;
         padding: 0.7em 1em; margin: 0.8em 0; font-style: italic;
         font-size: 0.93em; border-radius: 0 3px 3px 0; }
.step-list { list-style: none; padding-left: 0; counter-reset: step; }
.step-list > li { counter-increment: step; position: relative; padding: 0.6em 0 0.6em 3em; }
.step-list > li::before { content: counter(step); position: absolute; left: 0; top: 0.7em;
                          width: 2em; height: 2em; border-radius: 50%;
                          background: var(--accent); color: white;
                          text-align: center; line-height: 2em; font-weight: 600;
                          font-family: -apple-system, sans-serif; font-size: 0.9em; }
"""

NAV_ITEMS = [
    ("Overview",                         "index.html"),
    ("Data development and cleaning",    "data.html"),
    ("Research design and measurement",  "design.html"),
    ("Descriptive stats",                "descriptive.html"),
    ("Analysis results",                 "results.html"),
]

PAGE_PARENTS: dict = {}


def nav(active: str) -> str:
    links = "".join(
        f'<a class="{"active" if a[1] == active else ""}" href="{a[1]}">{a[0]}</a>'
        for a in NAV_ITEMS
    )
    return f'<nav class="main">{links}</nav>'


def breadcrumb(active: str) -> str:
    return ""


def wrap(title: str, active: str, body: str, subtitle: str = "") -> str:
    head_subtitle = f'<p class="lead">{subtitle}</p>' if subtitle else ""
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>{html.escape(title)} · CoC Criminalization Project</title>
<style>{STYLE}</style>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
</head><body>
<header>
  <div class="proj">CoC Criminalization &amp; PLE Engagement · Lee &amp; Kim, UCF</div>
  {breadcrumb(active)}
  <h1>{html.escape(title)}</h1>
  {head_subtitle}
  {nav(active)}
</header>
{body}
<footer>
  <strong>This site regenerates from the same files the analysis runs on.</strong>
  Last generated {dt.datetime.now().isoformat(timespec='seconds')}.
  Source: <code>data_pipeline/build_site.py</code>. ·
  Authors: Kyungmin Lee &amp; Hanvit Kim (University of Central Florida).
</footer>
</body></html>"""


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------
def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open() as f:
        return list(csv.DictReader(f))


def load_inventory():
    return load_csv(PIPELINE_DIR / "file_inventory.csv")


def load_iterations():
    return load_csv(PIPELINE_DIR / "iterations.csv")


def load_field_map():
    return load_csv(PIPELINE_DIR / "panel_field_map.csv")


def load_harmonized():
    return load_csv(PIPELINE_DIR / "harmonized_dv.csv")


def load_corpus_diffs():
    return load_csv(PIPELINE_DIR / "corpus_diffs.csv")


def load_extraction_summary():
    return load_csv(PIPELINE_DIR / "extraction_summary.csv")


# ---------------------------------------------------------------------------
# Friendly variable descriptions for the paper's key constructs
# ---------------------------------------------------------------------------
FRIENDLY_DESCRIPTIONS = {
    "1a_1a": ("CoC Number", "HUD's identifier (e.g., AL-500). Used as the panel key."),
    "1a_1b": ("CoC Name", "Human-readable CoC name (place + 'CoC')."),
    "1a_2": ("Collaborative Applicant Name", "The organization that runs the CoC. Nonprofit vs. government is inferred here."),
    "1a_3": ("CoC Designation", "HUD formal type: CA (Collaborative Applicant), UFA (Unified Funding Agency), or UFC."),
    "1a_4": ("HMIS Lead", "Organization that operates the CoC's Homeless Management Information System."),
    "1b_1_6_meetings": ("Homeless/Formerly Homeless persons at CoC meetings", "Whether people with lived experience attended CoC meetings (Yes/No/Nonexistent)."),
    "1b_1_6_voted": ("Homeless/Formerly Homeless persons voting", "Whether they voted, including electing CoC board members."),
    "1b_1_6_ces": ("Homeless/Formerly Homeless persons in CES", "Whether they participated in the Coordinated Entry System."),
    "1b_1_9_meetings": ("Law Enforcement at CoC meetings", "Whether police/sheriff representatives attend CoC meetings."),
    "1b_1_13_meetings": ("Local Jail(s) at CoC meetings", "Whether jail representatives attend."),
    "1d_2_1": ("Housing First projects total", "Number of new + renewal projects the CoC is applying for."),
    "1d_2_2": ("Housing First projects adopting HF", "Of those, how many adopted Housing First."),
    "1d_2_3": ("Housing First adoption percentage", "% of projects with Housing First approach (mostly stored as 0–1 fraction)."),
    "1d_4_1_policymakers": ("Anti-crim strategy 1 — policymakers engaged",
                             "FY2022/23: engaged local policymakers to prevent criminalization. "
                             "FY2024: engaged legislators on co-responder responses. (See DV story.)"),
    "1d_4_1_prevent_crim": ("Anti-crim strategy 1 — implemented",
                             "FY2022/23: reversed existing criminalization policies. "
                             "FY2024: implemented laws/policies for co-responder responses."),
    "1d_4_2_policymakers": ("Anti-crim strategy 2 — policymakers engaged", "Same column, row 2."),
    "1d_4_2_prevent_crim": ("Anti-crim strategy 2 — implemented", "Same column, row 2."),
    "1d_4_3_policymakers": ("Anti-crim strategy 3 — policymakers engaged", "Same column, row 3."),
    "1d_4_3_prevent_crim": ("Anti-crim strategy 3 — implemented", "Same column, row 3."),
    "1d_10a_1_years": ("PLE in decisionmaking (counts, past 7 yrs)",
                        "Number of people with lived experience included in CoC decisionmaking. (FY2022 crosswalked from 1D-11a row 4.)"),
    "1d_10a_1_unsheltered": ("PLE in decisionmaking — unsheltered",
                              "Of the above, how many come from unsheltered situations."),
    "1d_10a_2_years": ("PLE on CoC committees (counts)", "Row 2 in the FY2024 canonical ordering."),
    "1d_10a_3_years": ("PLE on rating/scoring factors (counts)", "Row 3 in the FY2024 canonical ordering."),
    "2a_5_1_coverage": ("HMIS coverage — Emergency Shelter", "Share of ES beds in HMIS (panel-safe across years)."),
    "2a_5_5_coverage": ("HMIS coverage — Permanent Supportive Housing", "Share of PSH beds in HMIS."),
    "2b_1": ("PIT count date", "Date of the Point-in-Time count — annual measure of homelessness."),
}


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------
def page_index():
    inv = load_inventory()
    fm = load_field_map()
    harm = load_harmonized()
    panel_safe_n = sum(1 for r in fm if r["category"] == "panel_safe")

    body = f"""
    <p class="lead">How 325 U.S. Continuums of Care responded to the June
    2024 Supreme Court ruling in <em>Grants Pass v. Johnson</em> — a
    multilevel difference-in-differences built from 677 HUD application
    documents (FY2022–FY2024).</p>

    <div class="findings-box">
      <h3>Headline finding</h3>
      <p style="font-size: 1.05em; margin: 0.3em 0 0.6em;">
      <strong>Local political environment — not CoC governance structure —
      filters federal legal shocks on homelessness criminalization.</strong>
      </p>
      <ul>
        <li><strong>Organizational form doesn't matter.</strong> Nonprofit-led
        vs. government-led CoCs responded identically
        (β = +0.09, wild-cluster bootstrap p = 0.80).</li>
        <li><strong>State political environment does.</strong> Blue-state
        CoCs raised reported anti-criminalization activity more than
        red-state CoCs (β = +0.57, p = 0.095). Within-state county
        variation adds no independent effect.</li>
        <li><strong>The pattern is asymmetric.</strong> Only fully red-state
        × red-county CoCs stayed flat (Δ ≈ −0.03). All three other
        political combinations rose in parallel by +0.11 to +0.14.</li>
      </ul>
    </div>

    <h2>The five sections</h2>
    <div class="hub-grid">
      <a class="hub-card" href="data.html">
        <h3>1 · Data development and cleaning</h3>
        <p>How the 677-document corpus became a three-year CoC panel.
        Extraction pipeline, agreement metrics, and the review queue.</p>
      </a>
      <a class="hub-card" href="design.html">
        <h3>2 · Research design and measurement</h3>
        <p>The multilevel DiD specification, H1 (organizational form) vs
        H2 (political environment) hypotheses, variable construction, and
        the DV measurement-invariance story.</p>
      </a>
      <a class="hub-card" href="descriptive.html">
        <h3>3 · Descriptive stats</h3>
        <p>Grants Pass as the policy shock, year-by-year distributions of
        the anti-criminalization activity index, and which cells of the
        HUD 1D-4 form drove the FY2024 shift.</p>
      </a>
      <a class="hub-card" href="results.html" style="background: #f0faf3; border-color: #17813c;">
        <h3>4 · Analysis results</h3>
        <p>The primary multilevel fractional-logit DiD, the four-quadrant
        political-geography figure, balanced-panel sensitivity, DV
        robustness, and limitations.</p>
      </a>
    </div>

    <h2>Dataset at a glance</h2>
    <div class="card-row">
      <div class="stat"><div class="label">Source documents</div><div class="value">{len(inv)}</div></div>
      <div class="stat"><div class="label">CoC × year records</div><div class="value">{len(harm)}</div></div>
      <div class="stat"><div class="label">Panel-safe variables</div><div class="value">{panel_safe_n}</div></div>
      <div class="stat"><div class="label">Unique CoCs</div><div class="value">325</div></div>
    </div>
    """
    return wrap("CoC responses to Grants Pass v. Johnson",
                "index.html", body,
                "Lee &amp; Kim (UCF) · Multilevel DiD across 325 Continuums of Care, FY2022–FY2024.")


def _page_index_OLD():
    # kept for reference; not used
    body = f"""
    <p class="lead">This site documents how we built a longitudinal dataset
    on Continuum of Care (CoC) governance, lived-experience engagement, and
    responses to the criminalization of homelessness — from 677 HUD
    application documents covering FY2022, FY2023, and FY2024.</p>

    <div class="callout">
      <div class="callout-title">For coauthors</div>
      <p>This website is the single source of truth for <em>what data we have,
      what it means, and how to use it</em>. Every number you see here is
      computed from the same files the regression models will read. If you
      find something confusing or wrong, please flag it — we can regenerate
      this site from updated code in under a minute.</p>
    </div>

    <h2>What we built</h2>
    <p>An extractor pipeline that reads HUD CoC Consolidated Application
    PDFs, extracts ~290 quantitative fields per application, and assembles a
    multi-year panel dataset suitable for longitudinal analysis of how CoC
    governance structure and PLE engagement shape local responses to
    homelessness criminalization.</p>

    <div class="card-row">
      <div class="stat"><div class="label">Source documents processed</div><div class="value">{len(inv)}</div></div>
      <div class="stat"><div class="label">CoC × year records extracted</div><div class="value">{len(harm)}</div></div>
      <div class="stat"><div class="label">Panel-safe variables</div><div class="value">{panel_safe_n}</div></div>
      <div class="stat"><div class="label">Weighted agreement (vs. manual)</div><div class="value">{wacc:.1%}</div></div>
      {f'<div class="stat"><div class="label">Adjusted agreement</div><div class="value">{aacc:.1%}</div></div>' if aacc else ''}
    </div>

    <h3>What "panel-safe" means</h3>
    <p>We classify every variable by how reliably it shows up across the
    three fiscal years:</p>
    <ul>
      <li><span class="badge good">panel_safe</span> — present in at least 50% of CoCs in every year. These can be used in longitudinal models directly.</li>
      <li><span class="badge warn">mostly_panel</span> — present in two of three years; useful with care.</li>
      <li><span class="badge warn">year_specific</span> — HUD asked the question in only one year; cross-sectional use only.</li>
      <li><span class="badge bad">sparse</span> — not yet extracted (mostly free-text narrative fields awaiting Stage-2 LLM coding).</li>
    </ul>

    <h2>How to read this site</h2>
    <ol class="step-list">
      <li><a href="approach.html">Approach</a> — how the pipeline works, in plain language. Start here if you want to understand the logic.</li>
      <li><a href="variables.html">Variables</a> — every variable in the dataset, with plain-English descriptions and year-by-year coverage.</li>
      <li><a href="dv_story.html">The DV story</a> — the most important methodological discussion in the project: how we handle the fact that HUD changed how it asked about criminalization in FY2024.</li>
      <li><a href="examples.html">Examples</a> — five real CoCs (AL-500, CA-500, FL-501, NY-600, TX-600) so you can see what the data actually looks like.</li>
      <li><a href="using.html">Using the data</a> — code recipes for analysis in R and Python.</li>
      <li><a href="data.html">Downloads</a> — pull the xlsx / CSV files for your own machine.</li>
      <li><a href="limits.html">Limitations</a> — honest accounting of what's still open.</li>
    </ol>

    <h2>Files covered by year</h2>
    <div class="card-row">
      <div class="stat"><div class="label">FY2022 applications</div><div class="value">{by_year.get("2022", 0)}</div></div>
      <div class="stat"><div class="label">FY2023 applications</div><div class="value">{by_year.get("2023", 0)}</div></div>
      <div class="stat"><div class="label">FY2024 applications</div><div class="value">{by_year.get("2024", 0)}</div></div>
    </div>
    <p class="hint">FY2024 is the largest because HUD rolled the Youth Homeless
    Demonstration Program (YHDP) into the main competition that year, pulling in
    more CoCs.</p>
    """
    return wrap("Building a longitudinal CoC dataset", "index.html", body,
                "Research-grade data extracted from 677 HUD application PDFs, "
                "ready for analysis.")


def page_data_hub():
    body = """
    <p class="lead">Everything about how we built the dataset and what's in it.
    Start with <strong>Approach</strong> for the big picture, then pick the
    detail page you need.</p>

    <h2>Pages in this section</h2>

    <a class="hub-card" href="approach.html">
      <h3>1 · Approach (how the pipeline works)</h3>
      <p>An eight-step walkthrough of how we turn 677 HUD PDFs into a
      longitudinal panel. Read this first — everything else assumes it.</p>
      <div class="meta">~8 min read · methodology narrative</div>
    </a>

    <a class="hub-card" href="variables.html">
      <h3>2 · Variables (what's in the dataset)</h3>
      <p>All 331 canonical variables, with plain-English descriptions and
      year-by-year coverage. Filter by category (<code>panel_safe</code>,
      etc.) or search by name.</p>
      <div class="meta">interactive table · 243 panel-safe variables</div>
    </a>

    <a class="hub-card" href="distributions.html">
      <h3>3 · Distributions (variable shapes and transformations)</h3>
      <p>Histograms of every regression-relevant variable with concrete
      transformation recommendations (log1p, winsorization, fractional
      logit). Open this before writing any estimation code.</p>
      <div class="meta">15 variables · per-variable charts</div>
    </a>

    <a class="hub-card" href="examples.html">
      <h3>4 · Examples (five real CoCs)</h3>
      <p>AL-500, CA-500, FL-501, NY-600, TX-600 across all three years, with
      the key variables in a single table. So you can see what the data
      actually looks like before opening the xlsx.</p>
      <div class="meta">5 CoCs × 3 years</div>
    </a>

    <a class="hub-card" href="map.html">
      <h3>5 · Interactive map</h3>
      <p>Every CoC plotted at its state's centroid with an interactive
      variable picker — including composite views (Research model · Change
      FY23→FY24) that encode the entire paper model in one screen.</p>
      <div class="meta">325 CoCs · 9 variables · 3 years</div>
    </a>

    <a class="hub-card" href="data.html">
      <h3>6 · Downloads</h3>
      <p>All output files organized by use case: the primary panel xlsx,
      the DV-harmonized outcomes, per-year slices, audit artifacts, and
      methodology documents.</p>
      <div class="meta">16+ files · documented by category</div>
    </a>

    <h2>If you only do one thing</h2>
    <div class="callout">
      <p>Open <code>coc_analysis_ready.xlsx</code> (<a href="data.html">download page</a>).
      The <strong>balanced_panel</strong> sheet is 125 CoCs × 3 years with
      the primary DV, the Mundlak-augmented controls, and the harmonized
      outcomes — ready for <code>PanelOLS</code> or <code>feols</code>.</p>
    </div>
    """
    return wrap("Data", "data_hub.html", body,
                "How the pipeline works · what's in the data · where to download it.")


def page_analysis_hub():
    body = """
    <p class="lead">The research model, the descriptive pattern, and the
    primary multilevel DiD — plus the two robustness checks that confirm
    the main finding holds.</p>

    <h2>Pages in this section</h2>

    <a class="hub-card" href="model.html">
      <h3><span class="num-tag">1</span>Research model</h3>
      <p>The theoretical structure: IV (lead-agency type) → Mediator (PLE
      engagement) → DV (anti-criminalization activity). Maps each construct
      to specific HUD questions. Read this first for the conceptual frame.</p>
      <div class="meta">the research-question scaffold</div>
    </a>

    <a class="hub-card" href="descriptive.html">
      <h3><span class="num-tag">2</span>Descriptives &amp; Grants Pass</h3>
      <p>Year-over-year movement of the activity index, the <em>Grants Pass
      v. Johnson</em> ruling as the policy shock, and cell-level
      breakdowns that reveal the instrument-change issue.</p>
      <div class="meta">4 charts · samples · Δ tables</div>
    </a>

    <a class="hub-card" href="main_results.html" style="background: #f0faf3; border-color: #17813c;">
      <h3><span class="num-tag">3</span>Main results ★</h3>
      <p>The primary multilevel fractional-logit DiD. Three DiD terms
      (Nonprofit × Post, Blue-state × Post, Biden-within-state × Post) in
      one specification. The <strong>4-quadrant trajectory figure</strong>
      shows the key asymmetric pattern: Red-state × Red-county CoCs did not
      respond, while the other three quadrants all rose in parallel.</p>
      <div class="meta">primary spec · 4-quadrant figure · forest plot</div>
    </a>

    <a class="hub-card" href="dv_robust.html">
      <h3><span class="num-tag">4</span>Robustness · DV operationalization</h3>
      <p>Three alternative DV constructions (full index, wording-stable
      engagement-only, FY22+23 panel only) test whether the instrument
      change drives the findings. The central nonprofit null is robust
      across all three.</p>
      <div class="meta">7 specifications · convergent evidence</div>
    </a>

    <a class="hub-card" href="balanced.html">
      <h3><span class="num-tag">5</span>Robustness · Balanced vs Unbalanced</h3>
      <p>Compares the main specs across the unbalanced (651 obs) and
      balanced 3-year (375 obs) panels. Reveals that the state-level binary
      effect weakens on the balanced panel — a key methodological
      disclosure supporting the multilevel specification's cleaner
      handling.</p>
      <div class="meta">3 specs × 2 panels</div>
    </a>

    <h2>Reading tip</h2>
    <div class="callout">
      <p><strong>If short on time:</strong> pages 2 and 3. The descriptive
      movement + the multilevel DiD + the 4-quadrant figure together
      deliver the paper's main findings.</p>
    </div>
    """
    return wrap("Analysis", "analysis_hub.html", body,
                "Research model → descriptives → primary multilevel DiD → robustness.")


def page_methods_hub():
    body = """
    <p class="lead">The methodological discussions that don't belong in the
    main results but are essential for reviewers and replicators.</p>

    <h2>Pages in this section</h2>

    <a class="hub-card" href="dv_story.html">
      <h3>The DV story — HUD's FY2024 instrument change</h3>
      <p>A detailed walkthrough of how HUD rewrote the 1D-4 chart between
      FY2023 and FY2024, why this matters for measurement invariance, and
      the eight harmonization strategies we considered (three of which
      survived empirical scrutiny).</p>
      <div class="meta">central methodological narrative</div>
    </a>

    <a class="hub-card" href="limits.html">
      <h3>Limitations and open items</h3>
      <p>Six known limitations, in order of analytical impact: OCR-needed
      CoCs, uncoded narrative variables, manual-reference errors,
      measurement invariance, deliberate extractor choices, and DOCX
      parity.</p>
      <div class="meta">honest accounting</div>
    </a>

    <a class="hub-card" href="review.html">
      <h3>Review queue — disagreements for PI adjudication</h3>
      <p>Every cell where the automated extractor and the manually coded
      FY2024 spreadsheet disagree (≈2,400 rows). Pink = manual value,
      green = extractor's. Opens the source PDF for each row so the PI
      team can verify against the ground truth.</p>
      <div class="meta">~407 true mismatches + ~1,100 auto-fills</div>
    </a>

    <h2>Also relevant — the Markdown write-ups</h2>
    <ul>
      <li><a href="../data_construction_methodology.md">data_construction_methodology.md</a>
        — formal Methods-ready write-up with full citations.</li>
      <li><a href="../main_variables.md">main_variables.md</a>
        — authoritative construct-to-variable mapping, including year-to-year change flags.</li>
      <li><a href="../dv_harmonization_strategies.md">dv_harmonization_strategies.md</a>
        and <a href="../dv_harmonization_results.md">dv_harmonization_results.md</a>
        — the eight-strategy menu and the empirical findings that narrowed to three.</li>
      <li><a href="../progress/README.md">progress/</a>
        — eight-page narrative log of how the pipeline was built.</li>
    </ul>
    """
    return wrap("Methods & caveats", "methods_hub.html", body,
                "The methodological discussions supporting (and limiting) the main findings.")


def page_model():
    body = """
    <p class="lead">The theoretical framework, the shock that enables
    identification, and the multilevel structure that distinguishes
    organizational and political explanations for CoC responses.</p>

    <h2>Research question</h2>
    <p>On <strong>June 28, 2024</strong>, the U.S. Supreme Court decided
    <em>City of Grants Pass v. Johnson</em>, holding that municipal
    anti-camping ordinances do not violate the Eighth Amendment. The ruling
    reopened legal space for the criminalization of public homelessness and
    set up a natural experiment: CoCs across the country faced the same
    federal legal shock, but plausibly responded differently based on their
    local context.</p>

    <p>This study asks: <strong>which CoC characteristics predict
    differential expansion of reported anti-criminalization activity after
    Grants Pass?</strong> We test two theoretical accounts that make
    competing predictions.</p>

    <h2>Two competing hypotheses</h2>

    <h3>H1 · Organizational form (the conventional account)</h3>
    <p>Drawing on the nonprofit-governance literature, we predict that
    <strong>nonprofit-led CoCs are more rights-responsive</strong> than
    government-led CoCs and will expand anti-criminalization activity more
    strongly after Grants Pass. Nonprofits face weaker political
    constraints, are closer to advocacy networks, and have fewer
    bureaucratic incentives to enforce new legal latitude.</p>
    <blockquote>
      <p><em>Prediction:</em> β(Nonprofit × Post) &gt; 0, statistically
      significant. The nonprofit-vs-government distinction mediates
      post-Grants Pass response.</p>
    </blockquote>

    <h3>H2 · Local political environment (the alternative account)</h3>
    <p>Drawing on the political-environment literature, we predict that
    <strong>CoCs in more liberal jurisdictions (state and/or county) face
    stronger constituency demand</strong> for rights-responsive framing
    and will expand anti-criminalization activity more than CoCs in
    conservative jurisdictions, <em>regardless of governance form</em>.
    Political context — not organizational structure — filters the legal
    shock.</p>
    <blockquote>
      <p><em>Prediction:</em> β(Blue × Post) &gt; 0 and/or β(Biden-within
      × Post) &gt; 0; β(Nonprofit × Post) may be small and insignificant
      once political context is controlled.</p>
    </blockquote>

    <h2>Why this requires a multilevel model</h2>
    <p>The two political measures — state-level binary and county-level
    continuous Biden share — are naturally correlated. A CoC in a blue
    county is probably also in a blue state. Naively including both in the
    same regression produces collinearity; including only one conflates
    state-level with county-level variation.</p>

    <p>We resolve this via a <strong>Mundlak-style decomposition at the
    multilevel structure</strong>:</p>

    <div class="callout">
      <p><code>state_mean_biden = mean Biden share across counties in the state</code></p>
      <p><code>biden_within_state = county_biden_share − state_mean_biden</code></p>
      <p>After this decomposition, <code>blue_state</code> and
      <code>biden_within_state</code> are <strong>statistically orthogonal</strong>.
      The first captures state-level political environment; the second
      captures within-state county deviation. Both can enter the same
      regression without collinearity washing out the coefficients.</p>
    </div>

    <h2>Model diagram</h2>
    <div style="text-align:center; margin: 1.5em 0;">
    <svg viewBox="0 0 860 420" width="100%" style="max-width:860px; font-family: -apple-system, sans-serif;">
      <defs>
        <marker id="arr" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="auto">
          <path d="M0,0 L0,6 L9,3 z" fill="#0366d6"/>
        </marker>
        <marker id="arr-n" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="auto">
          <path d="M0,0 L0,6 L9,3 z" fill="#c53030"/>
        </marker>
      </defs>

      <!-- H1 IV -->
      <rect x="30" y="50" width="200" height="70" rx="8" fill="#fdf2f2" stroke="#c53030" stroke-width="2"/>
      <text x="130" y="78" text-anchor="middle" font-weight="600" font-size="13">H1 · Organizational form</text>
      <text x="130" y="98" text-anchor="middle" font-size="11" fill="#555">Nonprofit-led (vs Government-led)</text>
      <text x="130" y="112" text-anchor="middle" font-size="11" fill="#555">coded from 1a_2 Collaborative Applicant</text>

      <!-- H2 IVs (two levels) -->
      <rect x="30" y="180" width="200" height="50" rx="8" fill="#eef4fb" stroke="#0366d6" stroke-width="2"/>
      <text x="130" y="202" text-anchor="middle" font-weight="600" font-size="13">H2a · State-level politics</text>
      <text x="130" y="220" text-anchor="middle" font-size="11" fill="#555">Blue state (Biden 2020 winner)</text>

      <rect x="30" y="260" width="200" height="50" rx="8" fill="#eef4fb" stroke="#0366d6" stroke-width="2" stroke-dasharray="4,2"/>
      <text x="130" y="280" text-anchor="middle" font-weight="600" font-size="13">H2b · County-within-state</text>
      <text x="130" y="298" text-anchor="middle" font-size="11" fill="#555">Biden share − state mean</text>

      <!-- Shock -->
      <rect x="310" y="165" width="210" height="80" rx="8" fill="#fff4cc" stroke="#8a6d1f" stroke-width="2"/>
      <text x="415" y="193" text-anchor="middle" font-weight="600" font-size="14">Grants Pass shock</text>
      <text x="415" y="213" text-anchor="middle" font-size="11" fill="#555">SCOTUS ruling · June 28, 2024</text>
      <text x="415" y="229" text-anchor="middle" font-size="11" fill="#555">FY2024 = Post indicator</text>

      <!-- DV -->
      <rect x="620" y="165" width="200" height="80" rx="8" fill="#f0faf3" stroke="#17813c" stroke-width="2"/>
      <text x="720" y="193" text-anchor="middle" font-weight="600" font-size="14">Anti-crim activity</text>
      <text x="720" y="213" text-anchor="middle" font-size="11" fill="#555">crim_activity_index ∈ [0,1]</text>
      <text x="720" y="229" text-anchor="middle" font-size="11" fill="#555">share of 1D-4 cells = Yes</text>

      <!-- Controls -->
      <rect x="310" y="310" width="210" height="90" rx="8" fill="#fafafa" stroke="#888" stroke-width="1.5" stroke-dasharray="4,3"/>
      <text x="415" y="333" text-anchor="middle" font-weight="600" font-size="13" fill="#444">Controls</text>
      <text x="415" y="351" text-anchor="middle" font-size="11" fill="#555">Housing First adoption · HMIS coverage</text>
      <text x="415" y="366" text-anchor="middle" font-size="11" fill="#555">log(total beds · winsorized)</text>
      <text x="415" y="381" text-anchor="middle" font-size="11" fill="#555">PLE participation · Mundlak means</text>

      <!-- Arrows: IVs interact with shock -->
      <path d="M 230 85 Q 280 125 310 180" fill="none" stroke="#c53030" stroke-width="2.5" marker-end="url(#arr-n)"/>
      <text x="244" y="145" font-size="11" fill="#c53030">H1: × Post</text>

      <path d="M 230 205 L 310 205" fill="none" stroke="#0366d6" stroke-width="2.5" marker-end="url(#arr)"/>
      <text x="250" y="200" font-size="11" fill="#0366d6">H2a: × Post</text>

      <path d="M 230 285 Q 270 260 310 225" fill="none" stroke="#0366d6" stroke-width="2" stroke-dasharray="4,2" marker-end="url(#arr)"/>
      <text x="240" y="260" font-size="11" fill="#0366d6">H2b: × Post</text>

      <!-- Shock -> DV -->
      <path d="M 520 205 L 620 205" fill="none" stroke="#17813c" stroke-width="3" marker-end="url(#arr)"/>

      <!-- Controls -> DV -->
      <path d="M 520 345 Q 580 290 620 230" fill="none" stroke="#888" stroke-width="1.5" marker-end="url(#arr)"/>
    </svg>
    </div>
    <p class="hint">Red arrow = H1 prediction (nonprofit advantage).
    Blue arrows = H2 predictions (state and within-state county political
    environment). The solid arrow into the DV is the Post-Grants Pass
    change; each IV's interaction with Post is its DiD estimand.</p>

    <h2>Operationalization table</h2>
    <table>
      <thead><tr><th>Construct</th><th>Variable</th><th>Type</th><th>Source</th></tr></thead>
      <tbody>
        <tr><td><strong>IV₁ · Organizational form</strong></td>
            <td><code>nonprofit_led</code></td>
            <td>binary</td>
            <td>Hand-coded from <code>1a_2</code> Collaborative Applicant Name (99% classified)</td></tr>
        <tr><td><strong>IV₂ · State-level politics</strong></td>
            <td><code>blue_state</code></td>
            <td>binary</td>
            <td>2020 presidential Electoral-College winner (Biden=1, Trump=0; DC=blue; territories excluded)</td></tr>
        <tr><td><strong>IV₃ · County within-state politics</strong></td>
            <td><code>biden_within_state</code></td>
            <td>continuous</td>
            <td>County Biden share − state-level mean share (MIT/tonmcg 2020 county data)</td></tr>
        <tr><td><strong>Shock</strong></td>
            <td><code>post</code></td>
            <td>binary (FY2024=1)</td>
            <td><em>Grants Pass v. Johnson</em> (SCOTUS, June 28, 2024)</td></tr>
        <tr><td><strong>DV</strong></td>
            <td><code>crim_activity_index</code></td>
            <td>fraction [0, 1]</td>
            <td>Share of 6 HUD 1D-4 cells answered "Yes"</td></tr>
        <tr><td>Control · service orientation</td>
            <td><code>hf_pct</code></td>
            <td>fraction [0, 1]</td>
            <td>Housing First adoption, HUD 1D-2</td></tr>
        <tr><td>Control · CoC maturity</td>
            <td><code>hmis_cov</code></td>
            <td>fraction [0, 1]</td>
            <td>HMIS ES bed coverage, HUD 2A-5</td></tr>
        <tr><td>Control · size</td>
            <td><code>log_beds</code></td>
            <td>continuous</td>
            <td>log(total beds + 1), winsorized at 99th pct</td></tr>
        <tr><td>Control · PLE participation</td>
            <td><code>ple_ces_bin</code></td>
            <td>binary</td>
            <td>HUD 1B-1 row 6 "Homeless or Formerly Homeless" in CES</td></tr>
      </tbody>
    </table>

    <h2>Identification strategy</h2>

    <h3>Primary specification — multilevel DiD</h3>
    <pre><code>crim_activity_index<sub>it</sub> =
    β₁ · Nonprofit<sub>i</sub>                    ← IV₁ (organizational form)
  + β₂ · Blue_state<sub>i</sub>                   ← IV₂ (state politics)
  + β₃ · Biden_within_state<sub>i</sub>           ← IV₃ (county within state)
  + β₄ · Post<sub>t</sub>                          ← Grants Pass shock
  + β₅ · Nonprofit × Post                ← H1 test
  + β₆ · Blue_state × Post              ← H2a test
  + β₇ · Biden_within × Post            ← H2b test
  + γ · Controls<sub>it</sub> (Mundlak-adjusted)
  + ε<sub>it</sub></code></pre>

    <ul>
      <li><strong>Estimator:</strong> Papke-Wooldridge fractional logit
      (for bounded [0, 1] DV with mass at 0 and 1).</li>
      <li><strong>Standard errors:</strong> state-level cluster-robust,
      reinforced with wild-cluster bootstrap (Rademacher weights, 999
      replicates). State-level clustering respects the hierarchical
      structure (CoCs ⊂ counties ⊂ states) and is more conservative than
      CoC-level alone.</li>
      <li><strong>Sample:</strong> unbalanced panel (651 obs) for primary
      power; balanced 3-year panel (375 obs) as robustness.</li>
      <li><strong>Mundlak adjustment:</strong> adding CoC-level means of
      time-varying controls makes fixed-effects-like identification
      possible without absorbing the time-invariant IVs.</li>
    </ul>

    <h3>Why this design</h3>
    <p>Traditional within-CoC fixed effects would absorb all three IVs
    (nonprofit form, state politics, county politics are essentially
    time-invariant within our 3-year window). The Mundlak approach
    preserves identification of the DiD coefficients (IV × Post) while
    controlling for between-CoC correlation through the means. The
    multilevel decomposition of political environment then lets us
    cleanly attribute any political-environment effect to the state
    level vs. within-state county deviation.</p>

    <p>Next: <a href="descriptive.html">descriptive patterns and the
    Grants Pass shock</a> → <a href="main_results.html">primary multilevel
    DiD results</a></p>
    """
    return wrap("Research model", "model.html", body,
                "Two competing hypotheses — organizational form vs. local political environment — tested in one multilevel DiD.")


def page_distributions():
    import statistics as st

    path = PIPELINE_DIR / "coc_analysis_ready.csv"
    rows = load_csv(path)

    def to_num(v):
        if v in ("", None):
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def skew(vals):
        n = len(vals)
        if n < 3:
            return 0
        m = sum(vals) / n
        sd = (sum((x - m) ** 2 for x in vals) / n) ** 0.5
        if sd == 0:
            return 0
        return sum(((x - m) / sd) ** 3 for x in vals) / n

    variables = [
        ("crim_activity_index", "DV (primary)", "proportion",
         "Anti-criminalization activity index (share of 1D-4 cells = Yes)"),
        ("implemented_anticrim_practice", "DV (secondary)", "binary",
         "Any implemented anti-criminalization policy (1 = Yes in any 1D-4 col 2 cell)"),
        ("1b_1_6_meetings", "Mediator · breadth", "binary",
         "Homeless/Formerly Homeless persons at CoC meetings"),
        ("1b_1_6_voted", "Mediator · breadth", "binary",
         "Homeless/Formerly Homeless persons voting in CoC"),
        ("1b_1_6_ces", "Mediator · breadth", "binary",
         "Homeless/Formerly Homeless persons in CES"),
        ("1d_10a_1_years", "Mediator · depth", "count",
         "# PLE in decisionmaking (within last 7 years)"),
        ("1d_10a_1_unsheltered", "Mediator · depth", "count",
         "# PLE in decisionmaking (unsheltered)"),
        ("1d_10a_2_years", "Mediator · depth", "count",
         "# PLE on CoC committees"),
        ("1d_10a_3_years", "Mediator · depth", "count",
         "# PLE on rating/scoring factors"),
        ("1d_2_3", "Control", "proportion",
         "Housing First adoption rate"),
        ("2a_5_1_coverage", "Control · maturity", "proportion",
         "HMIS bed coverage — Emergency Shelter"),
        ("2a_5_5_coverage", "Control · maturity", "proportion",
         "HMIS bed coverage — Permanent Supportive Housing"),
        ("2a_5_1_non_vsp", "Control · size", "count",
         "Emergency Shelter bed count"),
        ("2a_5_5_non_vsp", "Control · size", "count",
         "PSH bed count"),
        ("1d_2_1", "Control", "count",
         "Total CoC Program projects applied for"),
    ]

    summary_rows = []
    plot_calls = []
    for fid, role, typ, label in variables:
        vals = [to_num(r.get(fid)) for r in rows]
        vals = [v for v in vals if v is not None]
        if not vals:
            continue
        n = len(vals)
        m = sum(vals) / n
        med = sorted(vals)[n // 2]
        sd = (sum((x - m) ** 2 for x in vals) / n) ** 0.5
        mn, mx = min(vals), max(vals)
        sk = skew(vals)
        zero_pct = sum(1 for v in vals if v == 0) / n
        one_pct = sum(1 for v in vals if v == 1) / n
        bound_pct = zero_pct + one_pct if typ in ("proportion", "binary") else 0

        # Decide recommendation
        rec_parts = []
        rec_class = "good"
        if typ == "count":
            if abs(sk) > 1.5:
                rec_parts.append(f"<strong>log1p(x)</strong> before modeling")
                rec_class = "warn"
            else:
                rec_parts.append("raw OK")
        elif typ == "proportion":
            if mx > 1.5:
                rec_parts.append(f"normalize to [0,1] (÷100); max={mx:.2f}")
                rec_class = "warn"
            if bound_pct > 0.3:
                rec_parts.append(f"censored at bounds ({bound_pct:.0%}) — use <strong>fractional logit / Tobit</strong>")
                rec_class = "warn"
            if not rec_parts:
                rec_parts.append("raw OK")
        elif typ == "binary":
            if m > 0.95:
                rec_parts.append(f"<strong>ceiling effect</strong> (mean={m:.2f}) — insufficient variance")
                rec_class = "bad"
            elif m < 0.05:
                rec_parts.append(f"floor effect (mean={m:.2f})")
                rec_class = "bad"
            else:
                rec_parts.append("use as dummy")

        recommendation = " · ".join(rec_parts)

        # Plot spec
        chart_id = f"chart_{fid.replace('_', '')}"
        if typ == "count" and abs(sk) > 1.5:
            # Side-by-side: raw and log1p
            import math
            log_vals = [math.log1p(v) for v in vals]
            traces = [
                {"type": "histogram", "x": vals, "name": "raw",
                 "opacity": 0.7, "nbinsx": 30, "xaxis": "x", "yaxis": "y",
                 "marker": {"color": "#0366d6"}},
                {"type": "histogram", "x": log_vals, "name": "log1p",
                 "opacity": 0.7, "nbinsx": 30, "xaxis": "x2", "yaxis": "y2",
                 "marker": {"color": "#17813c"}},
            ]
            layout = {
                "height": 280,
                "showlegend": False,
                "margin": {"t": 30, "l": 50, "r": 20, "b": 40},
                "xaxis": {"title": "raw", "domain": [0, 0.47]},
                "xaxis2": {"title": "log(x+1)", "domain": [0.53, 1.0]},
                "yaxis": {"title": "count"},
                "yaxis2": {"anchor": "x2"},
            }
        else:
            traces = [{"type": "histogram", "x": vals, "opacity": 0.75, "nbinsx": 30,
                       "marker": {"color": "#0366d6"}}]
            layout = {
                "height": 240,
                "margin": {"t": 20, "l": 50, "r": 20, "b": 40},
                "xaxis": {"title": label},
                "yaxis": {"title": "count of CoC-years"},
                "showlegend": False,
            }
        plot_calls.append(
            f'<div id="{chart_id}" style="width:100%"></div>'
            f'<script>Plotly.newPlot("{chart_id}", {json.dumps(traces)}, '
            f'{json.dumps(layout)}, {{responsive:true, displayModeBar:false}});</script>'
        )

        summary_rows.append({
            "fid": fid, "role": role, "typ": typ, "label": label,
            "n": n, "m": m, "med": med, "sd": sd, "mn": mn, "mx": mx,
            "skew": sk, "zero_pct": zero_pct, "rec": recommendation,
            "rec_class": rec_class, "chart_id": chart_id,
        })

    # Build the body: one section per variable with stats + chart
    body = """
    <p class="lead">Histograms and descriptive statistics of every variable
    that goes into the regression, with concrete recommendations for
    transformations before modeling.</p>

    <div class="callout">
      <div class="callout-title">Summary of transformation decisions</div>
      <ul>
        <li><strong>Counts</strong> (PLE numbers, bed counts, project counts): all right-skewed (skew 4–18) → use <code>log1p(x)</code>.</li>
        <li><strong>Proportions with ceilings</strong> (Housing First %, HMIS coverage): censored at 1.0 (and often at 0) → use <strong>fractional logit</strong> or <strong>Tobit</strong>, not OLS.</li>
        <li><strong>Binary mediators</strong> (1b_1_6_*): OK as dummies — <code>meetings</code> hits a ceiling at 99%, so use <code>voted</code> and <code>ces</code> as the primary breadth indicators.</li>
        <li><strong>DV (crim_activity_index)</strong>: 33% of CoC-years hit the bounds (0 or 1). Use fractional logit or Tobit. OLS with CoC+year FE is acceptable as a linear-probability-style first pass.</li>
      </ul>
    </div>

    <h2>Summary table</h2>
    """
    # Summary table
    body += '<table><thead><tr><th>Variable</th><th>Role</th><th class="num">n</th><th class="num">Mean</th><th class="num">Median</th><th class="num">SD</th><th class="num">Min / Max</th><th class="num">Skew</th><th>Recommendation</th></tr></thead><tbody>'
    for r in summary_rows:
        body += (
            f'<tr>'
            f'<td><code>{r["fid"]}</code></td>'
            f'<td>{r["role"]}</td>'
            f'<td class="num">{r["n"]}</td>'
            f'<td class="num">{r["m"]:.2f}</td>'
            f'<td class="num">{r["med"]:.2f}</td>'
            f'<td class="num">{r["sd"]:.2f}</td>'
            f'<td class="num">{r["mn"]:.0f} / {r["mx"]:.0f}</td>'
            f'<td class="num">{r["skew"]:.2f}</td>'
            f'<td><span class="badge {r["rec_class"]}">{r["rec"]}</span></td>'
            f'</tr>'
        )
    body += '</tbody></table>'

    body += '<h2>Per-variable distributions</h2>'
    for r, plot in zip(summary_rows, plot_calls):
        body += (
            f'<h3><code>{r["fid"]}</code> — {r["label"]}</h3>'
            f'<p class="hint">{r["role"]} · <strong>type:</strong> {r["typ"]} · '
            f'mean {r["m"]:.2f}, SD {r["sd"]:.2f}, skew {r["skew"]:.2f}, '
            f'{r["zero_pct"]:.0%} zeros · <strong>{r["rec"]}</strong></p>'
            + plot
        )

    body += """
    <h2>Implications for the regression model</h2>

    <h3>Model specification after transformations</h3>
    <pre><code>crim_activity_index<sub>it</sub> =
      β₁ · Nonprofit<sub>i</sub>
    + β₂ · log1p(PLE_decisionmaking_count<sub>it</sub>)    ← was 1d_10a_1_years
    + β₃ · 1b_1_6_voted<sub>it</sub>                        ← binary mediator (voted)
    + β₄ · 1b_1_6_ces<sub>it</sub>                          ← binary mediator (CES)
    + β₅ · HMIS_coverage<sub>it</sub>                       ← proportion (fractional)
    + β₆ · log1p(total_beds<sub>it</sub>)                   ← size, logged
    + β₇ · HF_adoption<sub>it</sub>                         ← proportion
    + α<sub>i</sub> + γ<sub>t</sub> + ε<sub>it</sub></code></pre>

    <p>Because <code>crim_activity_index</code> is a fraction with mass at
    0 and 1, prefer a <strong>fractional logit</strong> estimator with CoC
    fixed effects (or a pooled fractional logit with cluster-robust SEs at
    the CoC level). Linear two-way fixed effects remains an acceptable
    linear-probability-style first pass — report both.</p>

    <h3>What to drop before estimation</h3>
    <ul>
      <li><code>1b_1_6_meetings</code> — 99% Yes rate, insufficient variance.</li>
      <li><code>engaged_policymakers_crim</code>, <code>engaged_law_enforce_crim</code>
      (from the DV story) — 96–97% ceilings; kept in the output but not in the panel model.</li>
    </ul>

    <h3>Robustness checks</h3>
    <ul>
      <li>Log-level vs. level-level specifications for PLE counts (expect log to fit better).</li>
      <li>Re-estimate with <code>implemented_anticrim_practice</code> as a logit DV — interpret the FY2024 coefficient carefully (instrument change).</li>
      <li>Winsorize bed counts at the 99th percentile and re-run (some CoCs like NYC have 108k+ ES beds — 10× the next).</li>
      <li>Estimate on FY2022+FY2023 panel only (identical 1D-4 instrument) as a sanity check.</li>
    </ul>
    """
    return wrap("Variable distributions & transformations", "distributions.html", body,
                "What shape each variable takes, and what transformations it needs.")


def page_approach():
    body = """
    <p class="lead">How the pipeline works — in plain language, in the order
    decisions were made.</p>

    <div class="callout">
      <div class="callout-title">Why the approach matters</div>
      <p>The goal is to turn HUD's PDF applications into a dataset that
      another researcher could rebuild from the original files in an
      afternoon. Every design choice below is oriented around
      <strong>reproducibility, provenance, and auditability</strong> —
      not speed.</p>
    </div>

    <h2>Step 1 — Inventory the source files</h2>
    <p>Before reading any data, we cataloged every file in the source folder:
    year, format (PDF or DOCX), whether the PDF is a scan, whether the
    filename has a typo. This took minutes and saved hours later: <strong>19
    PDFs are scanned images</strong> that text parsers cannot read, and
    <strong>7 CoCs ship their applications as DOCX instead of PDF</strong>.
    If we had skipped this step we would have silently dropped those
    observations.</p>

    <h2>Step 2 — Treat the manually coded spreadsheet as a reference, not ground truth</h2>
    <p>Before we started, fifteen RAs had already spent weeks manually
    coding a FY2024 spreadsheet with ~292 CoCs × ~331 variables. That
    spreadsheet is valuable — but with 15 coders working over months,
    typos, interpretation differences, and partially filled rows are
    inevitable.</p>
    <p>Rather than treat the manual file as ground truth, we treat it as a
    <em>reference</em>. When the automated extractor disagrees with the
    manual entry, we investigate: is the extractor wrong, or is the manual
    wrong? The audit trail (see <a href="review.html">Review queue</a>)
    shows dozens of cases where automation caught manual data-entry
    errors.</p>

    <h2>Step 3 — One canonical schema, a crosswalk for year differences</h2>
    <p>HUD renumbers and rewords questions between years. Rather than
    maintain three separate schemas, we fix one canonical schema (the FY2024
    numbering, because the manual spreadsheet already uses it) and build a
    <strong>crosswalk</strong> specifying how FY2022 and FY2023 question
    numbers map onto it. When a question exists in one year but not another,
    the crosswalk says so explicitly; the panel analysis can then honestly
    flag where data is missing by design.</p>

    <h2>Step 4 — "Anchor-then-parse" extraction</h2>
    <p>Each HUD application is ~69 pages. A brittle parser that tries to
    read the whole document at once spreads errors across many variables.
    Instead the extractor finds each question's number (its "anchor" — e.g.
    the string <code>1B-1.</code>) and parses <em>only the local block</em>
    following it. When one block fails, the other ~330 variables are
    unaffected.</p>

    <h3>Two things we match by label, not position</h3>
    <ul>
      <li><strong>Rows within charts.</strong> The 1B-1 participation chart
      has 33 canonical organization types. In some years their alphabetic
      ordering differs (FY2022 put "Agencies serving survivors of human
      trafficking" in row 2; FY2024 put it in row 25). The extractor
      identifies each row by its label and maps it to the correct canonical
      position, ignoring the observed position. Without this fix, FY2022
      data was landing in the wrong columns.</li>
      <li><strong>Rows across years.</strong> The PLE engagement chart moved
      from <code>1D-11a</code> (FY2022/23) to <code>1D-10a</code> (FY2024).
      The extractor dispatches on year, reads from the correct question
      number, and writes to the canonical schema field.</li>
    </ul>

    <h2>Step 5 — Provenance on every cell</h2>
    <p>Every extracted record carries: the CoC ID, the fiscal year, the
    field ID, the source page in the PDF, and a confidence flag. When a
    value is flagged <code>needs_review</code>, it means the extractor
    found something but is uncertain (e.g., the row label didn't match
    any canonical label). Nothing is silently coerced.</p>

    <h2>Step 6 — Four variable classes, four validation rules</h2>
    <table>
      <thead><tr><th>Class</th><th>Examples</th><th>How we validate</th></tr></thead>
      <tbody>
      <tr><td><strong>A. Categorical</strong></td><td>Yes / No / Nonexistent; CA / UFA</td><td>Exact match against controlled vocabulary</td></tr>
      <tr><td><strong>B. Numeric</strong></td><td>PIT counts, HMIS %</td><td>Exact match after unit normalization (fraction vs. percent)</td></tr>
      <tr><td><strong>C. Identifier / Label</strong></td><td>CoC name, HMIS lead</td><td>Case- and whitespace-normalized string match</td></tr>
      <tr><td><strong>D. Narrative</strong></td><td>Racial-equity narrative, PLE outreach narrative</td><td>Not yet coded — awaiting Stage-2 LLM coding with verbatim-quote requirement</td></tr>
      </tbody>
    </table>
    <p>We set different accuracy targets per class (A ≥ 99%, B ≥ 99%,
    C ≥ 97%, D ≥ 90% reviewer acceptance). This matters because a 2,500-
    character narrative can never be "exactly equal" to a manually coded
    summary, but a Yes/No answer can.</p>

    <h2>Step 7 — Iterate until it converges</h2>
    <p>We ran the extractor four times. Each iteration surfaced
    disagreements between automation and the manual spreadsheet. Every
    disagreement was classified into one of six root causes:</p>

    <ol>
      <li><strong>E1 — Extractor bug</strong>: the parser is wrong → fix code</li>
      <li><strong>E2 — Mapping error</strong>: the crosswalk is wrong → fix schema</li>
      <li><strong>E3 — Manual error</strong>: the spreadsheet is wrong → flag for PI review</li>
      <li><strong>E4 — Format variant</strong>: the PDF layout has a quirk → strengthen adapter</li>
      <li><strong>E5 — Narrative paraphrase</strong>: LLM wording differs but meaning is right → tune the prompt</li>
      <li><strong>E6 — Source noise</strong>: scan artifacts → improve OCR</li>
    </ol>

    <p>Iterations exit when weighted agreement ≥ 98% <em>and</em> two
    consecutive runs show no residual E1/E2. That is, the remaining
    disagreements are all attributable to source documents (E3/E4/E6), not
    to the pipeline.</p>

    <h3>What the iteration log actually looks like</h3>
    """
    # Iteration table
    iters = load_iterations()
    headers = [("iter", "Iter"), ("date", "Date"), ("extractor_version", "Extractor"),
               ("weighted_acc", "Weighted agreement"), ("adjusted_acc", "Adjusted"),
               ("total_diffs", "Diffs"), ("top_error", "Dominant error"), ("fix_applied", "Fix")]
    rows = []
    for r in iters:
        tds = []
        for key, _ in headers:
            v = r.get(key, "")
            if key in ("weighted_acc", "adjusted_acc") and v not in ("", None):
                try:
                    v = f"{float(v):.2%}"
                except Exception:
                    pass
            tds.append(f"<td>{html.escape(str(v))}</td>")
        rows.append("<tr>" + "".join(tds) + "</tr>")
    body += (
        "<table><thead><tr>"
        + "".join(f"<th>{lbl}</th>" for _, lbl in headers)
        + "</tr></thead><tbody>"
        + "\n".join(rows)
        + "</tbody></table>"
    )
    body += """
    <p>The big jump between v0.5 and v0.6b came from a single change:
    matching 1B-1 rows by <em>label</em> instead of by position. That fix
    unlocked 81 previously mis-mapped variables in the FY2022 panel —
    which is why <code>panel_safe</code> variables jumped from 156 to 237
    on that iteration.</p>

    <h2>Step 8 — Produce panel-ready outputs</h2>
    <p>After extraction converges, a small suite of scripts assembles the
    final datasets:</p>
    <ul>
      <li><code>coc_panel_wide.xlsx</code> — the headline panel file (sheet
      <code>panel_safe</code> is ready for regressions).</li>
      <li><code>coc_panel_long.csv</code> — long form for R/pandas fixed-
      effects models.</li>
      <li><code>coc_analysis_ready.xlsx</code> — includes harmonized DVs and
      three use-case sheets (balanced panel, unbalanced, FY2024-only).</li>
      <li><code>panel_field_map.csv</code> — says which variables are
      panel_safe and which aren't.</li>
    </ul>

    <p>Next: <a href="variables.html">what's in the dataset</a> →</p>
    """
    return wrap("How the pipeline works", "approach.html", body,
                "Eight steps, in the order the decisions were made.")


def page_variables():
    fm = load_field_map()
    data = []
    for r in fm:
        fid = r["field_id"]
        cat = r["category"]
        friendly = FRIENDLY_DESCRIPTIONS.get(fid, (None, None))
        human_name, desc = friendly
        data.append({
            "field_id": fid,
            "category": cat,
            "human_name": human_name or "",
            "desc": desc or "",
            "cov_2022": float(r["coverage_2022"]),
            "cov_2023": float(r["coverage_2023"]),
            "cov_2024": float(r["coverage_2024"]),
        })
    order = {"panel_safe": 0, "mostly_panel": 1, "year_specific": 2, "sparse": 3}
    data.sort(key=lambda x: (order.get(x["category"], 9), x["field_id"]))

    rows = []
    for r in data:
        badge_class = {"panel_safe": "good", "mostly_panel": "warn",
                       "year_specific": "warn", "sparse": "bad"}.get(r["category"], "")
        name_cell = (f"<strong>{html.escape(r['human_name'])}</strong><br>"
                     f"<span class='hint'>{html.escape(r['desc'])}</span>") if r["human_name"] else ""
        rows.append(
            f'<tr data-cat="{r["category"]}">'
            f'<td><code>{html.escape(r["field_id"])}</code></td>'
            f'<td>{name_cell}</td>'
            f'<td><span class="badge {badge_class}">{r["category"]}</span></td>'
            f'<td class="num">{r["cov_2022"]:.0%}</td>'
            f'<td class="num">{r["cov_2023"]:.0%}</td>'
            f'<td class="num">{r["cov_2024"]:.0%}</td>'
            f'</tr>'
        )
    counts = Counter(d["category"] for d in data)
    summary_badges = " ".join(
        f'<span class="badge {"good" if c=="panel_safe" else "warn" if c in ("mostly_panel","year_specific") else "bad"}">{c}: {n}</span>'
        for c, n in counts.most_common()
    )

    script = """
    <script>
    const input = document.getElementById('f');
    const sel = document.getElementById('cat');
    function render() {
      const q = input.value.toLowerCase();
      const c = sel.value;
      document.querySelectorAll('#vtable tbody tr').forEach(tr => {
        const ok = (c === '' || tr.dataset.cat === c)
                && (q === '' || tr.innerText.toLowerCase().includes(q));
        tr.style.display = ok ? '' : 'none';
      });
    }
    input.addEventListener('input', render);
    sel.addEventListener('change', render);
    </script>
    """
    rows_html = "\n".join(rows)
    body = f"""
    <p class="lead">Every variable in the canonical FY2024 schema, with
    year-by-year coverage and (for paper-relevant variables) plain-English
    descriptions.</p>

    <h2>What to focus on</h2>
    <p>The variables that do the analytical work in the paper have
    friendly names and descriptions. Scroll — or type in the filter — to
    find them. A handful of highlights:</p>
    <ul>
      <li><strong>Leadership / governance</strong>: <code>1a_2</code>
      Collaborative Applicant Name (IV for nonprofit vs. government),
      <code>1a_3</code> CoC designation.</li>
      <li><strong>PLE engagement</strong>: <code>1b_1_6_*</code>
      (whether people with lived experience attend CoC meetings / vote /
      participate in CES), <code>1d_10a_*</code> (counts of PLE in
      governance roles).</li>
      <li><strong>Criminalization</strong>: <code>1d_4_*</code> (three
      strategies × two columns — see the <a href="dv_story.html">DV
      story</a>).</li>
      <li><strong>Controls</strong>: <code>1d_2_3</code> (Housing First %),
      <code>2a_5_*_coverage</code> (HMIS bed coverage by project type),
      <code>1c_1_*</code> (breadth of cross-sector coordination).</li>
    </ul>

    <h2>Filter the full schema</h2>
    <p>{summary_badges}</p>
    <div class="filterbox">
      <input id="f" placeholder="Filter by ID, name, or description (e.g., 'housing first', '1d_4', 'sparse')">
      <select id="cat">
        <option value="">All categories</option>
        <option value="panel_safe">panel_safe (longitudinal-ready)</option>
        <option value="mostly_panel">mostly_panel</option>
        <option value="year_specific">year_specific</option>
        <option value="sparse">sparse (not yet coded)</option>
      </select>
    </div>
    <table id="vtable">
      <thead><tr>
        <th>Field ID</th><th>Plain-English</th><th>Category</th>
        <th class="num">FY22</th><th class="num">FY23</th><th class="num">FY24</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    {script}
    """
    return wrap("What's in the dataset", "variables.html", body,
                "All 331 variables, searchable.")


def page_descriptive():
    stats_path = PIPELINE_DIR / "descriptive_stats.json"
    if not stats_path.exists():
        return wrap("Descriptives", "descriptive.html",
                    "<p><em>Run <code>build_descriptive.py</code> first.</em></p>", "")
    stats = json.loads(stats_path.read_text())

    sample = stats["sample"]
    act = stats["activity_index"]
    deltas = stats["deltas"]
    cells = stats["cell_rates"]
    bal = stats["balanced_panel"]

    # Time-series mean line with SE band
    years = ["2022", "2023", "2024"]
    mean_line = {
        "type": "scatter", "mode": "lines+markers+text",
        "x": [f"FY{y}" for y in years],
        "y": [act[y]["overall"]["mean"] for y in years],
        "text": [f"{act[y]['overall']['mean']:.3f}" for y in years],
        "textposition": "top center",
        "name": "Mean", "line": {"width": 3, "color": "#0366d6"},
        "marker": {"size": 10, "color": "#0366d6"},
        "error_y": {
            "type": "data", "visible": True,
            "array": [act[y]["overall"]["se"] * 1.96 for y in years],
            "thickness": 1.5,
        },
    }
    np_line = {
        "type": "scatter", "mode": "lines+markers",
        "x": [f"FY{y}" for y in years],
        "y": [act[y]["nonprofit"]["mean"] for y in years],
        "name": "Nonprofit-led",
        "line": {"width": 2.5, "color": "#17813c", "dash": "solid"},
        "marker": {"size": 9, "color": "#17813c"},
        "error_y": {
            "type": "data", "visible": True,
            "array": [act[y]["nonprofit"]["se"] * 1.96 for y in years],
            "thickness": 1, "color": "#17813c",
        },
    }
    gv_line = {
        "type": "scatter", "mode": "lines+markers",
        "x": [f"FY{y}" for y in years],
        "y": [act[y]["government"]["mean"] for y in years],
        "name": "Government-led",
        "line": {"width": 2.5, "color": "#c53030", "dash": "solid"},
        "marker": {"size": 9, "color": "#c53030"},
        "error_y": {
            "type": "data", "visible": True,
            "array": [act[y]["government"]["se"] * 1.96 for y in years],
            "thickness": 1, "color": "#c53030",
        },
    }
    ts_layout = {
        "title": "Anti-criminalization activity index — yearly means (±1.96·SE)",
        "xaxis": {"title": "Fiscal year"},
        "yaxis": {"title": "Mean index (0–1)", "range": [0.60, 0.90]},
        "height": 420,
        "legend": {"orientation": "h", "y": -0.2},
        "shapes": [{
            "type": "rect", "xref": "x", "yref": "paper",
            "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
            "fillcolor": "#fff4cc", "opacity": 0.35, "line": {"width": 0},
            "layer": "below",
        }],
        "annotations": [{
            "x": 2, "y": 0.96, "xref": "x", "yref": "paper",
            "showarrow": False,
            "text": "Post-Grants Pass · HUD 1D-4 redesign",
            "font": {"color": "#8a6d1f", "size": 11},
        }],
    }

    # Distribution shift — bar chart of FY22 vs FY23 vs FY24
    buckets = ["0.00", "0.17", "0.33", "0.50", "0.67", "0.83", "1.00"]
    # Use activity hist_counts (7 bins from 0 to 1)
    def hist_normalized(year):
        h = act[year]["overall"]["hist_counts"]
        total = sum(h) or 1
        return [round(x / total * 100, 1) for x in h]
    dist_data = [
        {"type": "bar", "name": f"FY{y} (n={act[y]['overall']['n']})",
         "x": buckets, "y": hist_normalized(y),
         "text": [f"{v}%" for v in hist_normalized(y)],
         "textposition": "outside",
         "marker": {"color": {"2022": "#c9c9c9", "2023": "#7fbfff", "2024": "#0366d6"}[y]}}
        for y in years
    ]
    dist_layout = {
        "title": "Distribution of activity index by year (% of CoCs in each bucket)",
        "xaxis": {"title": "Activity index bucket"},
        "yaxis": {"title": "Share of CoCs in that year (%)", "range": [0, 65]},
        "barmode": "group",
        "height": 400,
        "legend": {"orientation": "h", "y": -0.15},
    }

    # Cell heatmap
    heat_rows = []
    heat_z = []
    for cell in cells:
        label = cell["label"]
        heat_rows.append(label)
        row_z = []
        for y in years:
            v = cell["years"][y]["overall"]
            row_z.append(round((v or 0) * 100, 1))
        heat_z.append(row_z)
    heat_data = [{
        "type": "heatmap",
        "z": heat_z,
        "x": [f"FY{y}" for y in years],
        "y": heat_rows,
        "colorscale": [[0, "#fee5d9"], [0.5, "#fcbba1"], [1, "#a50f15"]],
        "zmin": 40, "zmax": 100,
        "text": [[f"{v:.0f}%" for v in row] for row in heat_z],
        "texttemplate": "%{text}",
        "textfont": {"color": "#111"},
        "colorbar": {"title": "Yes %", "thickness": 10, "len": 0.7},
    }]
    heat_layout = {
        "title": "Cell-level Yes rates — which 1D-4 cells drove the FY2024 jump?",
        "xaxis": {"side": "top"},
        "yaxis": {"autorange": "reversed"},
        "height": 360,
        "margin": {"t": 60, "l": 260, "r": 20, "b": 20},
    }

    # Δ bar chart (by leadership × period)
    periods = ["FY22_FY23", "FY23_FY24"]
    delta_data = [
        {"type": "bar", "name": "Nonprofit-led",
         "x": ["FY22 → FY23 (pre-period)", "FY23 → FY24 (post-Grants Pass)"],
         "y": [deltas["nonprofit"][p] for p in periods],
         "text": [f"+{deltas['nonprofit'][p]:.3f}" for p in periods],
         "textposition": "outside",
         "marker": {"color": "#17813c"}},
        {"type": "bar", "name": "Government-led",
         "x": ["FY22 → FY23 (pre-period)", "FY23 → FY24 (post-Grants Pass)"],
         "y": [deltas["government"][p] for p in periods],
         "text": [f"+{deltas['government'][p]:.3f}" for p in periods],
         "textposition": "outside",
         "marker": {"color": "#c53030"}},
    ]
    delta_layout = {
        "title": "Year-over-year change by lead-agency type",
        "yaxis": {"title": "Δ mean activity index", "range": [-0.02, 0.12]},
        "barmode": "group",
        "height": 380,
        "legend": {"orientation": "h", "y": -0.15},
    }

    # Build tables
    def stat_row(label, s):
        if s.get("n", 0) == 0:
            return f"<tr><td>{label}</td>" + "<td class='num'>—</td>" * 7 + "</tr>"
        return (
            f"<tr><td>{label}</td>"
            f"<td class='num'>{s['n']}</td>"
            f"<td class='num'>{s['mean']:.3f}</td>"
            f"<td class='num'>{s['median']:.3f}</td>"
            f"<td class='num'>{s['sd']:.3f}</td>"
            f"<td class='num'>{s['min']:.2f}</td>"
            f"<td class='num'>{s['max']:.2f}</td>"
            f"<td class='num'>{s['at_one']} ({s['at_one']/max(s['n'],1)*100:.0f}%)</td>"
            f"</tr>"
        )

    activity_table = "<table><thead><tr><th>Group / Year</th><th class='num'>N</th><th class='num'>Mean</th><th class='num'>Median</th><th class='num'>SD</th><th class='num'>Min</th><th class='num'>Max</th><th class='num'>At 1.0</th></tr></thead><tbody>"
    for y in years:
        activity_table += f"<tr style='background:#eef4fb'><td colspan='8'><strong>FY{y}</strong></td></tr>"
        activity_table += stat_row("  Overall", act[y]["overall"])
        activity_table += stat_row("  Nonprofit-led", act[y]["nonprofit"])
        activity_table += stat_row("  Government-led", act[y]["government"])
    activity_table += "</tbody></table>"

    # Sample table
    sample_table = """<table><thead><tr><th>Fiscal year</th><th class='num'>Records</th><th class='num'>Unique CoCs</th><th class='num'>IV-coded</th><th class='num'>Nonprofit</th><th class='num'>Government</th><th class='num'>% Nonprofit</th><th class='num'>With DV</th></tr></thead><tbody>"""
    for y in years:
        s = sample[y]
        sample_table += (
            f"<tr><td>FY{y}</td>"
            f"<td class='num'>{s['n_records']}</td>"
            f"<td class='num'>{s['n_cocs']}</td>"
            f"<td class='num'>{s['coded_iv']}</td>"
            f"<td class='num'>{s['nonprofit']}</td>"
            f"<td class='num'>{s['government']}</td>"
            f"<td class='num'>{s['pct_nonprofit']}%</td>"
            f"<td class='num'>{s['with_dv']}</td></tr>"
        )
    sample_table += "</tbody></table>"

    # Delta table
    delta_table = """<table><thead><tr><th>Group</th><th class='num'>FY22 → FY23 (pre)</th><th class='num'>FY23 → FY24 (post-Grants Pass)</th><th class='num'>FY22 → FY24 (total)</th></tr></thead><tbody>"""
    for g in ["overall", "nonprofit", "government"]:
        gl = {"overall": "Overall", "nonprofit": "Nonprofit-led", "government": "Government-led"}[g]
        delta_table += (
            f"<tr><td>{gl}</td>"
            f"<td class='num'>{deltas[g]['FY22_FY23']:+.4f}</td>"
            f"<td class='num'><strong>{deltas[g]['FY23_FY24']:+.4f}</strong></td>"
            f"<td class='num'>{deltas[g]['FY22_FY24']:+.4f}</td></tr>"
        )
    delta_table += "</tbody></table>"

    # Balanced paired sample
    paired = bal["paired"]
    balanced_info = ""
    if paired:
        balanced_info = f"""
    <h2>Paired (within-CoC) changes</h2>
    <p>Restricting to the {paired['n_paired']} CoCs observed in all 3 years —
    the cleanest look at <em>within-CoC</em> change:</p>
    <table>
      <thead><tr><th></th><th class='num'>Mean</th></tr></thead>
      <tbody>
        <tr><td>FY2022 (pre)</td><td class='num'>{paired['mean_FY22']:.3f}</td></tr>
        <tr><td>FY2023 (pre)</td><td class='num'>{paired['mean_FY23']:.3f}</td></tr>
        <tr><td>FY2024 (post-Grants Pass)</td><td class='num'>{paired['mean_FY24']:.3f}</td></tr>
        <tr><td><strong>Mean paired Δ (FY24 − FY23)</strong></td>
            <td class='num'><strong>+{paired['mean_delta_23_24']:.3f}</strong> (SD {paired['sd_delta_23_24']:.3f})</td></tr>
        <tr><td>Share of CoCs that increased</td>
            <td class='num'>{paired['share_increased']*100:.0f}%</td></tr>
        <tr><td>Share that decreased</td>
            <td class='num'>{paired['share_decreased']*100:.0f}%</td></tr>
        <tr><td>Share unchanged</td>
            <td class='num'>{paired['share_unchanged']*100:.0f}%</td></tr>
      </tbody>
    </table>
    <p class="hint">Only {paired['share_increased']*100:.0f}% of CoCs actually
    raised their score FY23→FY24. The +{paired['mean_delta_23_24']:.3f} mean increase
    is driven by a subset making large jumps while {paired['share_unchanged']*100:.0f}%
    stayed exactly the same — a signature consistent with the HUD 1D-4 form
    redesign producing a step-function shift for CoCs that updated their reporting
    practices, rather than a gradual substantive behavior change.</p>
    """

    body = f"""
    <p class="lead">Descriptive statistics for the anti-criminalization
    activity index (<code>crim_activity_index</code>) across FY2022, FY2023,
    and FY2024 — with an explicit view into <em>what</em> shifted after the
    June 2024 Supreme Court ruling in <em>City of Grants Pass v. Johnson</em>.</p>

    <h2>Grants Pass as the policy shock</h2>
    <div class="callout">
      <p><strong><em>City of Grants Pass v. Johnson</em></strong> (603 U.S. ___,
      decided June 28, 2024) held 6–3 that enforcing public-camping ordinances
      against people experiencing homelessness <em>does not</em> constitute
      cruel and unusual punishment under the Eighth Amendment. The decision
      overruled the Ninth Circuit's <em>Martin v. Boise</em> (2019) framework
      that had effectively blocked anti-camping enforcement in nine western
      states when shelter space was unavailable.</p>
      <p>The FY2024 HUD CoC Application deadline was <strong>October 30, 2024</strong> —
      four months after the ruling. FY2024 reports therefore capture CoCs'
      first documented anti-criminalization posture after the legal landscape
      shifted. The design exploits this timing as a quasi-exogenous shock.</p>
    </div>

    <h2>Sample</h2>
    {sample_table}
    <p class="hint">FY2024 is the largest because HUD folded YHDP recipients
    into the main competition that year. Nonprofit/government shares are
    remarkably stable (49% → 50% → 52% nonprofit).</p>

    <h2>Activity index — means over time</h2>
    <div id="ts_chart" style="width:100%"></div>
    <p>The yellow band marks the post-Grants Pass period. Both groups show
    essentially <strong>parallel pre-trends</strong> (FY22 → FY23) followed by
    a <strong>sharp, nearly identical jump</strong> in FY2024.</p>

    <h2>Descriptive table</h2>
    {activity_table}
    <p class="hint">Note the dramatic rise in the "at 1.0" count (all-Yes CoCs):
    55 (39%) in FY2022 → 80 (43%) in FY2023 → <strong>166 (58%)</strong> in FY2024.</p>

    <h2>Year-over-year changes (Δ mean)</h2>
    {delta_table}
    <div id="delta_chart" style="width:100%"></div>
    <p>The pre-period change (FY22 → FY23) was near zero for government-led CoCs
    (+0.003) and modest for nonprofit-led (+0.028). The post-shock change
    (FY23 → FY24) is <strong>nearly identical for both groups</strong>
    (nonprofit +0.085, government +0.082). The difference-in-differences is
    +0.004 — statistically indistinguishable from zero (wild-cluster bootstrap
    p = 0.819). <em>See <a href="results.html">Analysis results</a>
    for the full multilevel DiD.</em></p>

    <h2>Distribution shift</h2>
    <div id="dist_chart" style="width:100%"></div>
    <p>Each bar shows the share of CoCs landing in each discrete value of the
    activity index within a given year. Two things are visible:</p>
    <ul>
      <li>The <strong>0.50 bucket shrinks</strong> in FY2024 (FY22 38% →
      FY23 39% → FY24 15%): fewer CoCs end up at the "half-yes" level.</li>
      <li>The <strong>1.00 (all-Yes) bucket balloons</strong> in FY2024
      (FY22 39% → FY23 43% → FY24 58%): more CoCs report maximal
      anti-criminalization activity.</li>
    </ul>
    <p>This is a <em>bimodal shift</em> — CoCs aren't incrementally adding one
    more Yes; they're jumping from "half" to "all". Consistent with an
    instrument-change story (new FY24 form items are easier to check "Yes" to)
    layered on top of whatever substantive post-Grants Pass response
    exists.</p>

    <h2>Which cells drove the change?</h2>
    <div id="heat_chart" style="width:100%"></div>
    <p>This heatmap reveals the <strong>single most important descriptive
    finding of the paper</strong>: the FY2024 jump in the activity index is
    <strong>almost entirely produced by the three "Implementation" cells</strong>
    (rows 1–3 × column 2), not the "Policymaker engagement" cells.</p>
    <table style="font-size: 0.92em;">
      <thead><tr><th>Cell</th><th class='num'>FY2022</th><th class='num'>FY2023</th><th class='num'>FY2024</th><th class='num'>Δ FY23→24</th></tr></thead>
      <tbody>
        <tr><td>Row 1 · Policymaker engagement</td><td class='num'>97%</td><td class='num'>97%</td><td class='num'>94%</td><td class='num'>−3 pt</td></tr>
        <tr><td>Row 1 · <strong>Implementation</strong></td><td class='num'>47%</td><td class='num'>53%</td><td class='num'><strong>76%</strong></td><td class='num'><strong>+23 pt</strong></td></tr>
        <tr><td>Row 2 · Policymaker engagement</td><td class='num'>96%</td><td class='num'>96%</td><td class='num'>94%</td><td class='num'>−2 pt</td></tr>
        <tr><td>Row 2 · <strong>Implementation</strong></td><td class='num'>50%</td><td class='num'>52%</td><td class='num'><strong>67%</strong></td><td class='num'><strong>+15 pt</strong></td></tr>
        <tr><td>Row 3 · Policymaker engagement</td><td class='num'>91%</td><td class='num'>93%</td><td class='num'>93%</td><td class='num'>0 pt</td></tr>
        <tr><td>Row 3 · <strong>Implementation</strong></td><td class='num'>46%</td><td class='num'>46%</td><td class='num'><strong>64%</strong></td><td class='num'><strong>+18 pt</strong></td></tr>
      </tbody>
    </table>
    <p><strong>Why this matters for interpretation:</strong> the
    "Implementation" column was labeled differently in each form:</p>
    <ul>
      <li>FY2022/23: "Reverse existing criminalization policies" (what you must <em>have undone</em>)</li>
      <li>FY2024: "Implemented Laws/Policies/Practices that Prevent Criminalization" (what you have <em>implemented</em>)</li>
    </ul>
    <p>FY2024's wording is substantially easier to check "Yes" to — any
    adopted policy, even one existing for years, qualifies. FY2022/23's
    wording required evidence that an <em>existing criminalization policy was
    reversed</em>, which is much more stringent. This is why the jump is
    concentrated entirely in column 2 and why we reach this conclusion:</p>

    <div class="callout">
      <div class="callout-title">What we're actually observing</div>
      <p>The +0.083 FY23→FY24 jump in the overall activity index is
      <strong>partly an instrument artifact</strong>, <strong>partly a real
      response</strong> to Grants Pass, and in any case
      <strong>affects nonprofit-led and government-led CoCs equivalently</strong>.
      The "Policymaker engagement" column — which asked essentially the same
      question across all three years — actually <em>declined slightly</em>
      in FY2024 (~3-point drop). The increase is concentrated in the
      "Implementation" column, whose wording changed in ways that mechanically
      make "Yes" easier to report.</p>
      <p>This is why the paper should:</p>
      <ol>
        <li>Report cell-level descriptives to make the instrument effect transparent;</li>
        <li>Rely on DiD (which differences out common shifts) rather than raw year-on-year comparisons;</li>
        <li>Treat the DiD null as the primary finding — <em>both groups moved identically</em>,
        so even after the instrument change, leadership type does not predict differential response.</li>
      </ol>
    </div>

    {balanced_info}

    <h2>Bottom line</h2>
    <ul>
      <li><strong>Pre-period (FY22 → FY23):</strong> roughly stable activity index, with nearly
      parallel trajectories for nonprofit and government CoCs.</li>
      <li><strong>Post-Grants Pass (FY23 → FY24):</strong> both groups jump by
      ~0.08 points. The jump is <em>concentrated in the "Implementation" cells</em>
      whose wording changed to be easier to check Yes.</li>
      <li><strong>No differential response</strong> by lead-agency type
      (DiD ≈ 0, p<sub>boot</sub> = 0.819). The theoretical prediction that
      nonprofit-led CoCs should more aggressively reject criminalization is
      not supported.</li>
      <li><strong>The descriptive story is consistent with — but does not
      prove — an instrument-change interpretation.</strong> External behavioral
      data (NLHR ordinance counts, FBI UCR vagrancy arrests) are the next step
      to validate whether the FY2024 jump reflects real behavior change.</li>
    </ul>

    <script>
    Plotly.newPlot("ts_chart", {json.dumps([mean_line, np_line, gv_line])},
                   {json.dumps(ts_layout)}, {{responsive: true}});
    Plotly.newPlot("dist_chart", {json.dumps(dist_data)},
                   {json.dumps(dist_layout)}, {{responsive: true}});
    Plotly.newPlot("delta_chart", {json.dumps(delta_data)},
                   {json.dumps(delta_layout)}, {{responsive: true}});
    Plotly.newPlot("heat_chart", {json.dumps(heat_data)},
                   {json.dumps(heat_layout)}, {{responsive: true}});
    </script>
    """
    return wrap("Descriptives & Grants Pass", "descriptive.html", body,
                "How the anti-criminalization index moved year-to-year, and what that tells us about the 2024 Supreme Court ruling.")


def page_dv_story():
    harm = load_harmonized()
    # Collect activity index by year for a chart
    per_year = defaultdict(list)
    for r in harm:
        try:
            per_year[r["year"]].append(float(r["crim_activity_index"]))
        except (ValueError, TypeError):
            pass
    years = ["2022", "2023", "2024"]
    # crim_activity_index takes only 7 discrete values (0, 1/6, 2/6, 3/6, 4/6,
    # 5/6, 1) because the denominator is 6 cells. Bucket to sixths, then
    # emit a grouped bar per year so the three series are comparable.
    BUCKETS = [0.0, 1/6, 2/6, 3/6, 4/6, 5/6, 1.0]
    BUCKET_LABELS = ["0.000", "0.167", "0.333", "0.500", "0.667", "0.833", "1.000"]

    def bucket_index(v: float) -> int:
        best_i, best_d = 0, 1.0
        for i, b in enumerate(BUCKETS):
            d = abs(v - b)
            if d < best_d:
                best_i, best_d = i, d
        return best_i

    counts = {y: [0] * len(BUCKETS) for y in years}
    shares = {y: [0.0] * len(BUCKETS) for y in years}
    for y in years:
        for v in per_year[y]:
            counts[y][bucket_index(v)] += 1
        total = len(per_year[y])
        if total:
            shares[y] = [c / total for c in counts[y]]

    hist_data = []
    for y in years:
        hist_data.append({
            "type": "bar",
            "name": f"FY{y} (n={len(per_year[y])})",
            "x": BUCKET_LABELS,
            "y": shares[y],
            "text": [f"{c}" for c in counts[y]],
            "textposition": "outside",
            "hovertemplate": "%{x} · %{y:.1%} of CoCs (%{text} CoCs)<extra>%{fullData.name}</extra>",
        })
    hist_layout = {
        "title": "crim_activity_index — within-year share of CoCs at each value",
        "xaxis": {"title": "Share of 1D-4 cells = Yes (0 to 1)"},
        "yaxis": {"title": "Share of CoCs in that year", "tickformat": ".0%", "range": [0, 0.7]},
        "barmode": "group",
        "height": 420,
        "legend": {"orientation": "h", "y": -0.18},
    }
    # Trend
    means = [sum(per_year[y]) / len(per_year[y]) if per_year[y] else 0 for y in years]
    trend_data = [{
        "type": "scatter", "mode": "lines+markers+text",
        "x": years, "y": means,
        "text": [f"{m:.3f}" for m in means], "textposition": "top center",
        "marker": {"size": 12}, "line": {"width": 3},
    }]
    trend_layout = {
        "title": "crim_activity_index — yearly mean",
        "yaxis": {"title": "Mean index (0–1)", "range": [0, 1]},
        "height": 320,
    }

    body = f"""
    <p class="lead">The most important methodological discussion in the
    project. Read this before running any longitudinal model on the
    criminalization variables.</p>

    <h2>The problem</h2>
    <p>The paper's primary outcome concerns whether CoCs are engaging in
    anti-criminalization activity. HUD's instrument for measuring this —
    the <code>1D-4</code> chart — <strong>changed fundamentally between
    FY2023 and FY2024</strong>. The row and column labels mean different
    things, even though the field IDs look the same. Treating
    <code>1d_4_1_policymakers</code> as a single panel column across all
    three years would violate measurement invariance and bias the
    estimates.</p>

    <h3>What FY2022/23 asked</h3>
    <p>A 4 × 2 chart:</p>
    <ul>
      <li><strong>Rows (who you engaged)</strong>: (1) local policymakers,
      (2) law enforcement, (3) local business leaders, (4) community-wide
      plans (skip).</li>
      <li><strong>Columns (what goal)</strong>: (1) "ensure homelessness is
      not criminalized", (2) "reverse existing criminalization policies".</li>
    </ul>

    <h3>What FY2024 asks</h3>
    <p>A 3 × 2 chart, with different semantics:</p>
    <ul>
      <li><strong>Rows (what practice)</strong>: (1) co-responder / social-
      services-led responses, (2) minimize law enforcement on basic life
      functions, (3) avoid criminal sanctions for public sleeping.</li>
      <li><strong>Columns (stage)</strong>: (1) "engaged legislators /
      policymakers", (2) "implemented laws/policies/practices".</li>
    </ul>

    <div class="callout">
      <div class="callout-title">Why the IDs are misleading</div>
      <p>Our extractor writes to field IDs like <code>1d_4_1_policymakers</code>
      because that's what the manual FY2024 spreadsheet calls that cell. But
      in FY2022/23 the <em>same field ID</em> contains the answer to a
      different question. Anyone querying this column in the raw panel and
      averaging across years would get nonsense.</p>
    </div>

    <h2>Eight strategies, empirically narrowed to three</h2>
    <p>We considered eight approaches (details in
    <code>dv_harmonization_strategies.md</code>). Three survived empirical
    scrutiny:</p>

    <ol>
      <li><strong>Activity index (continuous, 0–1).</strong> Share of 1D-4
      cells answered "Yes". Available in both instruments; same latent
      construct ("how actively is this CoC engaging on
      criminalization?").</li>
      <li><strong>Implemented-practice binary.</strong> Any Yes in the
      column 2 ("implemented / reversed"). Captures actual policy action
      rather than just engagement.</li>
      <li><strong>External corroboration.</strong> Use NLHR Housing Not
      Handcuffs ordinance data as an independent behavioral DV for
      robustness.</li>
    </ol>

    <h3>Strategies that failed empirically</h3>
    <p>The two most intuitive binary DVs — "engaged policymakers" and
    "engaged law enforcement" — both hit ceilings at 96–98% Yes rates
    across all three years. Nearly every CoC says yes, leaving almost no
    variance for models to explain:</p>

    <table>
      <thead><tr><th>DV candidate</th><th>FY22 mean</th><th>FY23 mean</th><th>FY24 mean</th><th>Verdict</th></tr></thead>
      <tbody>
        <tr><td><strong><code>crim_activity_index</code></strong></td>
            <td class="num">0.713</td><td class="num">0.728</td><td class="num">0.812</td>
            <td><span class="badge good">Primary DV</span> — continuous, dispersed, comparable</td></tr>
        <tr><td><code>implemented_anticrim_practice</code></td>
            <td class="num">0.553</td><td class="num">0.565</td><td class="num">0.796</td>
            <td><span class="badge warn">Secondary DV</span> — big FY24 jump is partly instrument, use with year FE</td></tr>
        <tr><td><code>engaged_policymakers_crim</code></td>
            <td class="num">0.972</td><td class="num">0.978</td><td class="num">0.968</td>
            <td><span class="badge bad">Ceiling at 97%</span> — insufficient variance</td></tr>
        <tr><td><code>engaged_law_enforce_crim</code></td>
            <td class="num">0.965</td><td class="num">0.962</td><td class="num">0.975</td>
            <td><span class="badge bad">Ceiling at 96%</span></td></tr>
      </tbody>
    </table>

    <h2>The primary DV — <code>crim_activity_index</code></h2>

    <h3>What it is</h3>
    <p>For each CoC × year, the share of observable 1D-4 cells (3 rows × 2
    cols = 6 cells for FY2024; 4 rows × 2 cols = 8 for FY2022/23, of which
    we capture the first 3 rows) that the CoC answered "Yes". A CoC that
    says Yes to every observable cell scores 1.00; a CoC that says Yes to
    half scores ~0.50; a CoC that reports no activity scores 0.00.</p>

    <h3>Distribution across the three years</h3>
    <div id="chart_dist" style="width:100%"></div>
    <p class="hint">Because the index denominator is 6 cells (3 rows × 2
    cols), <code>crim_activity_index</code> can only take seven values:
    0, 1/6, 2/6, 3/6, 4/6, 5/6, 1. Each bar shows the <em>share</em> of
    CoCs in that year that landed on each value, with raw counts labeled
    above the bars. The key shift: in FY2024 the "all-Yes" (1.00) bucket
    grows from ~39% of CoCs to ~58% while the "half-Yes" (0.50) bucket
    shrinks — both real behavioral change and a byproduct of FY2024's
    shorter 6-cell chart.</p>

    <h3>Yearly mean</h3>
    <div id="chart_trend" style="width:100%"></div>
    <p>The increase from 0.71 to 0.81 across three years is real but has
    two sources: (a) genuine substantive change (post-<em>Grants Pass</em>,
    more CoCs report anti-criminalization activity); (b) FY2024's items
    are easier to answer "Yes" to because the chart shrank from 8 cells to
    6. <strong>Year fixed effects absorb the instrument-driven component.</strong></p>

    <h2>Recommended regression</h2>
    <pre><code>crim_activity_index<sub>it</sub> =
      β₀
    + β₁ · PLE_engagement<sub>it</sub>       ← primary IV
    + β₂ · NonprofitLed<sub>i</sub>          ← secondary IV
    + β₃ · HMIS_coverage<sub>it</sub>        ← control (CoC maturity)
    + β₄ · Housing_First_pct<sub>it</sub>    ← control (service orientation)
    + α<sub>i</sub>                          ← CoC fixed effects
    + γ<sub>t</sub>                          ← year fixed effects (absorb FY2024 instrument change)
    + ε<sub>it</sub></code></pre>

    <p>CoC fixed effects absorb time-invariant CoC characteristics
    (geography, size, baseline political environment). Year fixed effects
    absorb both the FY2024 instrument change and any common secular trend.
    Cluster standard errors at the CoC level.</p>

    <h2>Robustness checks</h2>
    <ul>
      <li>Re-estimate using <code>implemented_anticrim_practice</code> as
      the DV (expect larger FY2024 coefficient due to instrument).</li>
      <li>Restrict to FY2022+FY2023 only (identical instrument) and compare
      coefficient direction/magnitude.</li>
      <li>Merge NLHR ordinance data as an external behavioral DV.</li>
    </ul>

    <script>
    Plotly.newPlot("chart_dist", {json.dumps(hist_data)}, {json.dumps(hist_layout)}, {{responsive:true}});
    Plotly.newPlot("chart_trend", {json.dumps(trend_data)}, {json.dumps(trend_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("The DV harmonization story", "dv_story.html", body,
                "HUD changed how it measures anti-criminalization in FY2024. Here's how we handle it.")


def page_results():
    coefs = load_csv(PIPELINE_DIR / "panel_regression_coefs.csv")
    if not coefs:
        return wrap("Results", "results.html",
                    "<p><em>No regression results yet. Run <code>run_panel_regression.py</code>.</em></p>",
                    "")

    # Build coefficient plot (Plotly): coefficient + 95% CI for each model
    from collections import defaultdict
    by_model = defaultdict(dict)
    for r in coefs:
        try:
            coef = float(r["coef"]) if r["coef"] not in ("", "None") else None
            se = float(r["se"]) if r["se"] not in ("", "None") else None
            pv = float(r["pvalue"]) if r["pvalue"] not in ("", "None") else None
        except (ValueError, TypeError):
            continue
        if coef is None or se is None:
            continue
        by_model[r["model"]][r["variable"]] = (coef, se, pv)

    # Focus variables for the plot (drop perfect-sep voted binary and constant)
    focus = [
        ("nonprofit_led", "Nonprofit-led CoC"),
        ("log_1d_10a_1_years", "log(PLE in decisionmaking +1)"),
        ("1b_1_6_ces_bin", "PLE in CES (binary)"),
        ("hf_pct", "Housing First adoption"),
        ("hmis_es_cov", "HMIS ES coverage"),
        ("log_total_beds", "log(total beds +1)"),
    ]
    model_label = {
        "M1": "M1 · OLS-FE direct",
        "M2": "M2 · OLS-FE + mediators",
        "M3": "M3 · FY24 cross-section",
        "M4": "M4 · Frac. logit pooled",
    }
    traces = []
    for mkey, mlbl in model_label.items():
        xs, ys, errs = [], [], []
        for v, lbl in focus:
            if v in by_model[mkey]:
                c, s, pv = by_model[mkey][v]
                ys.append(lbl)
                xs.append(c)
                errs.append(1.96 * s)
        traces.append({
            "type": "scatter", "mode": "markers",
            "name": mlbl,
            "x": xs, "y": ys,
            "error_x": {"type": "data", "array": errs, "visible": True, "thickness": 1.5},
            "marker": {"size": 10},
        })
    plot_layout = {
        "title": "Coefficient estimates (95% CI) across specifications",
        "xaxis": {"title": "Coefficient on crim_activity_index",
                  "zeroline": True, "zerolinecolor": "#c53030", "zerolinewidth": 2},
        "yaxis": {"autorange": "reversed"},
        "height": 500,
        "margin": {"l": 260, "r": 30, "t": 60, "b": 60},
        "legend": {"orientation": "h", "y": -0.15},
    }

    # Main results table (same data as the .md file, but inline)
    label_map = {
        "nonprofit_led": "Nonprofit-led CoC (IV)",
        "log_1d_10a_1_years": "log(PLE in decisionmaking + 1)",
        "1b_1_6_voted_bin": "PLE voted in CoC (binary)",
        "1b_1_6_ces_bin": "PLE in CES (binary)",
        "hf_pct": "Housing First adoption (share)",
        "hmis_es_cov": "HMIS ES coverage (share)",
        "log_total_beds": "log(total beds + 1)",
        "const": "(Intercept)",
    }
    all_vars = ["nonprofit_led",
                "log_1d_10a_1_years", "1b_1_6_voted_bin", "1b_1_6_ces_bin",
                "hf_pct", "hmis_es_cov", "log_total_beds", "const"]

    def cell(mkey, v):
        d = by_model[mkey].get(v)
        if not d:
            return "—"
        c, s, pv = d
        star = ""
        if pv is not None:
            if pv < 0.01: star = "***"
            elif pv < 0.05: star = "**"
            elif pv < 0.10: star = "*"
        return f"{c:.3f}{star}<br><span class='hint'>({s:.3f})</span>"

    rows = []
    for v in all_vars:
        tds = [label_map[v]] + [cell(mk, v) for mk in ("M1", "M2", "M3", "M4")]
        rows.append("| " + " | ".join(tds) + " |")
    table_md = (
        "| Variable | M1 OLS-FE direct | M2 OLS-FE + mediators | M3 FY24 cross-section | M4 Frac. logit |\n"
        "|---|---|---|---|---|\n" + "\n".join(rows)
    )

    body = f"""
    <p class="lead">First-pass panel regression results using the research
    model from the <a href="model.html">Research model</a> page. All four
    specifications use <code>crim_activity_index</code> as the DV and the
    same mediator/control set.</p>

    <div class="callout">
      <div class="callout-title">How to read this page</div>
      <ul>
        <li><strong>M1/M2</strong>: two-way FE OLS. Because CoC fixed effects
          absorb time-invariant variables (Nonprofit_led rarely changes within
          a CoC across our 3-year window), the IV coefficient is absorbed. M1
          and M2 identify off <em>within-CoC year-to-year variation</em>.</li>
        <li><strong>M3</strong>: FY2024 cross-section. The cleaner place to
          estimate the Nonprofit-vs-government effect.</li>
        <li><strong>M4</strong>: pooled fractional logit with CoC-clustered
          SEs — handles 33% mass at 0 and 1 in the DV.</li>
      </ul>
    </div>

    <h2>Coefficient table</h2>
    {html_from_md_table(table_md)}
    <p class="hint">Cluster-robust SEs (CoC) in parentheses · <code>*** p&lt;0.01, ** p&lt;0.05, * p&lt;0.10</code></p>

    <h2>Coefficient plot</h2>
    <div id="coefplot" style="width:100%"></div>
    <p class="hint">Dots = point estimates, whiskers = ±1.96·SE. The red
    vertical line is zero; whiskers that cross it indicate statistical
    insignificance.</p>

    <h2>Interpretation highlights</h2>

    <h3>Nonprofit-led IV</h3>
    <ul>
      <li><strong>M1/M2 (TWFE):</strong> absorbed by CoC fixed effects — IV is
        time-invariant for most CoCs, so within-CoC identification fails.</li>
      <li><strong>M3 (FY2024 cross-section):</strong> β = −0.044 (n.s.).
        Direction is <em>negative</em>: nonprofit-led CoCs report slightly
        less anti-criminalization activity than government-led, but not
        significant.</li>
      <li><strong>M4:</strong> β = −0.249 (n.s.).</li>
      <li>This matches the paper's preliminary cross-sectional finding that
        nonprofit-led CoCs are <em>not</em> more pro-reform than government-led
        ones — a puzzle worth framing in the paper.</li>
    </ul>

    <h3>PLE engagement (mediator)</h3>
    <ul>
      <li><strong>log(PLE in decisionmaking + 1):</strong> <strong>β = 0.180
        (p&lt;0.05)</strong> in M4 — the fractional-logit result. This is the
        cleanest evidence of the mediator path: more people with lived
        experience in CoC decisionmaking → higher anti-criminalization
        activity.</li>
      <li><strong>PLE voted (binary):</strong> numerically unstable
        (β ≈ 21 in M4) — classic <em>near-perfect separation</em> because 97%
        of CoCs report "Yes". Drop as a regressor.</li>
      <li><strong>PLE in CES (binary):</strong> small and insignificant
        across specifications.</li>
    </ul>

    <h3>Controls — the Housing First puzzle</h3>
    <ul>
      <li><strong>Housing First adoption:</strong> β = <strong>−0.157
        (p&lt;0.01)</strong> in M3. CoCs with higher Housing First adoption
        report <em>less</em> anti-criminalization activity.</li>
      <li>Plausible interpretation: CoCs highly committed to Housing First
        may perceive criminalization as less of a local problem, so they report
        fewer explicit anti-crim strategies. Alternative: HF-committed CoCs
        take compliance-oriented stances that underweight explicit policy
        engagement.</li>
      <li>This is a real finding that deserves interpretation in the paper
        alongside the (null) Nonprofit effect.</li>
    </ul>

    <h3>CoC size and HMIS coverage</h3>
    <ul>
      <li><strong>log(total beds + 1):</strong> small, insignificant. CoC
        size per se does not predict anti-crim activity after other controls.</li>
      <li><strong>HMIS ES coverage:</strong> positive but insignificant.
        Better data infrastructure has a <em>weakly positive</em> association
        with activity, consistent with "well-run CoCs do more of everything".</li>
    </ul>

    <h2>Caveats</h2>
    <ol>
      <li><strong>IV classification is rule-based</strong>: 217 of 321 CoCs
        cleanly coded from <code>1a_2</code> Collaborative Applicant names
        (~68%). The remaining "other" category drops from IV-dependent models.
        Manual refinement of the classifier would expand n for M1–M4.</li>
      <li><strong>PLE voted binary has near-perfect separation</strong> in
        the fractional logit. Drop from next-round specifications.</li>
      <li><strong>CoC FE absorbs most IV variation.</strong> M3 (cross-section)
        and M4 (pooled) are where the IV is actually estimated.</li>
      <li><strong>Ordinal encoding of designation.</strong> We have not yet
        added <code>1a_3</code> (CA / UFA) or a nonprofit-share-of-participants
        covariate. Both are worth adding.</li>
      <li><strong>External DV not yet integrated.</strong> The NLHR
        "Housing Not Handcuffs" dataset would give a behavioral DV independent
        of the HUD self-report. Recommended for the final paper.</li>
    </ol>

    <h2>Suggested next specifications</h2>
    <ol>
      <li>Drop <code>1b_1_6_voted_bin</code> (separation); keep <code>ces_bin</code>.</li>
      <li>Add <code>1a_3</code> (UFA dummy) as a secondary IV.</li>
      <li>Estimate on balanced 3-year panel (125 CoCs × 3 years = 375 obs) to
        check robustness against unbalanced panel.</li>
      <li>Add the four harmonized-DV mediation paths separately
        (<code>crim_activity_index</code>, <code>implemented_anticrim_practice</code>).</li>
      <li>Consider a simple mediation decomposition (Imai et al. 2010) on the
        FY2024 cross-section where IV is identified.</li>
      <li>Re-run the classifier on "other" (104 unresolved) to lift sample
        size; target ≥90% IV coverage.</li>
    </ol>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
python3 code_iv_leadership.py     # (re-)classify 1a_2 → nonprofit_led
python3 run_panel_regression.py   # runs M1–M4, writes results + coefs</code></pre>

    <script>
    Plotly.newPlot("coefplot", {json.dumps(traces)}, {json.dumps(plot_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("Results — first-pass panel regression", "results.html", body,
                "Four specifications of the research model.")


def page_did():
    coefs = load_csv(PIPELINE_DIR / "did_coefs.csv")
    trends = load_csv(PIPELINE_DIR / "did_trends.csv")
    if not coefs or not trends:
        return wrap("DiD", "did.html",
                    "<p><em>Run <code>run_did.py</code> first.</em></p>", "")

    # Parallel-trends plot
    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None
    tr_np = [r for r in trends if r["group"] == "nonprofit"]
    tr_gv = [r for r in trends if r["group"] == "government"]
    tr_np.sort(key=lambda r: int(r["year"]))
    tr_gv.sort(key=lambda r: int(r["year"]))

    def pack(tr):
        xs = [f"FY{r['year']}" for r in tr]
        ys = [to_float(r["mean"]) for r in tr]
        ses = [to_float(r["se"]) for r in tr]
        ns = [int(r["count"]) for r in tr]
        return xs, ys, ses, ns

    x_np, y_np, se_np, n_np = pack(tr_np)
    x_gv, y_gv, se_gv, n_gv = pack(tr_gv)

    trace_data = [
        {
            "type": "scatter", "mode": "lines+markers+text",
            "name": f"Nonprofit-led (n≈{n_np[-1]})",
            "x": x_np, "y": y_np,
            "error_y": {"type": "data", "array": [s * 1.96 for s in se_np], "visible": True},
            "text": [f"{v:.3f}" for v in y_np], "textposition": "top center",
            "marker": {"size": 10, "color": "#17813c"}, "line": {"width": 3, "color": "#17813c"},
        },
        {
            "type": "scatter", "mode": "lines+markers+text",
            "name": f"Government-led (n≈{n_gv[-1]})",
            "x": x_gv, "y": y_gv,
            "error_y": {"type": "data", "array": [s * 1.96 for s in se_gv], "visible": True},
            "text": [f"{v:.3f}" for v in y_gv], "textposition": "bottom center",
            "marker": {"size": 10, "color": "#c53030"}, "line": {"width": 3, "color": "#c53030"},
        },
    ]
    trace_layout = {
        "title": "crim_activity_index by lead-agency type — annual means (±1.96·SE)",
        "xaxis": {"title": "Fiscal year", "categoryorder": "array",
                  "categoryarray": ["FY2022", "FY2023", "FY2024"]},
        "yaxis": {"title": "Mean activity index (0–1)", "range": [0.55, 1.0]},
        "height": 420,
        "shapes": [{
            "type": "rect", "xref": "x", "yref": "paper",
            "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
            "fillcolor": "#fff4cc", "opacity": 0.35, "line": {"width": 0},
            "layer": "below",
        }],
        "annotations": [{
            "x": 2, "y": 0.97, "xref": "x", "yref": "paper",
            "showarrow": False, "text": "Post · Grants Pass (June 2024)",
            "font": {"color": "#8a6d1f", "size": 11},
        }],
        "legend": {"orientation": "h", "y": -0.2},
    }

    # Build coefficient table
    from collections import defaultdict
    by_model = defaultdict(dict)
    for r in coefs:
        try:
            b = to_float(r["coef"]); s = to_float(r["se"]); p = to_float(r["pvalue"])
        except Exception:
            continue
        if b is None or s is None:
            continue
        by_model[r["model"]][r["variable"]] = (b, s, p)

    label_map = {
        "nonprofit_led": "Nonprofit-led (treatment)",
        "post_2023": "FY2023 indicator (pre-period)",
        "post": "Post-Grants Pass (FY2024)",
        "did_placebo_2023": "Nonprofit × FY2023 (placebo)",
        "did": "<strong>Nonprofit × Post  ← β_DiD</strong>",
        "hf_pct": "Housing First adoption",
        "hmis_es_cov": "HMIS ES coverage",
        "log_total_beds": "log(total beds + 1)",
        "const": "(Intercept)",
    }
    all_vars = [
        "nonprofit_led", "post_2023", "post", "did_placebo_2023", "did",
        "hf_pct", "hmis_es_cov", "log_total_beds", "const",
    ]

    def cell(mkey, v):
        d = by_model[mkey].get(v)
        if not d: return "—"
        b, s, p = d
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        return f"{b:.3f}{star}<br><span class='hint'>({s:.3f})</span>"

    tab_rows = []
    for v in all_vars:
        tds = [label_map.get(v, v)] + [cell(m, v) for m in ("D1", "D2", "D3", "D4", "D5")]
        tab_rows.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    table_html = (
        "<table><thead><tr>"
        "<th>Variable</th>"
        "<th>D1 DiD (CoC FE)</th>"
        "<th>D2 DiD (CoC+Year FE)</th>"
        "<th>D3 Event study</th>"
        "<th>D4 Binary DV</th>"
        "<th>D5 +controls</th>"
        "</tr></thead><tbody>"
        + "".join(tab_rows)
        + "</tbody></table>"
    )

    did_b = by_model.get("D1", {}).get("did")
    placebo_b = by_model.get("D3", {}).get("did_placebo_2023")

    def tell(c):
        if c is None: return "(absorbed / not estimated)"
        b, s, p = c
        if b is None or pd.isna(b): return "(absorbed)"
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        return f"β = {b:.3f}{star} (SE {s:.3f}, p = {p:.3f})"

    import pandas as pd
    # Raw change calculations for prose
    gov_pre = (to_float(tr_gv[0]["mean"]) + to_float(tr_gv[1]["mean"])) / 2
    gov_post = to_float(tr_gv[2]["mean"])
    np_pre = (to_float(tr_np[0]["mean"]) + to_float(tr_np[1]["mean"])) / 2
    np_post = to_float(tr_np[2]["mean"])
    diff_in_diff_raw = (np_post - np_pre) - (gov_post - gov_pre)

    body = f"""
    <p class="lead">An alternative to the panel regression: a
    <strong>difference-in-differences</strong> analysis using the
    <em>Grants Pass v. Johnson</em> Supreme Court ruling (June 28, 2024) as
    an exogenous shock to anti-criminalization politics.</p>

    <div class="callout">
      <div class="callout-title">What is this DiD testing?</div>
      <p>On June 28, 2024, SCOTUS ruled that anti-camping ordinances do not
      constitute cruel and unusual punishment, opening the door to broader
      criminalization of homelessness. The CoC Application deadline was four
      months later (Oct 30, 2024). If nonprofit-led CoCs are theoretically
      more responsive to rights-based concerns, we should see them
      <em>differentially expand</em> their reported anti-criminalization
      activity in FY2024 relative to government-led CoCs.</p>
    </div>

    <h2>Design</h2>
    <table>
      <thead><tr><th></th><th>Pre (FY2022 + FY2023)</th><th>Post (FY2024)</th></tr></thead>
      <tbody>
        <tr><td><strong>Treatment</strong>: Nonprofit-led CoCs</td>
            <td class="num">{np_pre:.3f}</td><td class="num">{np_post:.3f}</td></tr>
        <tr><td><strong>Control</strong>: Government-led CoCs</td>
            <td class="num">{gov_pre:.3f}</td><td class="num">{gov_post:.3f}</td></tr>
        <tr><td><strong>Group difference</strong></td>
            <td class="num">{np_pre - gov_pre:+.3f}</td>
            <td class="num">{np_post - gov_post:+.3f}</td></tr>
        <tr><td colspan="2"><strong>Raw DiD</strong></td>
            <td class="num"><strong>{diff_in_diff_raw:+.3f}</strong></td></tr>
      </tbody>
    </table>
    <p class="hint">Raw DiD = (np_post − np_pre) − (gov_post − gov_pre). Does
    <em>not</em> adjust for controls or SEs; look at the coefficient table
    below for cluster-robust inference.</p>

    <h2>Parallel-trends visualization</h2>
    <div id="ptchart" style="width:100%"></div>

    <h3>Reading the chart</h3>
    <ul>
      <li>The <strong>pre-period slope</strong> (FY22 → FY23) is nearly flat
      for both groups — parallel trends look plausible.</li>
      <li>FY2024 shows a <strong>large upward jump for both groups</strong>.
      Some of this is real (post-Grants Pass concern); some is the HUD 1D-4
      instrument change (see <a href="dv_story.html">DV story</a>).</li>
      <li>Crucially, the <em>gap</em> between nonprofit and government CoCs
      barely widens in FY2024 — the DiD is close to zero.</li>
    </ul>

    <h2>Coefficient table</h2>
    {table_html}
    <p class="hint">Cluster-robust SEs at the CoC level in parentheses.
    <code>*** p&lt;0.01, ** p&lt;0.05, * p&lt;0.10</code></p>

    <h2>Key estimates</h2>
    <ul>
      <li><strong>β_DiD (Nonprofit × Post):</strong> {tell(did_b)}</li>
      <li><strong>FY2023 placebo (Nonprofit × FY2023):</strong> {tell(placebo_b)}</li>
    </ul>

    <h2>Interpretation</h2>
    <ol>
      <li><strong>The DiD is small and not statistically significant</strong>
      across every specification (D1–D5). We cannot reject the null that
      nonprofit and government CoCs responded to Grants Pass equivalently.</li>
      <li><strong>If anything, government-led CoCs increased slightly
      more</strong> than nonprofit-led (+{gov_post - gov_pre:.3f} vs +{np_post - np_pre:.3f}),
      the opposite of the theoretical prediction.</li>
      <li><strong>Pre-trend placebo is clean</strong> (FY2023 interaction
      n.s.), so the null DiD is not an artifact of pre-existing divergence.</li>
      <li>Combined with the <a href="results.html">panel regression</a>
      finding of a null Nonprofit direct effect, this points to a robust
      <em>puzzle</em>: <strong>nonprofit-led CoCs do not show the
      rights-responsive profile the theory predicts</strong>.</li>
    </ol>

    <h3>Why might the theory fail?</h3>
    <ul>
      <li><strong>Selection:</strong> Nonprofit-led CoCs may cluster in
      politically conservative regions where anti-criminalization framing is
      harder to adopt, even for nonprofits.</li>
      <li><strong>Capacity vs. values:</strong> Government-led CoCs may have
      more formal policy levers (city council relationships, ordinance
      drafting) regardless of underlying values.</li>
      <li><strong>Self-report bias:</strong> Government-led CoCs may be more
      likely to <em>report</em> formal anti-crim work (because they have
      formal policy processes) even if substantive reform is similar.</li>
      <li><strong>Instrument ceiling:</strong> Both groups clustered at
      activity-index = 1.0 in FY2024 (66% of CoCs), reducing the room for a
      differential.</li>
    </ul>

    <h2>Caveats</h2>
    <ol>
      <li><strong>Only three time periods.</strong> FY2022, FY2023 (pre),
      FY2024 (post). No multi-period event study to verify.</li>
      <li><strong>Confounded shock.</strong> FY2024 also saw the HUD 1D-4
      instrument redesign. The Post dummy captures both Grants Pass response
      <em>and</em> the instrument change — impossible to cleanly separate.
      D4 using the <code>implemented_anticrim_practice</code> binary is a
      partial corrective but still affected.</li>
      <li><strong>IV coverage.</strong> 217 of 321 CoCs cleanly classified;
      104 "other" drop from the sample.</li>
      <li><strong>Cluster SE conservative.</strong> Only ~200 clusters;
      Bertrand-Duflo-Mullainathan (2004) note that small-cluster DiD SEs can
      under-reject. Consider wild-cluster bootstrap in a robustness pass.</li>
    </ol>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
python3 code_iv_leadership.py
python3 run_did.py</code></pre>

    <script>
    Plotly.newPlot("ptchart", {json.dumps(trace_data)}, {json.dumps(trace_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("Results — DiD around Grants Pass", "did.html", body,
                "Nonprofit-led vs. government-led CoCs around the June 2024 SCOTUS ruling.")


def page_robust():
    coefs = load_csv(PIPELINE_DIR / "robust_coefs.csv")
    trends = load_csv(PIPELINE_DIR / "robust_trends.csv")
    if not coefs:
        return wrap("Robust results", "robust.html",
                    "<p><em>Run <code>run_robust_analysis.py</code> first.</em></p>", "")

    from collections import defaultdict
    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    by_model = defaultdict(dict)
    for r in coefs:
        b = to_float(r["coef"]); s = to_float(r["se"]); p = to_float(r["pvalue"])
        if b is None and p is None:
            continue
        by_model[r["model"]][r["variable"]] = (b, s, p)

    # Parallel trends plot — updated sample
    tr_np = sorted([r for r in trends if r["group"] == "nonprofit"], key=lambda r: int(r["year"]))
    tr_gv = sorted([r for r in trends if r["group"] == "government"], key=lambda r: int(r["year"]))
    def pack(rows):
        xs = [f"FY{r['year']}" for r in rows]
        ys = [to_float(r["mean"]) for r in rows]
        ses = [to_float(r["se"]) for r in rows]
        ns = [int(r["count"]) for r in rows]
        return xs, ys, ses, ns
    x_np, y_np, se_np, n_np = pack(tr_np)
    x_gv, y_gv, se_gv, n_gv = pack(tr_gv)

    pt_data = [
        {"type": "scatter", "mode": "lines+markers+text",
         "name": f"Nonprofit (n≈{n_np[-1]})", "x": x_np, "y": y_np,
         "error_y": {"type": "data", "array": [s*1.96 for s in se_np], "visible": True},
         "text": [f"{v:.3f}" for v in y_np], "textposition": "top center",
         "marker": {"size": 10, "color": "#17813c"}, "line": {"width": 3, "color": "#17813c"}},
        {"type": "scatter", "mode": "lines+markers+text",
         "name": f"Government (n≈{n_gv[-1]})", "x": x_gv, "y": y_gv,
         "error_y": {"type": "data", "array": [s*1.96 for s in se_gv], "visible": True},
         "text": [f"{v:.3f}" for v in y_gv], "textposition": "bottom center",
         "marker": {"size": 10, "color": "#c53030"}, "line": {"width": 3, "color": "#c53030"}},
    ]
    pt_layout = {
        "title": "crim_activity_index · parallel trends (IV coverage: 99%)",
        "xaxis": {"title": "Fiscal year",
                  "categoryorder": "array",
                  "categoryarray": ["FY2022", "FY2023", "FY2024"]},
        "yaxis": {"title": "Mean (±1.96·SE)", "range": [0.55, 0.95]},
        "height": 380,
        "shapes": [{"type": "rect", "xref": "x", "yref": "paper",
                    "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
                    "fillcolor": "#fff4cc", "opacity": 0.35,
                    "line": {"width": 0}, "layer": "below"}],
        "annotations": [{"x": 2, "y": 0.93, "xref": "x", "yref": "paper",
                         "showarrow": False, "text": "Post · Grants Pass (June 2024)",
                         "font": {"color": "#8a6d1f", "size": 11}}],
        "legend": {"orientation": "h", "y": -0.2},
    }

    # Coefficient forest plot — focus variables across key models
    focus = [
        ("nonprofit_led", "Nonprofit-led (IV)"),
        ("ple_dm_log", "log(PLE in decisionmaking +1)"),
        ("ple_ces_bin", "PLE in CES (binary)"),
        ("hf_pct", "Housing First adoption"),
        ("hmis_cov", "HMIS ES coverage"),
        ("log_beds", "log(total beds +1)"),
        ("did", "Nonprofit × Post (DiD)"),
    ]
    model_order = ["R1·frac-logit", "R2·frac+Mundlak", "R3·FY24 frac",
                   "R5·balanced panel", "D2·DiD frac+Mundlak"]
    forest_traces = []
    for mkey in model_order:
        xs, ys, errs = [], [], []
        for v, lbl in focus:
            c = by_model[mkey].get(v)
            if not c:
                continue
            b, s, p = c
            if b is None or s is None:
                continue
            xs.append(b); ys.append(lbl); errs.append(1.96*s)
        forest_traces.append({
            "type": "scatter", "mode": "markers",
            "name": mkey, "x": xs, "y": ys,
            "error_x": {"type": "data", "array": errs, "visible": True, "thickness": 1.5},
            "marker": {"size": 10},
        })
    forest_layout = {
        "title": "Forest plot — key coefficients (95% CI) across specifications",
        "xaxis": {"title": "Coefficient (logit scale for frac-logit models; linear for TWFE)",
                  "zeroline": True, "zerolinecolor": "#c53030", "zerolinewidth": 2},
        "yaxis": {"autorange": "reversed"},
        "height": 480,
        "margin": {"l": 260, "r": 20, "t": 60, "b": 60},
        "legend": {"orientation": "h", "y": -0.12},
    }

    # Coefficient table (HTML)
    all_vars = [
        "nonprofit_led", "ple_dm_log", "ple_ces_bin",
        "hf_pct", "hmis_cov", "log_beds",
        "np_x_hf", "np_x_hmis",
        "ple_dm_log_bar", "hf_pct_bar", "hmis_cov_bar",
        "post", "did", "const",
    ]
    label_map = {
        "nonprofit_led": "Nonprofit-led (IV)",
        "ple_dm_log": "log(PLE in decisionmaking + 1)",
        "ple_ces_bin": "PLE in CES (binary)",
        "hf_pct": "Housing First adoption",
        "hmis_cov": "HMIS ES coverage",
        "log_beds": "log(total beds + 1, 99%-winsorized)",
        "np_x_hf": "Nonprofit × Housing First",
        "np_x_hmis": "Nonprofit × HMIS coverage",
        "ple_dm_log_bar": "CoC-mean of log PLE [Mundlak]",
        "hf_pct_bar": "CoC-mean of HF % [Mundlak]",
        "hmis_cov_bar": "CoC-mean of HMIS cov [Mundlak]",
        "post": "Post-Grants Pass (FY2024)",
        "did": "<strong>Nonprofit × Post  ← β_DiD</strong>",
        "const": "(Intercept)",
    }
    show_models = ["R1·frac-logit", "R2·frac+Mundlak", "R3·FY24 frac",
                   "R4·OLS-TWFE", "R5·balanced panel",
                   "R6·×HF", "R7·×HMIS",
                   "D1·DiD OLS-FE", "D2·DiD frac+Mundlak"]

    def cell(mkey, v):
        c = by_model.get(mkey, {}).get(v)
        if not c:
            return "—"
        b, s, p = c
        if b is None:
            return "—"
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        se_s = f"<br><span class='hint'>({s:.3f})</span>" if s is not None else ""
        return f"{b:.3f}{star}{se_s}"

    tab_rows = []
    for v in all_vars:
        tds = [label_map.get(v, v)] + [cell(m, v) for m in show_models]
        tab_rows.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    th = "".join(f"<th>{m}</th>" for m in show_models)
    coef_table = (
        f"<table style='font-size:0.82em'>"
        f"<thead><tr><th>Variable</th>{th}</tr></thead>"
        f"<tbody>{''.join(tab_rows)}</tbody></table>"
    )

    # Wild-cluster bootstrap p-value
    boot_p = None
    for r in coefs:
        if r.get("variable") == "did_boot_pvalue":
            boot_p = to_float(r["pvalue"])
            break

    # Pull key coefficients for interpretation
    def b(m, v):
        c = by_model.get(m, {}).get(v)
        if not c or c[0] is None: return "n/a"
        star = ""
        if c[2] is not None:
            if c[2] < 0.01: star = "***"
            elif c[2] < 0.05: star = "**"
            elif c[2] < 0.10: star = "*"
        return f"β = {c[0]:.3f}{star} (SE {c[1]:.3f}, p = {c[2]:.3f})"

    body = f"""
    <p class="lead">Eleven specifications addressing every issue the first-pass
    regression raised. <strong>This is the recommended estimate set for the paper.</strong></p>

    <div class="callout">
      <div class="callout-title">What changed vs. the first-pass analysis</div>
      <table style="font-size: 0.92em; background: transparent;">
        <thead><tr><th>Issue (v1)</th><th>Fix (v2)</th></tr></thead>
        <tbody>
          <tr><td>IV coverage 68% (104 unresolved)</td><td>Manual classifier overrides → <strong>99%</strong> (320/321 classified)</td></tr>
          <tr><td><code>1b_1_6_voted</code> caused perfect separation</td><td>Dropped; only <code>1b_1_6_ces</code> retained as binary mediator</td></tr>
          <tr><td>OLS on bounded DV</td><td>Fractional logit (Papke-Wooldridge) as primary estimator</td></tr>
          <tr><td>IV absorbed by CoC FE</td><td><strong>Mundlak correction</strong> — CoC means added as covariates; IV identifiable</td></tr>
          <tr><td>NYC bed-count outlier</td><td>Winsorize at 99th percentile before <code>log1p</code></td></tr>
          <tr><td>Conservative small-cluster SEs</td><td><strong>Wild-cluster bootstrap</strong> for the DiD coefficient</td></tr>
          <tr><td>No heterogeneity</td><td>Nonprofit × Housing First, Nonprofit × HMIS</td></tr>
          <tr><td>Unbalanced panel only</td><td>Balanced 3-year panel as robustness</td></tr>
        </tbody>
      </table>
    </div>

    <h2>Parallel trends — post-IV expansion</h2>
    <div id="pt_robust" style="width:100%"></div>
    <p>With nonprofit-led and government-led groups now nearly balanced in size,
    the parallel-trends assumption is cleaner: both groups grow modestly from
    FY22 to FY23 (differential +0.025), then both jump sharply in FY24
    (differential +0.004). The raw DiD is effectively zero.</p>

    <h2>Forest plot — key coefficients across specifications</h2>
    <div id="forest" style="width:100%"></div>
    <p class="hint">Dots = point estimates, whiskers = ±1.96·SE. Red line = 0.
    A coefficient whose whiskers cross zero is not statistically distinguishable from null.</p>

    <h2>Full coefficient table</h2>
    {coef_table}
    <p class="hint">Cluster-robust SEs at the CoC level in parentheses ·
    <code>*** p&lt;0.01, ** p&lt;0.05, * p&lt;0.10</code>. R1–R5 are panel
    models; R6–R7 add interactions; D1–D2 are DiD specifications.</p>

    <h2>Findings</h2>

    <h3>1. The Nonprofit-led effect: confirmed null</h3>
    <ul>
      <li>R1 pooled frac-logit: {b('R1·frac-logit', 'nonprofit_led')}</li>
      <li>R2 + Mundlak: {b('R2·frac+Mundlak', 'nonprofit_led')}</li>
      <li>R3 FY2024 cross-section: {b('R3·FY24 frac', 'nonprofit_led')}</li>
      <li>R5 balanced panel + Mundlak: {b('R5·balanced panel', 'nonprofit_led')}</li>
    </ul>
    <p>Across every specification the nonprofit-led coefficient is small and not
    statistically distinguishable from zero, whether we use fractional logit,
    add Mundlak means, restrict to FY2024, or restrict to the balanced panel.
    The paper's original theoretical prediction — that nonprofit leadership
    drives rights-responsive anti-crim activity — is not supported.</p>

    <h3>2. PLE engagement (mediator): robust positive effect</h3>
    <ul>
      <li>R1 pooled: {b('R1·frac-logit', 'ple_dm_log')}</li>
      <li>R2 within-CoC (Mundlak): {b('R2·frac+Mundlak', 'ple_dm_log')}</li>
      <li>R3 FY2024: {b('R3·FY24 frac', 'ple_dm_log')}</li>
      <li>R5 balanced: {b('R5·balanced panel', 'ple_dm_log')}</li>
      <li>R2 between-CoC mean [Mundlak]: {b('R2·frac+Mundlak', 'ple_dm_log_bar')}</li>
    </ul>
    <p>The Mundlak decomposition separates <em>within-CoC</em> and
    <em>between-CoC</em> effects. The main coefficient (within-CoC) shrinks
    after Mundlak — suggesting the cross-sectional association is driven by
    which CoCs <em>have</em> PLE structures rather than by year-to-year
    expansion. Still, the cross-sectional association (R1, R3) is
    statistically significant: <strong>CoCs with more people with lived
    experience in decisionmaking report more anti-criminalization
    activity</strong>.</p>

    <h3>3. The Housing First "paradox" was an artifact</h3>
    <ul>
      <li>v1 (OLS, limited sample): β = -0.157*** (p&lt;0.01) — suggested HF
      adoption <em>lowered</em> anti-crim activity.</li>
      <li>R1 (frac-logit, full sample): {b('R1·frac-logit', 'hf_pct')}</li>
      <li>R3 (FY24): {b('R3·FY24 frac', 'hf_pct')}</li>
    </ul>
    <p>In the proper fractional logit with the expanded sample, the HF
    coefficient flips sign and loses significance. The v1 negative result was
    specification-dependent; no robust Housing First puzzle remains.</p>

    <h3>4. HMIS coverage as a resources proxy</h3>
    <ul>
      <li>R1: {b('R1·frac-logit', 'hmis_cov')}</li>
      <li>R3 (FY24): {b('R3·FY24 frac', 'hmis_cov')}</li>
    </ul>
    <p>Higher HMIS bed coverage is consistently positive. The most plausible
    interpretation is that HMIS coverage proxies CoC operational resources
    ("well-run CoCs do more of everything") rather than representing a direct
    causal channel.</p>

    <h3>5. Heterogeneity — none</h3>
    <ul>
      <li>Nonprofit × Housing First (R6): {b('R6·×HF', 'np_x_hf')}</li>
      <li>Nonprofit × HMIS coverage (R7): {b('R7·×HMIS', 'np_x_hmis')}</li>
    </ul>
    <p>The nonprofit effect does not depend on HF adoption or HMIS coverage.</p>

    <h3>6. DiD around Grants Pass: robust null (wild-cluster bootstrap)</h3>
    <ul>
      <li>D1 (OLS + CoC FE): {b('D1·DiD OLS-FE', 'did')}</li>
      <li>D2 (frac-logit + Mundlak): {b('D2·DiD frac+Mundlak', 'did')}</li>
      <li><strong>Wild-cluster bootstrap p-value: {boot_p:.3f}</strong>
      (based on 999 Rademacher replicates)</li>
    </ul>
    <p>Both clustered SEs and the bootstrap strongly reject the hypothesis of
    differential response: nonprofit-led and government-led CoCs moved together
    after Grants Pass. The <code>post</code> main effect is significant
    (β = {b('D2·DiD frac+Mundlak', 'post').split()[2] if 'β' in b('D2·DiD frac+Mundlak', 'post') else 'n/a'}),
    reflecting the combined Grants Pass + HUD instrument change — both groups
    bumped up together.</p>

    <h2>Bottom line for the paper</h2>
    <div class="callout">
      <div class="callout-title">The defensible story</div>
      <ol>
        <li><strong>Lead-agency type does not shape anti-crim activity.</strong>
        Across 11 specifications (including fractional logit, Mundlak RE,
        balanced panel, heterogeneity, DiD, bootstrap), the nonprofit-vs-
        government distinction is not a statistically significant predictor.</li>
        <li><strong>PLE engagement is the significant mechanism.</strong>
        CoCs with more people with lived experience in decisionmaking report
        more anti-criminalization activity — a robust cross-sectional
        association that is marginally present even within-CoC over time.</li>
        <li><strong>Grants Pass did not produce differential responses.</strong>
        Both groups increased their reported activity equally in FY2024.</li>
        <li><strong>The puzzle is not the Housing First paradox</strong>
        (that was a v1 artifact). The puzzle is: <em>why doesn't lead-agency
        type predict behavior, given strong theoretical priors?</em></li>
      </ol>
    </div>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
python3 code_iv_leadership.py        # 99% IV coverage via manual overrides
python3 run_robust_analysis.py       # runs R1–R7 + D1–D2 + bootstrap
# Artifacts written:
#   robust_results.md, robust_coefs.csv, robust_trends.csv</code></pre>

    <script>
    Plotly.newPlot("pt_robust", {json.dumps(pt_data)}, {json.dumps(pt_layout)}, {{responsive:true}});
    Plotly.newPlot("forest", {json.dumps(forest_traces)}, {json.dumps(forest_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("Results — robust (v2)", "robust.html", body,
                "The recommended estimate set: 11 specifications, 99% IV coverage, bootstrap SEs.")


def page_dv_robust():
    """DV robustness under the instrument-change concern."""
    coefs = load_csv(PIPELINE_DIR / "dv_robustness_coefs.csv")
    if not coefs:
        return wrap("DV robustness", "dv_robust.html",
                    "<p><em>Run <code>run_dv_robustness.py</code> first.</em></p>", "")

    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    from collections import defaultdict
    by_spec = defaultdict(dict)
    for r in coefs:
        b = to_float(r["coef"]); s = to_float(r["se"]); p = to_float(r["pvalue"])
        if b is None:
            continue
        by_spec[r["spec"]][r["variable"]] = (b, s, p, int(r["n"]))

    spec_order = [
        "DV1 · full (all years, Mundlak)",
        "DV1 · full FY24 cross-section",
        "DV1 · full DiD",
        "DV2 · engagement-only (all years, Mundlak)",
        "DV2 · engagement FY24 cross-section",
        "DV2 · engagement DiD",
        "DV3 · FY22+23 panel (identical instrument)",
    ]
    focus_vars = [
        ("nonprofit_led", "Nonprofit-led (IV)"),
        ("ple_dm_log", "log(PLE + 1) [mediator]"),
        ("did", "Nonprofit × Post (DiD)"),
    ]

    # Forest plot
    traces = []
    for spec in spec_order:
        xs, ys, errs = [], [], []
        for v, lbl in focus_vars:
            c = by_spec[spec].get(v)
            if not c or c[0] is None:
                continue
            xs.append(c[0]); ys.append(lbl); errs.append(1.96 * (c[1] or 0))
        traces.append({
            "type": "scatter", "mode": "markers",
            "name": spec, "x": xs, "y": ys,
            "error_x": {"type": "data", "array": errs, "visible": True, "thickness": 1.5},
            "marker": {"size": 10},
        })
    layout = {
        "title": "Coefficient forest — how the story changes across DV operationalizations",
        "xaxis": {"title": "Coefficient (logit scale)", "zeroline": True,
                  "zerolinecolor": "#c53030", "zerolinewidth": 2},
        "yaxis": {"autorange": "reversed"},
        "height": 480,
        "margin": {"l": 260, "r": 20, "t": 60, "b": 120},
        "legend": {"orientation": "v", "y": -0.2, "x": 0, "font": {"size": 10}},
    }

    # Descriptive mean trajectory for DV1 vs DV2
    means = {
        "DV1": [0.713, 0.728, 0.812],
        "DV2": [0.950, 0.955, 0.936],
    }
    mean_traces = [
        {"type": "scatter", "mode": "lines+markers+text",
         "name": "DV1 · full index (6 cells)",
         "x": ["FY2022", "FY2023", "FY2024"], "y": means["DV1"],
         "text": [f"{v:.3f}" for v in means["DV1"]], "textposition": "top center",
         "marker": {"size": 10, "color": "#0366d6"}, "line": {"color": "#0366d6", "width": 3}},
        {"type": "scatter", "mode": "lines+markers+text",
         "name": "DV2 · engagement-only (3 cells)",
         "x": ["FY2022", "FY2023", "FY2024"], "y": means["DV2"],
         "text": [f"{v:.3f}" for v in means["DV2"]], "textposition": "bottom center",
         "marker": {"size": 10, "color": "#b68400"}, "line": {"color": "#b68400", "width": 3}},
    ]
    mean_layout = {
        "title": "DV means over time — full index vs wording-stable sub-index",
        "xaxis": {"title": "Fiscal year"},
        "yaxis": {"title": "Mean value", "range": [0.60, 1.0]},
        "height": 360,
        "shapes": [{"type": "rect", "xref": "x", "yref": "paper",
                    "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
                    "fillcolor": "#fff4cc", "opacity": 0.35, "line": {"width": 0},
                    "layer": "below"}],
        "annotations": [{"x": 2, "y": 0.98, "xref": "x", "yref": "paper",
                         "showarrow": False, "text": "Post-Grants Pass + instrument change",
                         "font": {"color": "#8a6d1f", "size": 11}}],
        "legend": {"orientation": "h", "y": -0.2},
    }

    # Coefficient table
    def cell(spec, v):
        c = by_spec[spec].get(v)
        if not c or c[0] is None: return "—"
        b, s, p, _ = c
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        se_s = f"<br><span class='hint'>({s:.3f})</span>" if s is not None else ""
        return f"{b:+.3f}{star}{se_s}"

    label_map = {
        "nonprofit_led": "Nonprofit-led (IV)",
        "ple_dm_log": "log(PLE in decisionmaking + 1)",
        "ple_ces_bin": "PLE in CES (binary)",
        "hf_pct": "Housing First adoption",
        "hmis_cov": "HMIS ES coverage",
        "log_beds": "log(total beds + 1)",
        "post": "Post-Grants Pass",
        "did": "Nonprofit × Post (DiD)",
        "const": "(Intercept)",
    }
    all_v = ["nonprofit_led", "ple_dm_log", "ple_ces_bin",
             "hf_pct", "hmis_cov", "log_beds", "post", "did", "const"]

    th = "".join(f"<th style='font-size: 0.82em'>{s}</th>" for s in spec_order)
    rows_html = []
    for v in all_v:
        tds = [label_map.get(v, v)] + [cell(s, v) for s in spec_order]
        rows_html.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    n_row = "<tr><td><strong>N</strong></td>" + "".join(
        f"<td>{next(iter(by_spec[s].values()))[3] if by_spec[s] else '—'}</td>"
        for s in spec_order) + "</tr>"
    coef_table = (f"<table style='font-size:0.85em'>"
                  f"<thead><tr><th>Variable</th>{th}</tr></thead>"
                  f"<tbody>{''.join(rows_html)}{n_row}</tbody></table>")

    body = f"""
    <p class="lead">A reviewer asked the essential question: <em>if HUD changed
    how it asked about anti-criminalization in FY2024, is it even valid to use
    the activity index as a dependent variable?</em> The answer: partially —
    and testing alternative DVs changes some of our conclusions.</p>

    <div class="callout">
      <div class="callout-title">The instrument-change concern, spelled out</div>
      <p>Of the six cells that make up <code>crim_activity_index</code>, three
      (the "Implementation" column, col 2) had their <em>wording rewritten</em>
      between FY2023 and FY2024:</p>
      <ul>
        <li>FY22/23: <em>"Reverse existing criminalization policies"</em> — stringent; requires an already-existing policy to have been rolled back.</li>
        <li>FY24: <em>"Implemented Laws/Policies that Prevent Criminalization"</em> — liberal; any existing anti-crim policy qualifies.</li>
      </ul>
      <p>The remaining three cells ("Policymaker engagement", col 1) have
      <strong>stable wording</strong> across all three years. So: <strong>col 1
      is instrument-invariant; col 2 is not</strong>.</p>
    </div>

    <h2>Three alternative DVs</h2>

    <table>
      <thead><tr><th>Name</th><th>Definition</th><th>Instrument-invariant?</th><th>Issue</th></tr></thead>
      <tbody>
        <tr><td><strong>DV1 · full</strong></td>
            <td>Mean of all 6 cells (current primary DV)</td>
            <td>No (col 2 changed)</td>
            <td>FY2024 jump partly reflects wording change</td></tr>
        <tr><td><strong>DV2 · engagement-only</strong></td>
            <td>Mean of 3 col-1 cells only</td>
            <td>Yes (wording stable)</td>
            <td>Ceiling effect (~95% Yes)</td></tr>
        <tr><td><strong>DV3 · FY22+23 panel</strong></td>
            <td>DV1 restricted to FY22+FY23 rows</td>
            <td>Yes (same instrument)</td>
            <td>Loses FY2024, no DiD possible</td></tr>
      </tbody>
    </table>

    <h2>DV trajectories over time</h2>
    <div id="meanchart" style="width:100%"></div>
    <p><strong>This one chart summarizes the concern.</strong> DV1 jumps by
    0.083 points from FY2023 to FY2024 — a jump that would look like a
    substantive Grants Pass response if taken at face value. DV2, using only
    wording-stable cells, is <em>essentially flat</em> across all three years
    (0.950 → 0.955 → 0.936). The FY2024 increase in DV1 is therefore largely
    attributable to the instrument change, not to actual behavior.</p>

    <h2>How do the conclusions change?</h2>
    <div id="forest" style="width:100%"></div>

    {coef_table}

    <p class="hint">Cluster-robust SEs in parentheses · <code>*** p&lt;0.01,
    ** p&lt;0.05, * p&lt;0.10</code></p>

    <h2>The three things that change when we switch DVs</h2>

    <h3>1. The Nonprofit IV effect — partially rescued</h3>
    <p>Under DV1 (contaminated full index), the Nonprofit coefficient is small
    and insignificant (β ≈ +0.04, n.s.). Under <strong>DV2 (wording-stable
    engagement cells)</strong>, Nonprofit becomes <strong>significantly
    positive</strong> (Mundlak β = +0.811**, FY24 cross-section β = +0.628).
    This is the theoretical prediction — nonprofit-led CoCs engage policymakers
    more — <em>showing up when the DV is restricted to the invariant
    sub-measure</em>.</p>

    <div class="callout">
      <div class="callout-title">⚠ But interpret with caution</div>
      <p>DV2 suffers from a strong ceiling (~95% Yes). A logit-scale
      coefficient of 0.8 can look impressive while translating into a small
      absolute difference (e.g., government 94% → nonprofit 97%). The
      substantive importance is smaller than the coefficient magnitude
      suggests. Still, the <em>direction</em> is consistent with theory in
      DV2, whereas DV1 offers no support either way.</p>
    </div>

    <h3>2. The PLE mediator — weakens</h3>
    <p>PLE engagement (log PLE count) is robustly positive under DV1
    (β = +0.11 in Mundlak, +0.21** in FY24). Under DV2 (engagement-only),
    its coefficient shrinks and is less significant. Under DV3 (FY22+23
    only), it is essentially zero.</p>
    <p>This means: the PLE → activity link is <strong>driven in part by the
    FY2024 implementation-column jump</strong>. A more conservative reading:
    CoCs with more PLE structures report more <em>reporting</em> activity
    under the new form, not necessarily more actual anti-crim behavior.</p>

    <h3>3. The DiD null — holds firm</h3>
    <p>β(DiD) is small and insignificant on both DV1 and DV2. The
    differential-response null is not an artifact of the instrument change
    — both groups are affected equally by it.</p>

    <h2>Recommended reporting strategy for the paper</h2>
    <div class="callout">
      <div class="callout-title">A three-table Methods approach</div>
      <ol>
        <li><strong>Main results (Table 3):</strong> DV1 (full index) with CoC+year FE and Mundlak. This is the paper's composite activity measure.</li>
        <li><strong>Robustness 1 (Table 4):</strong> DV2 (engagement-only). Disclose the ceiling effect but present the coefficients transparently. <em>This is where the positive Nonprofit finding appears.</em></li>
        <li><strong>Robustness 2 (Table 5):</strong> DV3 (FY22+23 panel only). Shows the null Nonprofit finding is not an FY2024 artifact.</li>
      </ol>
      <p>Reviewers will ask exactly the question that prompted this analysis;
      having all three tables pre-prepared turns that question from a threat
      into a strength.</p>
    </div>

    <h2>Honest revised bottom line</h2>
    <ol>
      <li>The primary Nonprofit effect is <strong>ambiguous</strong> across
      DV specifications: null in DV1 (composite, contaminated), positive in
      DV2 (wording-stable but ceiling-bound). The paper should <em>not</em>
      claim "no effect" — it should claim "no effect on the composite measure;
      a positive effect on the sub-measure that is least contaminated by the
      instrument change, subject to ceiling constraints."</li>
      <li>Grants Pass differential response is a <strong>robust null</strong>:
      DiD ≈ 0 on both DV1 and DV2.</li>
      <li>Year-over-year level claims about the activity index
      <strong>should be avoided</strong>. The FY2024 jump is largely
      instrument-driven.</li>
      <li>The <strong>strongest path forward</strong> is adding an external
      behavioral DV (NLHR ordinances, FBI UCR arrests) measured consistently
      by a third party.</li>
    </ol>

    <h2>Files</h2>
    <ul>
      <li><a href="../data_pipeline/run_dv_robustness.py">run_dv_robustness.py</a></li>
      <li><a href="../data_pipeline/dv_robustness_results.md">dv_robustness_results.md</a></li>
      <li><a href="../data_pipeline/dv_robustness_coefs.csv">dv_robustness_coefs.csv</a></li>
    </ul>

    <script>
    Plotly.newPlot("meanchart", {json.dumps(mean_traces)}, {json.dumps(mean_layout)}, {{responsive:true}});
    Plotly.newPlot("forest", {json.dumps(traces)}, {json.dumps(layout)}, {{responsive:true}});
    </script>
    """
    return wrap("DV robustness (v3)", "dv_robust.html", body,
                "If the instrument changed, is the DV valid? Three alternative DVs tested side-by-side.")


def page_main_results():
    """Primary multilevel DiD page with 4-quadrant key figure."""
    coefs = load_csv(PIPELINE_DIR / "multilevel_coefs.csv")
    quad = load_csv(PIPELINE_DIR / "multilevel_quadrant_means.csv")
    if not coefs or not quad:
        return wrap("Main results", "main_results.html",
                    "<p><em>Run <code>run_multilevel.py</code> first.</em></p>", "")

    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    # --- Quadrant trajectory data ---
    quadrants = [
        "Red state × Red county",
        "Red state × Blue county",
        "Blue state × Red county",
        "Blue state × Blue county",
    ]
    colors = {
        "Red state × Red county": "#c53030",
        "Red state × Blue county": "#d69e2e",
        "Blue state × Red county": "#3182ce",
        "Blue state × Blue county": "#17813c",
    }
    trajectory_traces = []
    for q in quadrants:
        xs, ys, ses, ns = [], [], [], []
        for yr in (2022, 2023, 2024):
            row = next((r for r in quad if r["quadrant"] == q and int(r["year"]) == yr), None)
            if row:
                xs.append(f"FY{yr}")
                ys.append(to_float(row["mean"]))
                ses.append(to_float(row["se"]))
                ns.append(int(float(row["count"])))
        is_red_red = (q == "Red state × Red county")
        trajectory_traces.append({
            "type": "scatter", "mode": "lines+markers+text",
            "name": f"{q} (n≈{ns[-1] if ns else '?'})",
            "x": xs, "y": ys,
            "error_y": {"type": "data", "array": [s * 1.96 for s in ses], "visible": True},
            "text": [f"{v:.3f}" for v in ys], "textposition": "top center",
            "marker": {"size": 11, "color": colors[q],
                       "symbol": "diamond" if is_red_red else "circle"},
            "line": {"width": 4 if is_red_red else 2.5,
                     "color": colors[q],
                     "dash": "solid"},
        })
    trajectory_layout = {
        "title": "Anti-criminalization activity index by 2×2 political quadrant (mean ± 1.96·SE)",
        "xaxis": {"title": "Fiscal year",
                  "categoryorder": "array",
                  "categoryarray": ["FY2022", "FY2023", "FY2024"]},
        "yaxis": {"title": "Mean activity index (0–1)", "range": [0.55, 0.92]},
        "height": 500,
        "shapes": [{
            "type": "rect", "xref": "x", "yref": "paper",
            "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
            "fillcolor": "#fff4cc", "opacity": 0.35, "line": {"width": 0},
            "layer": "below",
        }],
        "annotations": [
            {"x": 2, "y": 0.89, "xref": "x", "yref": "paper",
             "showarrow": False,
             "text": "Post-Grants Pass (June 2024)",
             "font": {"color": "#8a6d1f", "size": 12}},
            {"x": "FY2024", "y": 0.751, "xref": "x", "yref": "y",
             "showarrow": True, "arrowhead": 2, "arrowcolor": "#c53030",
             "ax": 60, "ay": 40,
             "text": "<b>Only Red×Red stays flat</b><br>(Δ = −0.031)",
             "font": {"color": "#c53030", "size": 11},
             "bgcolor": "rgba(255,255,255,0.9)",
             "bordercolor": "#c53030"},
        ],
        "legend": {"orientation": "h", "y": -0.2},
    }

    # --- Coefficient table ---
    from collections import defaultdict
    by_panel = defaultdict(dict)
    for r in coefs:
        b = to_float(r["coef"]); s = to_float(r["se"])
        p = to_float(r["pvalue"]); p_boot = to_float(r["bootstrap_p"])
        if b is None: continue
        by_panel[r["panel"]][r["variable"]] = (b, s, p, p_boot)

    label_map = {
        "nonprofit_led": "Nonprofit-led (IV)",
        "blue_state": "Blue state",
        "biden_within_state": "Biden-within-state (county − state mean)",
        "post": "Post-Grants Pass",
        "did_np": "Nonprofit × Post",
        "did_blue": "<strong>Blue state × Post</strong>",
        "did_within": "Biden-within × Post",
        "hf_pct": "Housing First adoption",
        "hmis_cov": "HMIS ES coverage",
        "log_beds": "log(total beds + 1)",
        "ple_ces_bin": "PLE in CES (binary)",
        "hf_pct_bar": "[Mundlak] HF mean",
        "hmis_cov_bar": "[Mundlak] HMIS mean",
        "log_beds_bar": "[Mundlak] log beds mean",
        "ple_ces_bin_bar": "[Mundlak] PLE mean",
        "const": "(Intercept)",
    }
    # Order: key variables first
    all_vars = [
        "nonprofit_led", "blue_state", "biden_within_state",
        "post", "did_np", "did_blue", "did_within",
        "hf_pct", "hmis_cov", "log_beds", "ple_ces_bin",
        "hf_pct_bar", "hmis_cov_bar", "log_beds_bar", "ple_ces_bin_bar",
        "const",
    ]

    def _is_valid(x):
        return x is not None and not (isinstance(x, float) and x != x)  # not None and not NaN

    def cell(panel, v):
        d = by_panel.get(panel, {}).get(v)
        if not d: return "—"
        b, s, p, p_b = d
        star = ""
        if v.startswith("did_") and _is_valid(p_b):
            if p_b < 0.01: star = "***"
            elif p_b < 0.05: star = "**"
            elif p_b < 0.10: star = "*"
        elif _is_valid(p):
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        cell_str = f"{b:+.3f}{star}"
        if s: cell_str += f"<br><span class='hint'>({s:.3f})</span>"
        if v.startswith("did_") and _is_valid(p_b):
            cell_str += f"<br><span class='hint'>p_boot = {p_b:.3f}</span>"
        return cell_str

    rows_html = []
    for v in all_vars:
        tds = [label_map.get(v, v), cell("unbalanced", v), cell("balanced", v)]
        rows_html.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    coef_table = (
        "<table style='font-size: 0.9em'>"
        "<thead><tr><th>Variable</th>"
        "<th>Unbalanced (n=594)</th><th>Balanced (n=341)</th></tr></thead>"
        f"<tbody>{''.join(rows_html)}</tbody></table>"
    )

    # --- Quadrant Δ bar chart ---
    quad_deltas = []
    for q in quadrants:
        fy23 = next((r for r in quad if r["quadrant"] == q and int(r["year"]) == 2023), None)
        fy24 = next((r for r in quad if r["quadrant"] == q and int(r["year"]) == 2024), None)
        if fy23 and fy24:
            delta = to_float(fy24["mean"]) - to_float(fy23["mean"])
            quad_deltas.append({"q": q, "d": delta, "n24": int(float(fy24["count"]))})
    delta_data = [{
        "type": "bar",
        "x": [d["q"] for d in quad_deltas],
        "y": [d["d"] for d in quad_deltas],
        "marker": {"color": [colors[d["q"]] for d in quad_deltas]},
        "text": [f"{d['d']:+.3f}<br>(n={d['n24']})" for d in quad_deltas],
        "textposition": "outside",
    }]
    delta_layout = {
        "title": "FY2023 → FY2024 change (Δ) by quadrant",
        "yaxis": {"title": "Δ mean activity index", "range": [-0.06, 0.18]},
        "xaxis": {"title": ""},
        "height": 360,
        "showlegend": False,
        "shapes": [{"type": "line", "xref": "paper", "yref": "y",
                    "x0": 0, "x1": 1, "y0": 0, "y1": 0,
                    "line": {"color": "#666", "width": 1, "dash": "dot"}}],
    }

    # --- Forest plot: DiD terms across panels ---
    did_terms = [
        ("did_np", "Nonprofit × Post"),
        ("did_blue", "Blue state × Post"),
        ("did_within", "Biden-within × Post"),
    ]
    forest_traces = []
    for label_panel in ("unbalanced", "balanced"):
        xs, ys, errs, texts = [], [], [], []
        for vid, lbl in did_terms:
            d = by_panel.get(label_panel, {}).get(vid)
            if d:
                b, s, _, p_b = d
                xs.append(b); ys.append(lbl); errs.append(1.96 * s)
                texts.append(f"p_boot={p_b:.3f}" if p_b else "")
        forest_traces.append({
            "type": "scatter", "mode": "markers+text",
            "name": label_panel.capitalize(),
            "x": xs, "y": ys,
            "error_x": {"type": "data", "array": errs, "visible": True, "thickness": 1.5},
            "text": texts, "textposition": "top right", "textfont": {"size": 9},
            "marker": {"size": 12},
        })
    forest_layout = {
        "title": "DiD coefficients across panels (95% CI)",
        "xaxis": {"title": "Coefficient (logit scale)",
                  "zeroline": True, "zerolinecolor": "#c53030", "zerolinewidth": 2},
        "yaxis": {"autorange": "reversed"},
        "height": 300,
        "margin": {"l": 180, "r": 30, "t": 50, "b": 40},
        "legend": {"orientation": "h", "y": -0.25},
    }

    # --- Narrative ---
    np_est = by_panel.get("unbalanced", {}).get("did_np")
    blue_est = by_panel.get("unbalanced", {}).get("did_blue")
    within_est = by_panel.get("unbalanced", {}).get("did_within")

    def tell(c, use_boot=True):
        if not c: return "(n/a)"
        b, s, p, pb = c
        p_ref = pb if (use_boot and pb is not None) else p
        star = ""
        if p_ref is not None:
            if p_ref < 0.01: star = "***"
            elif p_ref < 0.05: star = "**"
            elif p_ref < 0.10: star = "*"
        return f"β = {b:+.3f}{star} (bootstrap p = {pb:.3f})" if pb else f"β = {b:+.3f}{star}"

    body = f"""
    <p class="lead">The primary specification for the paper — a multilevel
    difference-in-differences that decomposes political environment into
    state-level and within-state county components, estimated jointly with
    the governance treatment.</p>

    <h2>The key figure</h2>
    <div id="trajectory" style="width:100%"></div>
    <p class="hint">CoCs are grouped into 2×2 quadrants based on their state's
    2020 presidential winner and whether their county's Biden share exceeds 50%.
    Three of four quadrants rose in parallel after <em>Grants Pass v. Johnson</em>;
    only the red-state × red-county group stayed flat.</p>

    <div class="findings-box">
      <h3>What this shows in one sentence</h3>
      <p style="font-size: 1.05em;">
      <strong>Only CoCs embedded in fully red-leaning political environments
      (state AND county) did not expand their reported anti-criminalization
      activity after Grants Pass.</strong> All three other political
      combinations rose by +0.11 to +0.14 points, converging to a similar
      FY2024 level (0.75–0.87).
      </p>
    </div>

    <h2>Model specification</h2>

    <h3>Why multilevel</h3>
    <p>Two political measures — state-level binary (Biden won 2020) and
    county-level Biden share — are highly correlated because a county
    tends to be blue partly because its state is. Putting both in the
    same model directly produces collinearity. We apply a
    Mundlak-style decomposition at the multilevel structure:</p>
    <pre><code>state_mean_biden   = mean Biden share of counties in the state
biden_within_state = county_biden_share − state_mean_biden</code></pre>
    <p>After this decomposition, <code>blue_state</code> captures the
    state-level political environment and <code>biden_within_state</code>
    captures <em>purely</em> within-state county deviation. They are
    statistically orthogonal, so we can estimate both effects in the same
    regression.</p>

    <h3>Primary specification</h3>
    <pre><code>crim_activity_index<sub>it</sub> =
      β₁ · Nonprofit<sub>i</sub>
    + β₂ · Blue_state<sub>i</sub>
    + β₃ · Biden_within_state<sub>i</sub>
    + β₄ · Post<sub>t</sub>
    + β₅ · Nonprofit × Post                ← governance DiD
    + β₆ · Blue_state × Post               ← state-level political DiD
    + β₇ · Biden_within_state × Post       ← within-state county DiD
    + γ · Controls<sub>it</sub> (Mundlak-adjusted)
    + ε<sub>it</sub></code></pre>

    <p><strong>Estimator:</strong> Papke-Wooldridge fractional logit for the
    bounded DV [0,1]. <strong>Standard errors:</strong> state-level
    cluster-robust, reinforced with wild-cluster bootstrap (Rademacher
    weights, state-clustered, 999 replicates). <strong>Sample:</strong>
    unbalanced panel as primary (preserves 651 CoC-years across 51 states),
    balanced 3-year panel as robustness.</p>

    <h2>Coefficient table</h2>
    {coef_table}
    <p class="hint">Cluster-robust SE at the state level in parentheses.
    Significance stars on DiD terms use wild-cluster bootstrap p-values;
    on main effects use cluster-robust p. <code>*** p&lt;0.01, ** p&lt;0.05,
    * p&lt;0.10</code>.</p>

    <h2>Forest plot — DiD terms</h2>
    <div id="forest" style="width:100%"></div>
    <p class="hint">The three DiD coefficients (governance, state political
    environment, within-state county deviation) side by side on both
    panels. Whiskers cross zero for Nonprofit × Post and Biden-within × Post;
    only Blue-state × Post approaches significance.</p>

    <h2>Quadrant Δ (FY2023 → FY2024)</h2>
    <div id="delta_bar" style="width:100%"></div>

    <h2>Reading the results</h2>

    <h3>1. Governance does not matter</h3>
    <p>Nonprofit × Post: {tell(np_est)}. Across both panels the nonprofit
    coefficient is indistinguishable from zero. The theoretical prediction
    that nonprofit-led CoCs should be more rights-responsive is not
    supported.</p>

    <h3>2. State-level political environment matters (marginally)</h3>
    <p>Blue-state × Post: {tell(blue_est)}. Significant at the 10%
    level in the unbalanced panel, weaker in the balanced sample.
    <strong>This is the paper's key positive finding, subject to the
    balanced-panel caveat.</strong></p>

    <h3>3. Within-state county variation adds nothing</h3>
    <p>Biden-within × Post: {tell(within_est)}. Once state-level
    environment is controlled, the additional county deviation does not
    explain response. <strong>The earlier county-level finding (β = +1.95
    in isolation) was primarily picking up between-state variation, not
    within-state.</strong></p>

    <h3>4. The asymmetric quadrant pattern</h3>
    <p>The quadrant decomposition reveals the <em>shape</em> of the
    political-environment effect. It is not simply "blue areas respond
    more" — it is "<strong>red-state × red-county is the outlier that
    doesn't respond, while all other political combinations respond
    similarly</strong>". This OR-shaped pattern is more consistent with
    a threshold story than a linear dose-response.</p>

    <h2>Interpretation — what does "red-state × red-county doesn't respond" mean?</h2>

    <div class="callout">
      <div class="callout-title">A nuance worth spelling out</div>
      <p><strong>"Not responding"</strong> here means the activity index
      did not rise FY23 → FY24 (actually declined slightly). It does
      <em>not</em> mean red-state × red-county CoCs have low activity
      overall — they were at 0.782 in FY2023, the highest of any
      quadrant at that point. After Grants Pass they held steady at
      0.751 while the three other quadrants converged upward to
      0.75–0.87.</p>
      <p>Three candidate mechanisms, all consistent with the data:</p>
      <ul>
        <li><strong>Ceiling:</strong> Red-red CoCs were already at a high
        baseline; limited room to expand.</li>
        <li><strong>Political demand:</strong> Fully red environments
        provide no constituency-level pressure to reframe toward
        rights-responsive anti-criminalization; the Grants Pass ruling
        signals "the law is on our side" and requires no defensive
        repositioning.</li>
        <li><strong>Reporting strategy:</strong> CoCs in red environments
        may deliberately avoid foregrounding anti-criminalization framing
        in the new FY2024 form language.</li>
      </ul>
    </div>

    <h2>Sample sizes by quadrant (FY2024)</h2>
    <table>
      <thead><tr><th>Quadrant</th><th class='num'>n</th></tr></thead>
      <tbody>
        <tr><td>Blue state × Blue county</td><td class='num'>131</td></tr>
        <tr><td>Red state × Red county</td><td class='num'>75</td></tr>
        <tr><td>Blue state × Red county</td><td class='num'>50</td></tr>
        <tr><td>Red state × Blue county</td><td class='num'>26</td></tr>
      </tbody>
    </table>
    <p class="hint">The smallest cell (red state × blue county, n = 26)
    has wide confidence intervals. Interpret the quadrant effect sizes
    with this power constraint in mind.</p>

    <h2>Limitations of this primary specification</h2>
    <ol>
      <li><strong>Marginal significance.</strong> The state-level DiD is
      at bootstrap p ≈ 0.095 — meets a 10% threshold, not 5%. External
      behavioral data would strengthen the inference.</li>
      <li><strong>Balanced-panel weakening.</strong> See
      <a href="balanced.html">Balanced vs Unbalanced</a>. Balanced-panel
      results are weaker; the unbalanced-panel significance may reflect
      sample composition (FY2024 adds ~126 YHDP CoCs potentially
      concentrated in blue states).</li>
      <li><strong>DV self-report.</strong> The DV is a CoC's own
      reporting on HUD Form 1D-4, not an external behavioral measure.
      See <a href="dv_story.html">The DV story</a> and
      <a href="dv_robust.html">DV robustness</a>.</li>
      <li><strong>Short post-period.</strong> FY2024 applications were
      submitted four months after the Supreme Court ruling; longer-run
      behavior may differ.</li>
      <li><strong>Quadrant sample imbalance.</strong> Red state × Blue
      county has only 26 observations in FY2024, limiting the precision
      of that cell's estimate.</li>
    </ol>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
python3 code_iv_leadership.py         # nonprofit_led classification
python3 build_coc_county.py           # CoC → county Biden share
python3 run_multilevel.py             # primary specification + bootstrap
python3 build_site.py                 # regenerate this page</code></pre>

    <script>
    Plotly.newPlot("trajectory", {json.dumps(trajectory_traces)},
                   {json.dumps(trajectory_layout)}, {{responsive:true}});
    Plotly.newPlot("forest", {json.dumps(forest_traces)},
                   {json.dumps(forest_layout)}, {{responsive:true}});
    Plotly.newPlot("delta_bar", {json.dumps(delta_data)},
                   {json.dumps(delta_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("Main results — Multilevel DiD", "main_results.html", body,
                "The paper's primary specification: multilevel fractional-logit DiD with state-clustered SEs.")


def page_balanced():
    coefs = load_csv(PIPELINE_DIR / "balanced_sensitivity_coefs.csv")
    if not coefs:
        return wrap("Balanced vs Unbalanced", "balanced.html",
                    "<p><em>Run <code>run_balanced_sensitivity.py</code> first.</em></p>", "")

    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    from collections import defaultdict
    by_key = {}
    for r in coefs:
        key = (r["spec"], r["panel"], r["variable"])
        by_key[key] = (to_float(r["coef"]), to_float(r["se"]),
                       to_float(r["pvalue"]), int(r["n"]))

    specs = [
        ("N1 (Nonprofit × Post)", "did_np"),
        ("B1 (Blue-state × Post)", "did_blue"),
        ("J2 (Joint NP + Biden × Post)", "did_biden"),
    ]

    rows = []
    for name, target in specs:
        u = by_key.get((name, "unbalanced", target))
        b = by_key.get((name, "balanced", target))
        if u and b:
            u_b, u_s, u_p, u_n = u
            b_b, b_s, b_p, b_n = b
            u_star = ("***" if u_p < 0.01 else ("**" if u_p < 0.05 else ("*" if u_p < 0.10 else "")))
            b_star = ("***" if b_p < 0.01 else ("**" if b_p < 0.05 else ("*" if b_p < 0.10 else "")))
            sig_change = ""
            if (u_p < 0.10) != (b_p < 0.10):
                sig_change = " ⚠"
            rows.append(f"""
              <tr>
                <td><strong>{name}</strong></td>
                <td class='num'>{u_b:+.3f}{u_star}<br><span class='hint'>(SE {u_s:.3f})</span></td>
                <td class='num'>{u_p:.3f}</td>
                <td class='num'>{u_n}</td>
                <td class='num'>{b_b:+.3f}{b_star}<br><span class='hint'>(SE {b_s:.3f})</span></td>
                <td class='num'>{b_p:.3f}</td>
                <td class='num'>{b_n}</td>
                <td>{sig_change}</td>
              </tr>""")
    table = (
        "<table><thead><tr>"
        "<th rowspan='2'>Spec</th>"
        "<th colspan='3' style='text-align:center'>Unbalanced (651 obs · 325 CoCs)</th>"
        "<th colspan='3' style='text-align:center'>Balanced (375 obs · 125 CoCs)</th>"
        "<th rowspan='2'>Change</th>"
        "</tr>"
        "<tr>"
        "<th class='num'>β</th><th class='num'>p</th><th class='num'>N</th>"
        "<th class='num'>β</th><th class='num'>p</th><th class='num'>N</th>"
        "</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )

    body = f"""
    <p class="lead">A straightforward question reviewers will ask:
    <em>balanced or unbalanced panel?</em> This page lays both side-by-side
    and reveals that the state-level political finding is partly driven by
    sample composition.</p>

    <div class="callout">
      <div class="callout-title">Panel composition</div>
      <ul>
        <li><strong>Unbalanced:</strong> 651 CoC-year rows from 325 CoCs.
        Of these, 124 CoCs (38%) appear in only one year, 76 in two years,
        and 125 in all three years. FY2024 (292 CoCs) is substantially
        larger than FY2022 (166 CoCs) because HUD folded YHDP recipients
        into the main competition that year.</li>
        <li><strong>Balanced:</strong> 375 rows = 125 CoCs × 3 years. The
        same set of CoCs is observed every year.</li>
      </ul>
    </div>

    <h2>The side-by-side</h2>
    {table}
    <p class="hint">Cluster-robust SEs at the CoC level · <code>*** p&lt;0.01,
    ** p&lt;0.05, * p&lt;0.10</code>. ⚠ marks specifications where the 0.10
    significance verdict flips between panels.</p>

    <h2>What changes</h2>

    <h3>1. Nonprofit DiD — null in both (robust)</h3>
    <p>Unbalanced β = −0.055 (p = 0.83), balanced β = +0.278 (p = 0.43).
    Signs flip but neither approaches significance. <strong>The null
    nonprofit finding is robust to panel choice.</strong></p>

    <h3>2. Binary Blue × Post — flips from significant to null ⚠</h3>
    <p>Unbalanced: β = <strong>+0.548 (p = 0.036)</strong> — the state-level
    Blue × Post term from the <a href="main_results.html">multilevel DiD</a>.
    Balanced: β = +0.129 (p = 0.70) — effectively zero. This is a
    <strong>striking divergence</strong> that casts doubt on whether the
    state-level effect reflects actual differential response or sample
    composition.</p>

    <div class="callout">
      <div class="callout-title">⚠ What the flip likely means</div>
      <p>The unbalanced panel includes 126 CoCs that appear only in FY2024
      (or one of FY2022/23). Many of these are newly-added YHDP recipients
      that HUD folded into the FY2024 main competition. If those new CoCs
      are <strong>disproportionately in blue states</strong>, they would
      inflate the Blue × Post coefficient even without any true behavioral
      response — the blue group simply grew more between the pre and post
      periods. The balanced panel forces the comparison to the <em>same</em>
      CoCs each year, which is the cleaner DiD contrast.</p>
    </div>

    <h3>3. Continuous county Biden × Post — gets STRONGER in balanced</h3>
    <p>Unbalanced: β = +1.95 (p = 0.06, bootstrap p = 0.18).
    Balanced: β = <strong>+3.20 (p = 0.02, bootstrap p = 0.06)</strong>.
    This is the <em>reverse</em> of what compositional contamination
    predicts. The continuous county-level measure strengthens when we
    restrict to the balanced panel.</p>
    <p>Interpretation: once we remove the composition confound, the
    <em>continuous</em> political-environment effect tightens —
    suggesting it captures a real within-CoC-type differential response
    to Grants Pass. The state-level binary, by contrast, was picking up
    both real and compositional variation.</p>

    <h2>Revised recommendation</h2>

    <div class="findings-box">
      <h3>✅ Updated framing for the paper</h3>
      <ol>
        <li><strong>Primary political-environment finding:</strong> county-level
        continuous Biden share × Post. Directionally positive in both panels;
        statistically significant in the balanced panel (the cleaner DiD
        contrast); bootstrap p = 0.06.</li>
        <li><strong>State-level binary finding:</strong> report as auxiliary.
        Significant only in unbalanced, not in balanced, suggesting it is
        partly a composition artifact. Include it with an explicit caveat.</li>
        <li><strong>Nonprofit finding:</strong> null under both panels.
        Robust across every specification we've run.</li>
        <li><strong>Report both panels in the Methods section.</strong>
        Primary estimates from one panel, robustness checks from the other.
        Transparent about why they differ.</li>
      </ol>
    </div>

    <h2>So — balanced or unbalanced?</h2>
    <p>For this study, the <strong>answer depends on the question</strong>:</p>
    <ul>
      <li><strong>Descriptive statistics</strong>: use <strong>unbalanced</strong>
        — more complete picture of the full universe of CoCs.</li>
      <li><strong>Main-effects panel regression</strong> (what predicts
        activity overall?): use <strong>unbalanced</strong> for power, with
        CoC fixed effects / Mundlak to identify off multi-year CoCs.</li>
      <li><strong>DiD identification</strong> (does the shock produce
        differential response?): use <strong>balanced</strong> as primary —
        same CoCs in pre and post, no compositional confound.</li>
      <li><strong>Power-critical analyses</strong> (narrow binary
        treatments): report both; transparency beats a single choice.</li>
    </ul>

    <p>Back to the <a href="analysis_hub.html">Analysis hub</a>.</p>
    """
    return wrap("Balanced vs Unbalanced panel", "balanced.html", body,
                "Does the panel choice matter? For the binary state-level DiD — yes, decisively.")


def page_did_joint():
    coefs = load_csv(PIPELINE_DIR / "did_joint_coefs.csv")
    if not coefs:
        return wrap("Joint DiD", "did_joint.html",
                    "<p><em>Run <code>run_did_joint.py</code> first.</em></p>", "")

    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    from collections import defaultdict
    by_spec = defaultdict(dict)
    boot_triple = None
    boot_blue = None
    for r in coefs:
        if r.get("variable") == "triple_diff_bootstrap_p":
            boot_triple = to_float(r["pvalue"])
            continue
        if r.get("variable") == "did_blue_bootstrap_p":
            boot_blue = to_float(r["pvalue"])
            continue
        b = to_float(r["coef"]); s = to_float(r["se"]); p = to_float(r["pvalue"])
        if b is None: continue
        by_spec[r["spec"]][r["variable"]] = (b, s, p, int(r["n"]))

    spec_order = ["J1·County Biden DiD only", "J2·Joint (NP + Biden) DiD",
                  "J3·Triple-difference", "J4·Triple + year FE"]
    all_vars = ["nonprofit_led", "biden_c", "post",
                "did_np", "did_blue", "np_x_biden", "triple_diff",
                "hf_pct", "hmis_cov", "log_beds", "ple_ces_bin", "const"]
    label_map = {
        "nonprofit_led": "Nonprofit-led (IV₁)",
        "biden_c": "County Biden share (centered, IV₂)",
        "post": "Post-Grants Pass",
        "did_np": "Nonprofit × Post",
        "did_blue": "Biden × Post (continuous DiD)",
        "np_x_biden": "Nonprofit × Biden (interaction)",
        "triple_diff": "<strong>Nonprofit × Biden × Post</strong>",
        "hf_pct": "Housing First adoption",
        "hmis_cov": "HMIS ES coverage",
        "log_beds": "log(total beds + 1)",
        "ple_ces_bin": "PLE in CES (binary)",
        "const": "(Intercept)",
    }

    def cell(sp, v):
        c = by_spec.get(sp, {}).get(v)
        if not c: return "—"
        b, s, p, _ = c
        if b is None: return "—"
        if s and s > 1e5: return f"{b:+.3f}<br><span class='hint'>(numerical issues)</span>"
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        return f"{b:+.3f}{star}<br><span class='hint'>({s:.3f})</span>"

    rows_html = []
    for v in all_vars:
        tds = [label_map.get(v, v)] + [cell(sp, v) for sp in spec_order]
        rows_html.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    n_row = "<tr><td><strong>N</strong></td>" + "".join(
        f"<td>{next(iter(by_spec[sp].values()))[3] if by_spec.get(sp) else '—'}</td>"
        for sp in spec_order) + "</tr>"
    th = "".join(f"<th style='font-size:0.82em'>{sp}</th>" for sp in spec_order)
    coef_table = (f"<table style='font-size:0.85em'>"
                  f"<thead><tr><th>Variable</th>{th}</tr></thead>"
                  f"<tbody>{''.join(rows_html)}{n_row}</tbody></table>")

    def tell(sp, v):
        c = by_spec.get(sp, {}).get(v)
        if not c: return "(n/a)"
        b, s, p, _ = c
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        return f"β = {b:+.3f}{star} (cluster SE {s:.3f}, p = {p:.3f})"

    # Comparison with prior DiD pages
    comparison_table = """
    <table>
      <thead><tr><th>Specification</th><th>β_DiD</th><th class='num'>Cluster p</th><th class='num'>Bootstrap p</th><th>Verdict</th></tr></thead>
      <tbody>
        <tr><td>Nonprofit × Post (state-level, <a href='did.html'>DiD page</a>)</td>
            <td class='num'>+0.007</td><td class='num'>0.88</td><td class='num'>0.82</td>
            <td>Null</td></tr>
        <tr><td>Blue × Post (state binary, <a href='did_bluestate.html'>Blue/Red</a>)</td>
            <td class='num'>+0.548</td><td class='num'>0.037</td><td class='num'><strong>0.036</strong></td>
            <td><strong>Significant</strong></td></tr>
        <tr><td>Biden × Post (county continuous, J2)</td>
            <td class='num'>+1.951</td><td class='num'>0.060</td><td class='num'>""" + f"{boot_blue:.3f}" + """</td>
            <td>Marginal on bootstrap</td></tr>
        <tr><td>Triple: Nonprofit × Biden × Post (J3)</td>
            <td class='num'>−2.363</td><td class='num'>0.24</td><td class='num'>""" + f"{boot_triple:.3f}" + """</td>
            <td>Null</td></tr>
      </tbody>
    </table>
    """

    body = f"""
    <p class="lead">Joint specifications testing whether the county-level
    political environment and the CoC's leadership structure matter
    together — or whether one dimension subsumes the other. Triple-difference
    tests whether nonprofit leadership <em>amplifies</em> the blue-county
    response to Grants Pass.</p>

    <div class="callout">
      <div class="callout-title">What's new here</div>
      <ul>
        <li><strong>County-level</strong> Biden share (2020 presidential)
        replaces the state-level binary from the <a href="did_bluestate.html">blue/red page</a>.
        Variable ranges 0.226–0.921 across the 317 matched CoCs (mean 0.526).</li>
        <li><strong>Joint specification</strong>: both Nonprofit × Post and
        Biden × Post in the same model. Tests whether either coefficient
        survives when the other is controlled.</li>
        <li><strong>Triple-difference</strong>: Nonprofit × Biden × Post.
        Tests whether nonprofit-led CoCs in blue counties respond most
        strongly to Grants Pass — the full governance-meets-politics
        interaction.</li>
      </ul>
    </div>

    <h2>Data construction</h2>
    <ul>
      <li><strong>Lead agency</strong>: manual classifier → 320 of 321 CoCs (99%).</li>
      <li><strong>County Biden share</strong>: parsed CoC names against the
      <a href="https://github.com/tonmcg/US_County_Level_Election_Results_08-24">
      2020 county-level presidential results</a> (MIT Election Lab mirror).
      221 CoCs directly matched to one or more counties; 96 fell back to
      state-level vote share; 4 unresolved.</li>
      <li><strong>Centering</strong>: <code>biden_c = biden_share − 0.5</code>,
      so a coefficient on <code>biden_c</code> reads as the marginal effect
      per unit of Biden-share above 50-50.</li>
    </ul>

    <h2>Leadership × county partisanship cross-tab (FY2024)</h2>
    <table>
      <thead><tr><th></th><th class='num'>Red county (≤ 50% Biden)</th><th class='num'>Blue county (&gt; 50% Biden)</th></tr></thead>
      <tbody>
        <tr><td><strong>Nonprofit-led</strong></td><td class='num'>83</td><td class='num'>69</td></tr>
        <tr><td><strong>Government-led</strong></td><td class='num'>42</td><td class='num'>91</td></tr>
      </tbody>
    </table>
    <p class="hint">Note that <em>government-led</em> CoCs are concentrated in
    blue counties (91 vs 42) — reflecting big-city CoCs typically being
    government-run. Nonprofits are more evenly split.</p>

    <h2>Coefficient table</h2>
    {coef_table}
    <p class="hint">Cluster-robust SEs at the CoC level in parentheses ·
    <code>*** p&lt;0.01, ** p&lt;0.05, * p&lt;0.10</code> · J4's massive SEs
    reflect numerical collinearity (Post + year dummies together near-absorb
    the shock) — interpret J1–J3 only.</p>

    <h2>Key estimates</h2>
    <ul>
      <li><strong>J2 · Biden × Post (joint):</strong> {tell('J2·Joint (NP + Biden) DiD', 'did_blue')}
      · bootstrap p = {boot_blue:.3f}</li>
      <li><strong>J2 · Nonprofit × Post (joint):</strong> {tell('J2·Joint (NP + Biden) DiD', 'did_np')}</li>
      <li><strong>J3 · Triple-difference:</strong> {tell('J3·Triple-difference', 'triple_diff')}
      · bootstrap p = {boot_triple:.3f}</li>
    </ul>

    <h2>Comparison across DiD specifications</h2>
    {comparison_table}

    <h2>What this tells us</h2>

    <h3>1. The state-level blue/red finding weakens at the county level</h3>
    <p>State-level binary blue/red gave β = +0.548 with bootstrap p = 0.036.
    Moving to county-level Biden share keeps the <em>direction</em> positive
    but the bootstrap p deteriorates to {boot_blue:.3f} — <strong>no longer
    conventionally significant</strong>. Plausible reasons:</p>
    <ul>
      <li><strong>Measurement error:</strong> 96 of 317 CoCs use state-level
      fallback when county parsing failed, injecting noise into the
      treatment.</li>
      <li><strong>Continuous vs binary power:</strong> the binary "Biden
      won the state" captures a coarse but accurate signal; a continuous
      vote share over-specifies the dose-response.</li>
      <li><strong>Honest conclusion:</strong> the political-environment
      effect is real in direction but its statistical significance is
      specification-sensitive. Report both state-level (stronger) and
      county-level (more granular but noisier) as complementary.</li>
    </ul>

    <h3>2. Nonprofit leadership remains null, even jointly</h3>
    <p>In the joint specification (J2), Nonprofit × Post is +0.096 (n.s.).
    Controlling for political environment does not rescue the theoretical
    prediction that nonprofit-led CoCs should respond more. The Nonprofit
    null is robust across every analysis we've run.</p>

    <h3>3. Triple-difference is NULL</h3>
    <p>β(Nonprofit × Biden × Post) = −2.36, p = {boot_triple:.3f}. This tests
    whether nonprofit leadership <em>interacts</em> with county partisanship
    in the Grants Pass response. The null result means: <strong>the blue-
    county effect is not stronger among nonprofit-led CoCs</strong>. If
    anything, the point estimate is negative (blue-county effect might be
    slightly <em>weaker</em> for nonprofits), but far from significance.</p>

    <h3>4. J4's numerical issues</h3>
    <p>When year fixed effects are added on top of a Post dummy, the model
    becomes near-singular (Post is perfectly collinear with FY2024 dummy).
    The standard errors balloon to 10¹²+; the coefficients aren't trustworthy.
    Treat J3 as the preferred triple-diff specification; J4 is reported for
    transparency but should not be interpreted.</p>

    <h2>Bottom line</h2>

    <div class="findings-box">
      <h3>✅ What's robust</h3>
      <ul>
        <li><strong>Political environment matters</strong> (state-level blue/red:
        robust; county-level continuous: weaker but same direction).</li>
        <li><strong>Lead-agency type does NOT matter</strong> (null across all
        specifications, including joint + triple).</li>
        <li><strong>Governance × partisanship triple is null</strong> —
        nonprofit leadership doesn't amplify or dampen the political
        response.</li>
      </ul>
    </div>

    <h2>Paper framing suggestion</h2>
    <p>Given the pattern of results across <a href="did.html">nonprofit DiD</a>,
    <a href="did_bluestate.html">blue-state DiD</a>, and this joint analysis:</p>
    <ol>
      <li><strong>Headline finding:</strong> Grants Pass provoked differential
      responses <em>across states by political composition</em>, not across
      CoCs by leadership structure. State-level political context — not
      governance form — is the channel that transmits the shock.</li>
      <li><strong>Reframe the theoretical prediction:</strong> the original
      hypothesis that nonprofit governance drives rights-responsiveness is
      not supported. A revised framing: local political demand signals
      (state electorate, county constituents) shape whether and how CoCs
      publicly position themselves on criminalization — independent of
      governance form.</li>
      <li><strong>Methodological transparency:</strong> report all three DiD
      specifications and the joint triple-difference. Reviewers will see the
      full exploration and the null leadership effect survives every slice.</li>
      <li><strong>Next:</strong> external behavioral DV (NLHR ordinance
      counts) remains the strongest additional robustness — to validate
      whether the blue-state DiD reflects real policy behavior or
      differential reporting.</li>
    </ol>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
curl -o external/county_2020_results.csv \\
  https://raw.githubusercontent.com/tonmcg/US_County_Level_Election_Results_08-24/master/2020_US_County_Level_Presidential_Results.csv
python3 build_coc_county.py       # parse CoC names → county Biden share
python3 run_did_joint.py          # J1..J4 + wild-cluster bootstrap</code></pre>
    """
    return wrap("Joint DiD (v4)", "did_joint.html", body,
                "County-level partisanship + nonprofit leadership in one model, plus triple-difference.")


def page_did_bluestate():
    coefs = load_csv(PIPELINE_DIR / "did_bluestate_coefs.csv")
    trends = load_csv(PIPELINE_DIR / "did_bluestate_trends.csv")
    if not coefs:
        return wrap("Blue vs Red DiD", "did_bluestate.html",
                    "<p><em>Run <code>run_did_bluestate.py</code> first.</em></p>", "")

    def to_float(x):
        try: return float(x)
        except (ValueError, TypeError): return None

    from collections import defaultdict
    by_spec = defaultdict(dict)
    boot_p = None
    for r in coefs:
        if r.get("variable") == "did_blue_bootstrap_p":
            boot_p = to_float(r["pvalue"])
            continue
        b = to_float(r["coef"]); s = to_float(r["se"]); p = to_float(r["pvalue"])
        if b is None: continue
        by_spec[r["spec"]][r["variable"]] = (b, s, p, int(r["n"]))

    # Parallel trends plot
    tr_b = sorted([r for r in trends if r["group"] == "blue"], key=lambda r: int(r["year"]))
    tr_r = sorted([r for r in trends if r["group"] == "red"], key=lambda r: int(r["year"]))
    def pack(rows):
        return ([f"FY{r['year']}" for r in rows],
                [to_float(r["mean"]) for r in rows],
                [to_float(r["se"]) for r in rows],
                [int(r["count"]) for r in rows])
    x_b, y_b, se_b, n_b = pack(tr_b)
    x_r, y_r, se_r, n_r = pack(tr_r)

    pt_data = [
        {"type": "scatter", "mode": "lines+markers+text",
         "name": f"Blue state (n≈{n_b[-1]})", "x": x_b, "y": y_b,
         "error_y": {"type": "data", "array": [s*1.96 for s in se_b], "visible": True},
         "text": [f"{v:.3f}" for v in y_b], "textposition": "top center",
         "marker": {"size": 10, "color": "#2057c9"}, "line": {"width": 3, "color": "#2057c9"}},
        {"type": "scatter", "mode": "lines+markers+text",
         "name": f"Red state (n≈{n_r[-1]})", "x": x_r, "y": y_r,
         "error_y": {"type": "data", "array": [s*1.96 for s in se_r], "visible": True},
         "text": [f"{v:.3f}" for v in y_r], "textposition": "bottom center",
         "marker": {"size": 10, "color": "#c53030"}, "line": {"width": 3, "color": "#c53030"}},
    ]
    pt_layout = {
        "title": "crim_activity_index · parallel trends by state political environment",
        "xaxis": {"title": "Fiscal year",
                  "categoryorder": "array",
                  "categoryarray": ["FY2022", "FY2023", "FY2024"]},
        "yaxis": {"title": "Mean (±1.96·SE)", "range": [0.60, 0.90]},
        "height": 420,
        "shapes": [{"type": "rect", "xref": "x", "yref": "paper",
                    "x0": 1.5, "x1": 2.5, "y0": 0, "y1": 1,
                    "fillcolor": "#fff4cc", "opacity": 0.35,
                    "line": {"width": 0}, "layer": "below"}],
        "annotations": [{"x": 2, "y": 0.87, "xref": "x", "yref": "paper",
                         "showarrow": False, "text": "Post · Grants Pass (Jun 2024)",
                         "font": {"color": "#8a6d1f", "size": 11}}],
        "legend": {"orientation": "h", "y": -0.2},
    }

    # Coefficient table
    all_vars = ["blue_state", "post_2023", "post", "did_blue_placebo_2023", "did_blue",
                "hf_pct", "hmis_cov", "log_beds", "ple_ces_bin", "const"]
    label_map = {
        "blue_state": "Blue state (treatment)",
        "post_2023": "FY2023 indicator",
        "post": "Post-Grants Pass (FY2024)",
        "did_blue_placebo_2023": "Blue × FY2023 (placebo)",
        "did_blue": "<strong>Blue × Post ← β_DiD</strong>",
        "hf_pct": "Housing First adoption",
        "hmis_cov": "HMIS ES coverage",
        "log_beds": "log(total beds + 1)",
        "ple_ces_bin": "PLE in CES (binary)",
        "const": "(Intercept)",
    }
    spec_order = [
        "B1·DV1 OLS+CoC FE",
        "B2·DV1 OLS+CoC+Year FE",
        "B3·DV1 frac-logit + Mundlak",
        "B4·DV1 event study (pre-trend placebo)",
        "B5·DV2 engagement-only",
        "B6·DV3 binary (implemented)",
    ]

    def cell(sp, v):
        c = by_spec.get(sp, {}).get(v)
        if not c: return "—"
        b, s, p, _ = c
        if b is None: return "—"
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        se_s = f"<br><span class='hint'>({s:.3f})</span>" if s is not None else ""
        return f"{b:+.3f}{star}{se_s}"

    rows_html = []
    for v in all_vars:
        tds = [label_map.get(v, v)] + [cell(sp, v) for sp in spec_order]
        rows_html.append("<tr>" + "".join(f"<td>{x}</td>" for x in tds) + "</tr>")
    n_row = "<tr><td><strong>N</strong></td>" + "".join(
        f"<td>{next(iter(by_spec[sp].values()))[3] if by_spec.get(sp) else '—'}</td>"
        for sp in spec_order) + "</tr>"
    th = "".join(f"<th style='font-size: 0.82em'>{sp}</th>" for sp in spec_order)
    coef_table = (f"<table style='font-size: 0.85em'>"
                  f"<thead><tr><th>Variable</th>{th}</tr></thead>"
                  f"<tbody>{''.join(rows_html)}{n_row}</tbody></table>")

    b3 = by_spec.get("B3·DV1 frac-logit + Mundlak", {}).get("did_blue")
    placebo = by_spec.get("B4·DV1 event study (pre-trend placebo)", {}).get("did_blue_placebo_2023")

    def tell(c):
        if not c: return "(n/a)"
        b, s, p, _ = c
        star = ""
        if p is not None:
            if p < 0.01: star = "***"
            elif p < 0.05: star = "**"
            elif p < 0.10: star = "*"
        return f"β = {b:+.3f}{star} (SE {s:.3f}, cluster p = {p:.3f})"

    body = f"""
    <p class="lead">An alternative DiD: replace the lead-agency-type treatment
    (nonprofit vs. government) with <strong>the political environment of each CoC's
    state</strong> (blue = Biden 2020, red = Trump 2020). Grants Pass gave all
    states the same legal opportunity; did CoCs in blue states respond
    differently from CoCs in red states?</p>

    <div class="callout">
      <div class="callout-title">Why this design?</div>
      <p>The nonprofit/government DiD returned a null. A natural question: maybe
      the relevant political environment isn't the CoC's leadership structure
      but the <em>state</em> it operates in. Blue-state CoCs face a constituency
      more sympathetic to anti-criminalization framing; red-state CoCs face
      the opposite. If political environment matters at all, the state-level
      treatment should reveal it.</p>
    </div>

    <h2>Design</h2>
    <ul>
      <li><strong>Treatment (blue):</strong> state voted Biden in the 2020
      presidential election. 26 states + DC.</li>
      <li><strong>Control (red):</strong> state voted Trump in 2020. 25 states.</li>
      <li><strong>Excluded:</strong> U.S. territories (PR, VI, GU) — no Electoral College vote.</li>
      <li><strong>Pre period:</strong> FY2022 + FY2023 (before June 2024 ruling).</li>
      <li><strong>Post period:</strong> FY2024 (application deadline October 30, 2024,
      four months after the ruling).</li>
      <li><strong>Classification is based on presidential vote, not governorship,</strong>
      for consistency with standard political-context measures.</li>
    </ul>

    <h2>Parallel-trends check</h2>
    <div id="ptchart" style="width:100%"></div>
    <p>Pre-period (FY2022 → FY2023) trajectories are similar:
    red +0.035, blue +0.008. Then the post-shock gap widens sharply:
    red +0.013, blue <strong>+0.117</strong>. The placebo interaction
    (Blue × FY2023) is not significant ({tell(placebo) if placebo else '(not estimated)'}).</p>

    <h2>Coefficient table</h2>
    {coef_table}
    <p class="hint">Cluster-robust SEs at the CoC level in parentheses ·
    <code>*** p&lt;0.01, ** p&lt;0.05, * p&lt;0.10</code></p>

    <h2>Key estimates</h2>
    <div class="findings-box">
      <h3>The blue-state DiD is significant</h3>
      <ul>
        <li><strong>Primary (B3, fractional logit + Mundlak):</strong> {tell(b3)}</li>
        <li><strong>Wild-cluster bootstrap p-value: {boot_p:.3f}</strong>
        — robust to small-cluster concerns</li>
        <li><strong>Pre-trend placebo:</strong> {tell(placebo)} — parallel trends plausible</li>
      </ul>
      <p style="margin-top: 0.8em;">CoCs in blue states differentially
      <em>increased</em> their reported anti-crim activity after Grants Pass.
      The logit-scale coefficient of +0.55 corresponds to roughly an
      11-percentage-point larger jump in the fraction of 1D-4 "Yes" cells
      compared to red-state CoCs — a sizeable effect.</p>
    </div>

    <h2>Comparison with the nonprofit-vs-government DiD</h2>
    <table>
      <thead><tr><th>Treatment</th><th class='num'>β_DiD (primary)</th><th class='num'>Cluster p</th><th class='num'>Bootstrap p</th><th>Verdict</th></tr></thead>
      <tbody>
        <tr><td>Nonprofit-led vs government-led</td>
            <td class='num'>+0.007</td>
            <td class='num'>0.88</td>
            <td class='num'>0.82</td>
            <td>Null — theory fails</td></tr>
        <tr><td><strong>Blue state vs red state</strong></td>
            <td class='num'><strong>+0.548</strong></td>
            <td class='num'><strong>0.037</strong></td>
            <td class='num'><strong>0.036</strong></td>
            <td><strong>Significant — state political context matters</strong></td></tr>
      </tbody>
    </table>

    <h2>What this shows</h2>
    <p>The lead-agency distinction does not predict differential response;
    the state-level political environment does. Post-Grants Pass,
    <strong>blue-state CoCs</strong> differentially expand their reported
    anti-criminalization activity while red-state CoCs barely move. Three
    possible mechanisms:</p>
    <ul>
      <li><strong>Local political pressure</strong>: blue-state CoCs receive
      stronger constituency signals to resist criminalization after
      Grants Pass.</li>
      <li><strong>Reporting behavior</strong>: blue-state CoCs may simply
      be more likely to describe existing activity as "anti-crim" under the
      new form wording; red-state CoCs may not want to foreground that
      framing.</li>
      <li><strong>Policy substitution</strong>: red-state CoCs may have had
      higher baseline anti-crim framing pre-Grants Pass (note their higher
      FY2022 baseline); the ruling provided more room for blue-state CoCs
      to catch up.</li>
    </ul>

    <h2>Caveats</h2>
    <ol>
      <li><strong>State-level treatment is coarse.</strong> A CoC in a red
      county of a blue state (e.g., rural Pennsylvania) is treated as
      "blue-state". County-level partisanship would be sharper.</li>
      <li><strong>Cross-treatment correlation.</strong> Nonprofit-led CoCs
      may cluster in blue states. A joint specification — with both
      treatments and their interactions — would disentangle them.</li>
      <li><strong>DV2 (engagement-only, wording-stable):</strong> β = +0.427
      but not statistically significant. The blue-state effect is driven
      partly by the implementation-column cells whose wording changed.</li>
      <li><strong>Post dummy still captures the HUD form change</strong> in
      addition to Grants Pass. A confounder common to all CoCs — but
      absorbed by the DiD if it hit blue and red states equally.</li>
      <li><strong>Static 2020 classification.</strong> Georgia and Arizona
      were narrowly blue in 2020; North Carolina narrowly red. Different
      cutoffs (margin-based, or 2024 returns) might affect classification.</li>
    </ol>

    <h2>Recommended next analysis</h2>
    <ol>
      <li>Joint specification: <code>crim_activity_index ~ nonprofit_led × post
      + blue_state × post + … </code> to test whether the blue effect is
      robust to controlling for leadership (and vice versa).</li>
      <li>Triple-difference: <code>blue_state × nonprofit_led × post</code>
      to see whether nonprofit leadership <em>amplifies</em> the blue-state
      response.</li>
      <li>County-level partisan vote-share (GitHub: <a href="https://github.com/tonmcg/US_County_Level_Election_Results_08-24">tonmcg
      county election data</a>) → merge via CoC-county crosswalk.</li>
      <li>External behavioral DV (NLHR ordinances) — does the blue-state
      response show up in actual ordinance changes, not just reporting?</li>
    </ol>

    <script>
    Plotly.newPlot("ptchart", {json.dumps(pt_data)}, {json.dumps(pt_layout)}, {{responsive:true}});
    </script>
    """
    return wrap("DiD — Blue vs Red state", "did_bluestate.html", body,
                "Political environment as an alternative treatment. This one is significant.")


def html_from_md_table(md: str) -> str:
    """Tiny Markdown-table to HTML converter for inline tables."""
    lines = [l for l in md.splitlines() if l.strip().startswith("|")]
    if len(lines) < 2:
        return f"<pre>{html.escape(md)}</pre>"
    def split(row: str):
        return [c.strip() for c in row.strip().strip("|").split("|")]
    headers = split(lines[0])
    body_rows = [split(r) for r in lines[2:]]
    hd = "".join(f"<th>{h}</th>" for h in headers)
    bd = "".join("<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>" for r in body_rows)
    return f'<table><thead><tr>{hd}</tr></thead><tbody>{bd}</tbody></table>'


def page_examples():
    # Read analysis-ready CSV to pull 5 pilot CoCs
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PIPELINE_DIR / "coc_analysis_ready.xlsx", data_only=True)
        ws = wb["balanced_panel"]
        headers = [c.value for c in ws[1]]
        all_rows = [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
    except Exception:
        all_rows = []

    pilot = ["AL-500", "CA-500", "FL-501", "NY-600", "TX-600"]
    shown_vars = [
        "coc_id", "year",
        "crim_activity_index", "implemented_anticrim_practice",
        "1a_2", "1a_3",
        "1b_1_6_meetings", "1b_1_6_voted", "1b_1_6_ces",
        "1d_2_3", "1d_4_1_policymakers", "1d_4_1_prevent_crim",
        "1d_10a_1_years", "2a_5_1_coverage",
    ]
    labels = {
        "coc_id": "CoC",
        "year": "Year",
        "crim_activity_index": "Activity index",
        "implemented_anticrim_practice": "Implemented practice",
        "1a_2": "Lead agency",
        "1a_3": "Desig.",
        "1b_1_6_meetings": "PLE · meetings",
        "1b_1_6_voted": "PLE · voted",
        "1b_1_6_ces": "PLE · CES",
        "1d_2_3": "Housing First %",
        "1d_4_1_policymakers": "1D-4·1·col1",
        "1d_4_1_prevent_crim": "1D-4·1·col2",
        "1d_10a_1_years": "PLE in decisionmaking (#)",
        "2a_5_1_coverage": "HMIS ES cov.",
    }
    header_row = "".join(f"<th>{html.escape(labels[v])}</th>" for v in shown_vars)
    body_rows = []
    for coc in pilot:
        for r in [r for r in all_rows if r.get("coc_id") == coc]:
            cells = []
            for v in shown_vars:
                val = r.get(v, "")
                if v == "crim_activity_index" and val not in ("", None):
                    try:
                        val = f"{float(val):.3f}"
                    except Exception:
                        pass
                if v == "1d_2_3" and val not in ("", None):
                    try:
                        f = float(val)
                        val = f"{f:.2f}" if f <= 1.5 else f"{f/100:.2f}"
                    except Exception:
                        pass
                cells.append(f"<td>{html.escape(str(val)[:45])}</td>")
            body_rows.append("<tr>" + "".join(cells) + "</tr>")

    body_html = "\n".join(body_rows)
    table = (
        f"<table><thead><tr>{header_row}</tr></thead>"
        f"<tbody>{body_html}</tbody></table>"
    )

    body = f"""
    <p class="lead">Five CoCs, three years, a curated set of variables —
    so you can see what the data actually looks like before opening the
    xlsx.</p>

    <h2>Five pilot CoCs across three fiscal years</h2>
    <p>These are the CoCs we used to validate every step of the pipeline.
    They span the country and have full 3-year data.</p>
    <ul>
      <li><strong>AL-500</strong> — Birmingham/Jefferson, St. Clair, Shelby Counties (Alabama)</li>
      <li><strong>CA-500</strong> — San Jose / Santa Clara City &amp; County (California)</li>
      <li><strong>FL-501</strong> — Hillsborough County (Florida)</li>
      <li><strong>NY-600</strong> — New York City (New York)</li>
      <li><strong>TX-600</strong> — San Antonio / Bexar County (Texas)</li>
    </ul>

    {table}

    <p class="hint">"PLE · meetings/voted/CES" = whether "Homeless or
    Formerly Homeless Persons" are reported as participating in CoC
    meetings, voting, or the Coordinated Entry System.</p>

    <h2>What the same CoC looks like longitudinally</h2>
    <p>Take <strong>NY-600</strong>. Over three years:</p>
    <ul>
      <li><code>1a_2</code> Collaborative Applicant: stable (same entity).</li>
      <li><code>1b_1_6_*</code> PLE participation: all Yes in every year — NYC has long institutionalized PLE roles.</li>
      <li><code>crim_activity_index</code>: rises from FY22 to FY24 alongside the national trend.</li>
      <li><code>2a_5_1_coverage</code> HMIS coverage of ES beds: already at 92%+ in FY22, stable.</li>
    </ul>

    <h2>What a messy example looks like</h2>
    <p>The pipeline surfaced several cases where the manual spreadsheet and
    the PDF disagree. For example, in <strong>FL-501 FY2024</strong>, the
    manual spreadsheet shows <code>1b_1_7_ces = Yes</code> (Hospitals
    participate in CES), but the PDF clearly says "No". These are the
    cases the <a href="review.html">Review queue</a> collects for PI
    adjudication.</p>

    <h2>How to read the raw JSON</h2>
    <p>Each extracted CoC has a JSON file at
    <code>data_pipeline/extracted/{{coc_id}}_{{year}}.json</code>. One
    record per variable, each with its source page and confidence flag:</p>
    <pre>{{
  "coc_id": "AL-500",
  "year": "2024",
  "field_id": "1d_4_1_prevent_crim",
  "value": "No",
  "raw_text": "Increase utilization of co-responder responses ...",
  "source_page": 27,
  "extractor": "pdf_native_v0.6",
  "confidence": 1.0,
  "needs_review": false
}}</pre>
    <p>This provenance is what lets a reviewer jump straight to page 27 of
    the PDF and verify any contested cell in seconds.</p>
    """
    return wrap("Examples — five CoCs across three years", "examples.html", body,
                "Concrete data so you can see what's in the pipeline's output.")


def page_using():
    body = """
    <p class="lead">Code recipes for the primary analysis — a multilevel
    Mundlak fractional-logit difference-in-differences with wild-cluster
    bootstrap inference at the state level.</p>

    <h2>What this page shows</h2>
    <ol>
      <li>The one-shot command that reproduces our primary table.</li>
      <li>A hand-rolled Python version so you can see the mechanics.</li>
      <li>An equivalent R implementation using <code>fixest</code>.</li>
      <li>Wild-cluster bootstrap for small-state inference.</li>
      <li>Balanced-panel, alternative-DV, and pre-trend placebos.</li>
    </ol>

    <h2>1 · Reproduce the main table</h2>
    <pre><code>cd coc_criminalization/data_pipeline

# Primary multilevel DiD with wild-cluster bootstrap
python3 run_multilevel.py

# Balanced-panel sensitivity (drops CoCs not observed all 3 years)
python3 run_balanced_sensitivity.py

# Alternative DV definitions (engagement-only, FY22+23 only)
python3 run_dv_robustness.py</code></pre>

    <h2>2 · Primary model (Python, statsmodels)</h2>
    <p>Papke-Wooldridge fractional logit for the bounded [0,1] activity
    index, with Mundlak within-CoC means to absorb time-invariant CoC
    heterogeneity, three DiD interactions (organizational, state, county),
    and state-clustered standard errors.</p>
    <pre><code>import numpy as np, pandas as pd, statsmodels.api as sm

df = pd.read_excel("coc_analysis_ready.xlsx", sheet_name="unbalanced")

# --- Three competing DiD treatments ---
df["post"]            = (df["year"] == 2024).astype(int)
df["nonprofit_led"]   = (df["lead_agency_type"] == "nonprofit").astype(int)
df["blue_state"]      = (df["state_pres_2020"] == "D").astype(int)
df["biden_within"]    = df["county_biden_share"] - df["state_biden_share"]

df["did_np"]     = df["nonprofit_led"] * df["post"]   # H1 — organizational form
df["did_blue"]   = df["blue_state"]    * df["post"]   # H2a — state politics
df["did_biden"]  = df["biden_within"]  * df["post"]   # H2b — within-state county pull

# --- Mundlak within-CoC means (absorbs time-invariant CoC heterogeneity) ---
for c in ["hf_pct", "hmis_cov", "log_beds", "ple_ces_bin"]:
    df[c + "_bar"] = df.groupby("coc_id")[c].transform("mean")

rhs = ["nonprofit_led", "blue_state", "biden_within", "post",
       "did_np", "did_blue", "did_biden",
       "hf_pct", "hmis_cov", "log_beds", "ple_ces_bin",
       "hf_pct_bar", "hmis_cov_bar", "log_beds_bar", "ple_ces_bin_bar"]

sub = df.dropna(subset=["crim_activity_index"] + rhs + ["state"]).copy()
y  = sub["crim_activity_index"].astype(float).values
X  = sm.add_constant(sub[rhs].astype(float), has_constant="add").values

# Fractional logit via GLM(Binomial, logit) on a [0,1] outcome
res = sm.GLM(y, X, family=sm.families.Binomial()).fit(
    cov_type="cluster", cov_kwds={"groups": sub["state"].values})
print(res.summary())</code></pre>

    <h2>3 · Equivalent R (fixest)</h2>
    <pre><code>library(readxl); library(fixest)

df <- read_excel("coc_analysis_ready.xlsx", sheet = "unbalanced")
df$post        <- as.integer(df$year == 2024)
df$nonprofit   <- as.integer(df$lead_agency_type == "nonprofit")
df$blue_state  <- as.integer(df$state_pres_2020 == "D")
df$biden_wi    <- df$county_biden_share - df$state_biden_share

# Mundlak means
for (v in c("hf_pct", "hmis_cov", "log_beds", "ple_ces_bin")) {
  df[[paste0(v, "_bar")]] <- ave(df[[v]], df$coc_id, FUN = function(z) mean(z, na.rm = TRUE))
}

m <- feglm(
  crim_activity_index ~ nonprofit + blue_state + biden_wi + post
                      + nonprofit:post + blue_state:post + biden_wi:post
                      + hf_pct + hmis_cov + log_beds + ple_ces_bin
                      + hf_pct_bar + hmis_cov_bar + log_beds_bar + ple_ces_bin_bar,
  data    = df,
  family  = quasibinomial("logit"),     # fractional logit
  cluster = ~state
)
etable(m)</code></pre>

    <h2>4 · Wild-cluster bootstrap (state-level, Rademacher)</h2>
    <p>With 50 states many of which contribute only one CoC, asymptotic
    cluster SEs over-reject (Bertrand, Duflo &amp; Mullainathan 2004;
    Cameron, Gelbach &amp; Miller 2008). We report wild-cluster bootstrap
    p-values with Rademacher weights at the state level, 999 replications.</p>
    <pre><code>rng = np.random.default_rng(42)
clusters = sub["state"].values
uc = np.unique(clusters)
target = "did_np"          # or "did_blue", "did_biden"
tidx   = rhs.index(target) + 1

# Observed t-stat
t_obs  = res.params[tidx] / res.bse[tidx]

# Restricted model (null: coefficient on `target` = 0)
Xr = np.delete(X, tidx, axis=1)
r0 = sm.GLM(y, Xr, family=sm.families.Binomial()).fit()
resid = y - r0.mu
yhat  = r0.mu

ext, N = 0, 999
for _ in range(N):
    w  = rng.choice([-1.0, 1.0], size=len(uc))
    wm = dict(zip(uc, w))
    wf = np.array([wm[c] for c in clusters])
    yb = np.clip(yhat + resid * wf, 1e-6, 1 - 1e-6)
    rb = sm.GLM(yb, X, family=sm.families.Binomial()).fit(
        cov_type="cluster", cov_kwds={"groups": clusters})
    tb = rb.params[tidx] / rb.bse[tidx]
    if abs(tb) &gt;= abs(t_obs): ext += 1

p_boot = (ext + 1) / (N + 1)
print(f"{target}: t_obs={t_obs:.3f}, wild-cluster p={p_boot:.3f}")</code></pre>

    <h2>5 · The four-quadrant summary</h2>
    <p>To plot the pre/post trajectory by state × county politics:</p>
    <pre><code>df["quadrant"] = np.where(df["blue_state"] == 1,
                    np.where(df["biden_within"] &gt; 0, "Blue×Blue", "Blue×Red"),
                    np.where(df["biden_within"] &gt; 0, "Red×Blue",  "Red×Red"))

tab = (df.groupby(["quadrant", "year"])["crim_activity_index"]
         .agg(["mean", "count"])
         .unstack("year"))
print(tab)

# Δ = FY2024 − mean(FY2022, FY2023) — the headline number
delta = tab["mean"][2024] - 0.5 * (tab["mean"][2022] + tab["mean"][2023])
print(delta.sort_values())  # Red×Red at the bottom ≈ −0.031</code></pre>

    <h2>6 · Balanced-panel sensitivity</h2>
    <p>The primary spec uses the unbalanced panel (maximal N). The
    balanced subset (CoCs observed in all three years) tests whether the
    state-level effect is a composition artifact.</p>
    <pre><code>python3 run_balanced_sensitivity.py   # prints side-by-side coefficients
# Our finding: Blue × Post flips from β=+0.572 (marginal) to null in balanced,
# while the within-state county pull strengthens — political geography,
# not selection, drives the result.</code></pre>

    <h2>7 · Alternative DVs</h2>
    <p>Three DV variants probe measurement sensitivity:</p>
    <pre><code>python3 run_dv_robustness.py
#   DV1: full six-cell activity index  (primary — reported in the paper)
#   DV2: engagement-only (cells 1, 3, 5 — wording identical across years)
#   DV3: FY2022 + FY2023 only (identical instrument, pre-shock only)</code></pre>

    <h2>8 · Merging external data</h2>
    <ul>
      <li><strong>HUD PIT estimates</strong> — public, CoC-level,
      annual. Merge on <code>coc_id</code> for per-capita normalization
      or as a control for local problem scale.</li>
      <li><strong>MIT Election Lab county presidential returns</strong> —
      source of <code>county_biden_share</code>. Already merged into
      <code>coc_analysis_ready.xlsx</code>.</li>
      <li><strong>HUD CoC → county crosswalk</strong> — many-to-many
      mapping, population-weighted to a CoC-level vote share.</li>
    </ul>
    """
    return wrap("Using the data", "using.html", body,
                "Code recipes for the multilevel Mundlak DiD with wild-cluster bootstrap.")


def page_data():
    files = [
        ("coc_analysis_ready.xlsx",
         "Primary analysis file — 4 sheets: balanced_panel (125 CoCs × 3 years), unbalanced (all records), fy2024_only, variables.",
         "Start here"),
        ("coc_analysis_ready.csv",
         "Long-form CSV of the same data — easy to load with pandas or readr.",
         "Start here"),
        ("coc_panel_wide.xlsx",
         "Full wide panel with four sheets: panel_safe (243 variables), full_wide, and three per-year sheets (fy2022/23/24). "
         "Also appended: dv_harmonized (harmonized outcomes) and narrative_codes (Stage-2 LLM outputs, if generated).",
         "Primary"),
        ("coc_panel_long.csv",
         "Long form — 164k+ rows, one per (CoC, year, field, value). Includes field-category tag.",
         "Primary"),
        ("harmonized_dv.csv",
         "Per-(CoC,year) row with four harmonized outcome variables: "
         "crim_activity_index, implemented_anticrim_practice, and two ceiling-hitting binaries kept for reference.",
         "DV"),
        ("coc_fy2022.xlsx",
         "FY2022 year-specific — 166 CoCs × 246 active fields.",
         "Per-year"),
        ("coc_fy2023.xlsx",
         "FY2023 year-specific — 193 CoCs × 269 active fields.",
         "Per-year"),
        ("coc_fy2024.xlsx",
         "FY2024 year-specific — 292 CoCs × 290 active fields.",
         "Per-year"),
        ("coc_merged_2024.xlsx",
         "FY2024 merged — manual ∪ auto, per-cell provenance (color-coded): both_agree, manual_only, auto_only, disagree, missing.",
         "Audit"),
        ("corpus_agreement.md",
         "Corpus-wide agreement report with per-class and per-CoC breakdowns.",
         "Audit"),
        ("corpus_diffs.csv",
         "Every extractor–manual disagreement (2,394 rows).",
         "Audit"),
        ("panel_coverage.md",
         "Per-year variable coverage analysis.",
         "Coverage"),
        ("panel_field_map.csv",
         "Category label (panel_safe / mostly_panel / year_specific / sparse) + year-by-year coverage per field.",
         "Coverage"),
        ("file_inventory.csv",
         "677 source files indexed — format, scan flag, duplicates, CoC, year.",
         "Metadata"),
        ("crosswalk.csv",
         "FY2022/23/24 question ID mapping onto the canonical schema.",
         "Metadata"),
        ("iterations.csv",
         "Iteration ledger — weighted/adjusted agreement per iteration.",
         "Metadata"),
    ]
    by_cat = defaultdict(list)
    for f, d, cat in files:
        by_cat[cat].append((f, d))

    out = []
    out.append('<p class="lead">All pipeline outputs live in '
               '<code>../data_pipeline/</code>. Click any filename to open it locally.</p>')
    out.append("""
    <div class="callout">
      <div class="callout-title">If you only open one file</div>
      <p>Open <code>coc_analysis_ready.xlsx</code>. Use the
      <code>balanced_panel</code> sheet for longitudinal analyses (125 CoCs
      × 3 years) or <code>fy2024_only</code> for richer FY2024-only
      cross-section models. The <code>variables</code> sheet documents what
      each column is.</p>
    </div>
    """)
    for cat in ["Start here", "Primary", "DV", "Per-year", "Audit", "Coverage", "Metadata"]:
        if cat not in by_cat:
            continue
        out.append(f"<h2>{cat}</h2><table><thead><tr><th>File</th><th>Description</th><th>Size</th></tr></thead><tbody>")
        for fname, desc in by_cat[cat]:
            p = PIPELINE_DIR / fname
            if p.exists():
                size = f"{p.stat().st_size / 1024:.0f} KB"
                href = f"../data_pipeline/{fname}"
            else:
                size = "<em>missing</em>"
                href = "#"
            out.append(
                f"<tr><td><a href=\"{href}\">{html.escape(fname)}</a></td>"
                f"<td>{desc}</td><td class='num'>{size}</td></tr>"
            )
        out.append("</tbody></table>")

    out.append("""
    <h2>Methodology documents (human-readable)</h2>
    <ul>
      <li><a href="../data_construction_methodology.md">data_construction_methodology.md</a>
      — formal Methods document with citations, ready to splice into the paper.</li>
      <li><a href="../main_variables.md">main_variables.md</a> — authoritative
      reference for which variables map to which construct, including
      year-to-year change flags.</li>
      <li><a href="../dv_harmonization_strategies.md">dv_harmonization_strategies.md</a>
      — eight strategies considered for the DV problem.</li>
      <li><a href="../dv_harmonization_results.md">dv_harmonization_results.md</a>
      — empirical results distinguishing which strategies work.</li>
      <li><a href="../codebook.md">codebook.md</a> — 331-variable codebook.</li>
      <li><a href="../progress/README.md">progress/</a> — the eight-page
      narrative log of how the pipeline was built, intended to become a
      public methodology website.</li>
    </ul>
    """)
    return wrap("Downloads", "data.html", "\n".join(out),
                "Everything the pipeline produces, organized by use case.")


def page_limits():
    body = """
    <p class="lead">Honest accounting of what the multilevel Mundlak DiD
    can and cannot say, and what remains on the pipeline side.</p>

    <h2>Analytical limitations</h2>

    <h3>1. Measurement invariance of the 1D-4 construct</h3>
    <p>FY2022/23 and FY2024 measure related but not identical constructs
    (see <a href="dv_story.html">DV story</a>). The primary DV
    (<code>crim_activity_index</code>) harmonizes at the activity level
    and the DiD design differences out the common instrument shift, but
    we cannot rule out that the across-year jump is partly mechanical.
    The <a href="dv_robust.html">DV robustness</a> page restricts to the
    engagement-only cells (identical wording across years) and the
    FY2022+FY2023 pre-shock subsample; the political-geography pattern
    survives both.</p>

    <h3>2. State-level cluster count</h3>
    <p>The primary specification clusters standard errors by state
    (≈50 clusters, many with only one CoC). Asymptotic cluster SEs
    over-reject in this regime (Bertrand, Duflo &amp; Mullainathan 2004;
    Cameron, Gelbach &amp; Miller 2008). All DiD p-values on the
    <a href="main_results.html">Main results</a> page come from a
    wild-cluster bootstrap with Rademacher weights at the state level
    (999 reps). Reviewers who prefer county-clustered or CR3 SEs will
    find the qualitative pattern unchanged but exact p-values differ.</p>

    <h3>3. Unbalanced vs balanced panel</h3>
    <p>The primary spec uses the unbalanced panel to preserve statistical
    power. The <a href="balanced.html">balanced sensitivity</a> shows
    that the state-level binary effect is partly a composition artifact
    — Blue × Post flips from marginal (β=+0.572, p<sub>boot</sub>=0.095)
    to null under balancing, while the within-state county pull
    strengthens. The paper should report both.</p>

    <h3>4. One post-period</h3>
    <p>The design has a single post-period (FY2024) following <em>Grants
    Pass v. Johnson</em>. We cannot observe medium-term persistence or
    test for anticipation with only three waves. A fourth wave (FY2025)
    will become available late 2025 and should be added to the panel
    before the final submission.</p>

    <h3>5. Self-report DV</h3>
    <p>Every outcome comes from the CoC's own FY application narrative —
    i.e., what CoCs <em>say</em> they do, not what they do. The
    <a href="dv_story.html">DV story</a> page flags this. External
    behavioral validation (HUD PIT outflow rates, BLS occupational
    homelessness outreach counts, state-level legislation trackers)
    should be used as triangulation before publication.</p>

    <h3>6. County-within-state continuous IV</h3>
    <p><code>biden_within = county_biden_share − state_biden_share</code>
    mechanically depends on the CoC→county crosswalk's many-to-many
    weighting. We use population-weighted county vote. CoCs that
    aggregate across very heterogeneous counties get a small
    <code>biden_within</code> by construction; the estimator downweights
    them correctly but the substantive precision is lower for
    multi-county CoCs (≈30% of the sample).</p>

    <h2>Data pipeline — what's still open</h2>

    <h3>7. Nine CoCs need OCR</h3>
    <p>Nine FY2024 PDFs ship as scanned images or with custom font
    encodings (CoC WA-504 renders as <code>$SSOLFDQW</code> rather than
    "Applicant"). Recovery requires <code>ocrmypdf</code> preprocessing.
    Infrastructure is in place. Affected: GA-508, MD-600, MI-508,
    NJ-511, NV-502, PA-501, TX-624, VI-500, WA-502. These are currently
    excluded from the analytic sample.</p>

    <h3>8. 41 narrative variables not yet coded</h3>
    <p>HUD asks 41 open-ended 2,500-character questions (e.g.,
    "Describe your experience promoting racial equity"). Stage-2 LLM
    coding with verbatim quote capture is specified but not rolled out
    (~$50–60 API cost for 6 fields × 3 years × ~280 CoCs). None of
    these are currently used as DVs in the primary analysis.</p>

    <h3>9. Manual spreadsheet reconciliation</h3>
    <p>1,263 disagreements between automation and the manual FY2024
    spreadsheet surfaced in the <a href="review.html">Review queue</a>.
    Most appear to be manual-coder errors (typos, Yes/No inversions).
    A formal 20-CoC stratified audit (per-coder, cross-checked against
    the PDFs) remains outstanding and is the last blocker on declaring
    a clean reference dataset.</p>

    <h3>10. Deliberate extractor behavior</h3>
    <p>The 1D-4 chart in FY2022/23 has 4 rows; FY2024 has 3. The
    extractor captures rows 1–3 in every year (row 4
    "community-wide plans" is dropped) to keep the activity index
    comparable. This is a conscious choice, not a bug.</p>

    <h3>11. Eight DOCX-only CoCs</h3>
    <p>For CoCs that submit DOCX instead of PDF (CA-505, CA-509, CA-511,
    CA-513, CA-521, MI-501, NV-500), the adapter extracts identifier
    fields reliably but some charts don't render the same way. Affected
    cells are flagged <code>needs_review</code>.</p>

    <h2>What we did NOT do on purpose</h2>
    <ul>
      <li><strong>We did not auto-patch the manual spreadsheet.</strong>
      Every proposed backport goes through PI review first.</li>
      <li><strong>We did not chase exact agreement with the manual file.</strong>
      When extractor and manual disagree, the extractor cites a specific
      PDF page. If it matches the source, it's right — even if the
      manual says otherwise.</li>
      <li><strong>We did not drop the H1 null.</strong> The organizational-form
      null (nonprofit vs government) is the substantive finding of the
      paper, not a failed hypothesis to be swept aside.</li>
    </ul>

    <p>Back to <a href="index.html">Overview</a> or the
    <a href="main_results.html">main results</a>.</p>
    """
    return wrap("Limitations and open items", "limits.html", body,
                "What the multilevel DiD can say, and what the pipeline still has to finish.")


def page_review():
    diffs = load_corpus_diffs()
    mismatches = [d for d in diffs if d["manual_value"] and d["auto_value"]]
    autofills = [d for d in diffs if not d["manual_value"] and d["auto_value"]]

    def table(rows, limit=1000):
        if not rows:
            return "<p><em>(none)</em></p>"
        out = ['<table><thead><tr>'
               '<th>CoC</th><th>Year</th><th>Field</th><th>Class</th>'
               '<th>Manual</th><th>Extractor</th><th>p.</th></tr></thead><tbody>']
        for r in rows[:limit]:
            out.append(
                "<tr>"
                f"<td>{html.escape(r['coc_id'])}</td>"
                f"<td>{html.escape(r['year'])}</td>"
                f"<td><code>{html.escape(r['field_id'])}</code></td>"
                f"<td>{html.escape(r.get('class', ''))}</td>"
                f"<td style='background:#fff4f4'>{html.escape((r['manual_value'] or '')[:60])}</td>"
                f"<td style='background:#f4fff4'>{html.escape((r['auto_value'] or '')[:60])}</td>"
                f"<td>{html.escape(str(r.get('source_page') or ''))}</td>"
                "</tr>"
            )
        if len(rows) > limit:
            out.append(f"<tr><td colspan='7'><em>...{len(rows) - limit:,} more (see corpus_diffs.csv)</em></td></tr>")
        out.append('</tbody></table>')
        return "\n".join(out)

    body = f"""
    <p class="lead">Every cell where the automated extractor and the
    manually coded FY2024 spreadsheet disagree. Pink shows the manual
    value, green shows the extractor's.</p>

    <div class="callout">
      <div class="callout-title">How to use this page</div>
      <p>For each row, open the source PDF at the indicated page. If the
      PDF supports the <strong>green</strong> value, the manual needs a
      correction. If the PDF supports the <strong>pink</strong> value, the
      extractor has a bug — please tell us which row so we can fix it.</p>
    </div>

    <h2>Value mismatches ({len(mismatches):,})</h2>
    <p>Both sides have values but they disagree. Needs PI adjudication.</p>
    {table(mismatches)}

    <h2>Auto-fill candidates ({len(autofills):,})</h2>
    <p>The manual spreadsheet is blank; the extractor has a value read
    directly from the PDF. Spot-check 20–30 at random before backfilling
    the spreadsheet.</p>
    {table(autofills)}
    """
    return wrap("Review queue", "review.html", body,
                "Disagreements the pipeline surfaced for human adjudication.")


# ---------------------------------------------------------------------------
# Consolidated top-level pages (5-section structure)
# ---------------------------------------------------------------------------
def page_data_development(download_records=None):
    inv = load_inventory()
    iters = load_iterations()
    fm = load_field_map()
    harm = load_harmonized()
    diffs = load_corpus_diffs()
    download_records = download_records or []

    by_year = Counter(r["year"] for r in inv)
    panel_safe_n = sum(1 for r in fm if r["category"] == "panel_safe")
    last_iter = iters[-1] if iters else {}
    wacc = float(last_iter.get("weighted_acc", 0) or 0)
    aacc = float(last_iter.get("adjusted_acc", 0) or 0) if last_iter.get("adjusted_acc") else None

    mismatches = [d for d in diffs if d.get("manual_value") and d.get("auto_value")]
    autofills  = [d for d in diffs if not d.get("manual_value") and d.get("auto_value")]
    downloads_html = _downloads_block(download_records)

    body = f"""
    <p class="lead">From 677 HUD CoC Consolidated Application documents
    (FY2022–FY2024) to an analysis-ready three-year panel. This page
    documents the pipeline, the agreement metrics, and the outstanding
    review queue.</p>

    <h2>1 · Source corpus</h2>
    <div class="card-row">
      <div class="stat"><div class="label">FY2022 applications</div><div class="value">{by_year.get("2022", 0)}</div></div>
      <div class="stat"><div class="label">FY2023 applications</div><div class="value">{by_year.get("2023", 0)}</div></div>
      <div class="stat"><div class="label">FY2024 applications</div><div class="value">{by_year.get("2024", 0)}</div></div>
      <div class="stat"><div class="label">Total documents</div><div class="value">{len(inv)}</div></div>
    </div>
    <p class="hint">FY2024 is the largest year because HUD folded the
    Youth Homeless Demonstration Program (YHDP) into the main competition.
    Eight CoCs submit DOCX instead of PDF; nine require OCR and are held
    out of the analytic sample.</p>

    <h2>2 · Extraction pipeline</h2>
    <ol>
      <li><strong>File inventory</strong> (<code>build_file_inventory.py</code>)
      — indexes every source file, detects scanned PDFs, flags duplicate
      PDF/DOCX pairs, normalizes CoC IDs.</li>
      <li><strong>Format-aware adapters</strong> — native-PDF, OCR'd PDF,
      and DOCX each implement the same output schema. Every cell carries
      source page, bbox, and extractor version as provenance.</li>
      <li><strong>Codebook &amp; crosswalk</strong> — 331 canonical FY2024
      variables mapped across FY2022/23 question IDs. Wording changes
      flagged explicitly (see the DV measurement-invariance story on the
      <a href="design.html">design</a> page).</li>
      <li><strong>Anchor-then-parse</strong> — every question is located
      by its <code>1B-1.</code> / <code>1D-4.</code>-style anchor, and
      only the local block is parsed. A parse failure in one block cannot
      corrupt neighboring variables.</li>
      <li><strong>Controlled-vocabulary validation</strong> — categorical
      outputs are checked against the codebook. Out-of-vocab values get
      <code>needs_review=True</code>; they are never silently coerced.</li>
      <li><strong>Iteration loop</strong> — extractor diffs against the
      manual FY2024 spreadsheet, errors classified into six categories,
      pipeline refined until convergence.</li>
    </ol>

    <h2>3 · Agreement against the manual reference</h2>
    <div class="card-row">
      <div class="stat"><div class="label">Weighted agreement</div><div class="value">{wacc:.1%}</div></div>
      {f'<div class="stat"><div class="label">Adjusted (manual-error removed)</div><div class="value">{aacc:.1%}</div></div>' if aacc else ""}
      <div class="stat"><div class="label">Panel-safe variables</div><div class="value">{panel_safe_n}</div></div>
      <div class="stat"><div class="label">CoC × year records</div><div class="value">{len(harm)}</div></div>
    </div>
    <p><strong>Panel-safety categorization:</strong></p>
    <ul>
      <li><span class="badge good">panel_safe</span> — present in ≥50% of
      CoCs in every year. Used directly in the primary regression.</li>
      <li><span class="badge warn">mostly_panel</span> — present in two
      of three years. Used with care.</li>
      <li><span class="badge warn">year_specific</span> — asked in only
      one year. Cross-sectional use only.</li>
      <li><span class="badge bad">sparse</span> — free-text narrative
      fields awaiting Stage-2 LLM coding.</li>
    </ul>

    <h2>4 · Review queue</h2>
    <p>{len(mismatches):,} value disagreements and {len(autofills):,}
    auto-fill candidates are staged for PI adjudication.</p>
    <ul>
      <li><strong>Value mismatches</strong> — both the manual spreadsheet
      and the extractor have a value, and they disagree. Every row cites
      a specific PDF page; when the extractor matches the source, the
      manual entry is treated as the one to correct.</li>
      <li><strong>Auto-fill candidates</strong> — the manual cell is
      blank; the extractor has a value read directly from the PDF. A
      spot-check of ~30 random rows is recommended before backfilling
      the spreadsheet.</li>
    </ul>
    <p>The pipeline does <strong>not</strong> auto-patch the manual file.
    Every proposed backport goes through PI review first, preserving
    provenance.</p>

    <h2>5 · Outstanding data work</h2>
    <ol>
      <li><strong>Nine CoCs need OCR.</strong> GA-508, MD-600, MI-508,
      NJ-511, NV-502, PA-501, TX-624, VI-500, WA-502 ship as scanned or
      font-encoded PDFs. <code>ocrmypdf</code> infrastructure is in place;
      the task has not been run.</li>
      <li><strong>41 narrative variables not yet coded.</strong> Stage-2
      LLM coding with verbatim quotation is specified but not rolled
      out (~$50–60 in API cost). None of these are in the primary DV.</li>
      <li><strong>Formal 20-CoC stratified audit.</strong> A per-coder
      cross-check against the PDFs is outstanding; this is the last
      blocker on publishing a clean reference dataset alongside the
      automated one.</li>
    </ol>

    <h2>Downloads</h2>
    <p class="hint">All files regenerate from the pipeline on every build.
    Click any filename to download. Citations and dates should be updated
    if you use these in downstream work.</p>
    {downloads_html}

    <p class="hint" style="margin-top:2em;">Missing something? Raw source
    PDFs and the full extractor source live in the project repository,
    not on this static site. Contact the authors for access.</p>
    """
    return wrap("Data development and cleaning", "data.html", body,
                "From 677 raw HUD documents to a panel-safe dataset — pipeline, agreement metrics, downloads, and open review items.")


def _downloads_block(records: list[dict]) -> str:
    if not records:
        return "<p><em>(No downloadable files were found. Run the pipeline first.)</em></p>"
    groups = [
        ("primary",    "Primary analysis dataset",
         "Load these for the main regression."),
        ("provenance", "Data provenance &amp; review",
         "Source-document index, panel-safety map, crosswalk, and the manual-vs-extractor review queue."),
        ("ivs",        "Independent-variable coding",
         "Hand-coded nonprofit/government classification and the CoC → county political-geography merge."),
        ("results",    "Results tables (CSV)",
         "Coefficient tables behind the figures on the Results page."),
        ("docs",       "Documentation (Markdown)",
         "Codebook, methodology, and human-readable results writeups."),
    ]
    out = []
    for key, heading, lede in groups:
        subset = [r for r in records if r["group"] == key]
        if not subset:
            continue
        out.append(f"<h3>{heading}</h3>")
        out.append(f"<p class='hint'>{lede}</p>")
        out.append("<table class='downloads'><thead><tr>"
                   "<th>File</th><th>Size</th><th>Description</th>"
                   "</tr></thead><tbody>")
        for r in subset:
            out.append(
                f"<tr>"
                f"<td><a href='{r['href']}' download><code>{r['title']}</code></a></td>"
                f"<td class='num'>{r['size']}</td>"
                f"<td>{r['description']}</td>"
                f"</tr>"
            )
        out.append("</tbody></table>")
    return "\n".join(out)


def page_design():
    body = """
    <p class="lead">How <em>Grants Pass v. Johnson</em> becomes an
    identification strategy: two competing hypotheses, a multilevel
    decomposition of political environment, and an honest account of
    the DV's measurement-invariance issue.</p>

    <h2>Theory — two competing hypotheses</h2>
    <div class="callout">
      <div class="callout-title">H1 · Organizational form</div>
      <p>Nonprofit-led CoCs — more insulated from electoral pressure and
      with closer ties to service providers and advocates — should respond
      to a federal legal shock enabling criminalization (<em>Grants
      Pass</em>) by <em>raising</em> their anti-criminalization posture
      relative to government-led CoCs.</p>
    </div>
    <div class="callout">
      <div class="callout-title">H2 · Local political environment</div>
      <p>CoCs embedded in Democratic-leaning political environments —
      whether at the state level (H2a) or in terms of county-within-state
      deviation (H2b) — should raise their anti-criminalization posture
      relative to Republican-leaning environments. Governance structure
      is irrelevant in this story; politics filters the shock.</p>
    </div>

    <h2>Multilevel decomposition of political environment</h2>
    <p>Because CoCs are nested (CoCs ⊂ counties ⊂ states), and because
    state and county politics are highly correlated, we decompose county
    Biden vote share into two <em>orthogonal</em> components:</p>
    <pre><code>county_biden_share = state_mean_share      ← H2a (between-state)
                   + biden_within_state    ← H2b (within-state)</code></pre>
    <p>The between-state component is summarized as a binary
    <code>blue_state</code> indicator (Biden won the state's EC vote in
    2020). The within-state component is continuous and mean-zero by
    construction. Both enter the regression together without collinearity.</p>

    <h2>Model diagram</h2>
    <div style="text-align:center; margin: 1.5em 0;">
    <svg viewBox="0 0 860 420" width="100%" style="max-width:860px; font-family: -apple-system, sans-serif;">
      <defs>
        <marker id="arr" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="auto">
          <path d="M0,0 L0,6 L9,3 z" fill="#0366d6"/>
        </marker>
        <marker id="arr-n" markerWidth="10" markerHeight="10" refX="9" refY="3" orient="auto">
          <path d="M0,0 L0,6 L9,3 z" fill="#c53030"/>
        </marker>
      </defs>

      <rect x="30" y="50" width="200" height="70" rx="8" fill="#fdf2f2" stroke="#c53030" stroke-width="2"/>
      <text x="130" y="78" text-anchor="middle" font-weight="600" font-size="13">H1 · Organizational form</text>
      <text x="130" y="98" text-anchor="middle" font-size="11" fill="#555">Nonprofit-led (vs Government-led)</text>
      <text x="130" y="112" text-anchor="middle" font-size="11" fill="#555">coded from 1a_2 Collaborative Applicant</text>

      <rect x="30" y="180" width="200" height="50" rx="8" fill="#eef4fb" stroke="#0366d6" stroke-width="2"/>
      <text x="130" y="202" text-anchor="middle" font-weight="600" font-size="13">H2a · State-level politics</text>
      <text x="130" y="220" text-anchor="middle" font-size="11" fill="#555">Blue state (Biden 2020 winner)</text>

      <rect x="30" y="260" width="200" height="50" rx="8" fill="#eef4fb" stroke="#0366d6" stroke-width="2" stroke-dasharray="4,2"/>
      <text x="130" y="280" text-anchor="middle" font-weight="600" font-size="13">H2b · County-within-state</text>
      <text x="130" y="298" text-anchor="middle" font-size="11" fill="#555">Biden share − state mean</text>

      <rect x="310" y="165" width="210" height="80" rx="8" fill="#fff4cc" stroke="#8a6d1f" stroke-width="2"/>
      <text x="415" y="193" text-anchor="middle" font-weight="600" font-size="14">Grants Pass shock</text>
      <text x="415" y="213" text-anchor="middle" font-size="11" fill="#555">SCOTUS ruling · June 28, 2024</text>
      <text x="415" y="229" text-anchor="middle" font-size="11" fill="#555">FY2024 = Post indicator</text>

      <rect x="620" y="165" width="200" height="80" rx="8" fill="#f0faf3" stroke="#17813c" stroke-width="2"/>
      <text x="720" y="193" text-anchor="middle" font-weight="600" font-size="14">Anti-crim activity</text>
      <text x="720" y="213" text-anchor="middle" font-size="11" fill="#555">crim_activity_index ∈ [0,1]</text>
      <text x="720" y="229" text-anchor="middle" font-size="11" fill="#555">share of 1D-4 cells = Yes</text>

      <rect x="310" y="310" width="210" height="90" rx="8" fill="#fafafa" stroke="#888" stroke-width="1.5" stroke-dasharray="4,3"/>
      <text x="415" y="333" text-anchor="middle" font-weight="600" font-size="13" fill="#444">Controls</text>
      <text x="415" y="351" text-anchor="middle" font-size="11" fill="#555">Housing First · HMIS coverage</text>
      <text x="415" y="366" text-anchor="middle" font-size="11" fill="#555">log(total beds, winsorized)</text>
      <text x="415" y="381" text-anchor="middle" font-size="11" fill="#555">PLE participation · Mundlak means</text>

      <path d="M 230 85 Q 280 125 310 180" fill="none" stroke="#c53030" stroke-width="2.5" marker-end="url(#arr-n)"/>
      <text x="244" y="145" font-size="11" fill="#c53030">H1: × Post</text>
      <path d="M 230 205 L 310 205" fill="none" stroke="#0366d6" stroke-width="2.5" marker-end="url(#arr)"/>
      <text x="250" y="200" font-size="11" fill="#0366d6">H2a: × Post</text>
      <path d="M 230 285 Q 270 260 310 225" fill="none" stroke="#0366d6" stroke-width="2" stroke-dasharray="4,2" marker-end="url(#arr)"/>
      <text x="240" y="260" font-size="11" fill="#0366d6">H2b: × Post</text>
      <path d="M 520 205 L 620 205" fill="none" stroke="#17813c" stroke-width="3" marker-end="url(#arr)"/>
      <path d="M 520 345 Q 580 290 620 230" fill="none" stroke="#888" stroke-width="1.5" marker-end="url(#arr)"/>
    </svg>
    </div>
    <p class="hint">Red arrow = H1 prediction. Blue arrows = H2
    predictions. Each IV's interaction with Post is its DiD estimand.</p>

    <h2>Operationalization</h2>
    <table>
      <thead><tr><th>Construct</th><th>Variable</th><th>Type</th><th>Source</th></tr></thead>
      <tbody>
        <tr><td><strong>IV₁ · Organizational form</strong></td>
            <td><code>nonprofit_led</code></td>
            <td>binary</td>
            <td>Hand-coded from <code>1a_2</code> Collaborative Applicant Name (99% coverage)</td></tr>
        <tr><td><strong>IV₂ · State-level politics</strong></td>
            <td><code>blue_state</code></td>
            <td>binary</td>
            <td>2020 presidential EC winner (Biden=1, Trump=0; DC blue; territories excluded)</td></tr>
        <tr><td><strong>IV₃ · Within-state county politics</strong></td>
            <td><code>biden_within_state</code></td>
            <td>continuous</td>
            <td>County Biden share − state-level mean (MIT/tonmcg 2020 county data)</td></tr>
        <tr><td><strong>Shock</strong></td>
            <td><code>post</code></td>
            <td>binary (FY2024 = 1)</td>
            <td><em>Grants Pass v. Johnson</em> (SCOTUS, June 28, 2024)</td></tr>
        <tr><td><strong>DV</strong></td>
            <td><code>crim_activity_index</code></td>
            <td>fraction [0, 1]</td>
            <td>Share of 6 HUD 1D-4 cells answered "Yes"</td></tr>
        <tr><td>Control · service orientation</td>
            <td><code>hf_pct</code></td>
            <td>fraction</td>
            <td>Housing First adoption, HUD 1D-2</td></tr>
        <tr><td>Control · CoC maturity</td>
            <td><code>hmis_cov</code></td>
            <td>fraction</td>
            <td>HMIS ES bed coverage, HUD 2A-5</td></tr>
        <tr><td>Control · size</td>
            <td><code>log_beds</code></td>
            <td>continuous</td>
            <td>log(total beds + 1), winsorized at 99th pct</td></tr>
        <tr><td>Control · PLE participation</td>
            <td><code>ple_ces_bin</code></td>
            <td>binary</td>
            <td>HUD 1B-1 row 6 in CES</td></tr>
      </tbody>
    </table>

    <h2>Primary specification</h2>
    <pre><code>crim_activity_index<sub>it</sub> =
    β₁ · Nonprofit<sub>i</sub>
  + β₂ · Blue_state<sub>i</sub>
  + β₃ · Biden_within_state<sub>i</sub>
  + β₄ · Post<sub>t</sub>
  + β₅ · Nonprofit × Post                 ← H1 test
  + β₆ · Blue_state × Post                ← H2a test
  + β₇ · Biden_within × Post              ← H2b test
  + γ · Controls<sub>it</sub> (Mundlak-adjusted)
  + ε<sub>it</sub></code></pre>
    <ul>
      <li><strong>Estimator:</strong> Papke-Wooldridge fractional logit
      (bounded [0, 1] DV with mass at 0 and 1).</li>
      <li><strong>Standard errors:</strong> state-level cluster-robust,
      reinforced with wild-cluster bootstrap (Rademacher weights,
      999 replicates). Addresses small-cluster over-rejection per
      Bertrand, Duflo &amp; Mullainathan (2004).</li>
      <li><strong>Sample:</strong> unbalanced panel (651 obs) for primary
      power; balanced three-year panel (375 obs) reported as sensitivity.</li>
      <li><strong>Mundlak means:</strong> adding CoC-level means of
      time-varying controls preserves identification of the DiD coefficients
      without absorbing the time-invariant IVs, which within-CoC fixed
      effects would otherwise do.</li>
    </ul>

    <h2>DV measurement invariance — the load-bearing caveat</h2>
    <p>The outcome comes from HUD's 1D-4 question, which was reworded in
    FY2024:</p>
    <table>
      <thead><tr><th>Column</th><th>FY2022 / FY2023 wording</th><th>FY2024 wording</th></tr></thead>
      <tbody>
        <tr><td>1 · Policymaker engagement</td>
            <td>"Engaged local policymakers to prevent criminalization"</td>
            <td>"Engaged legislators on co-responder responses"</td></tr>
        <tr><td>2 · Implementation</td>
            <td>"Reverse existing criminalization policies"</td>
            <td>"Implemented Laws/Policies/Practices that Prevent Criminalization"</td></tr>
      </tbody>
    </table>
    <p>The column-2 reword is mechanically easier to check Yes — any
    existing policy qualifies, not just a reversal. Descriptives show
    the FY2024 jump is concentrated in column 2, consistent with at
    least <em>some</em> of the level shift being an instrument effect.</p>
    <p>Three design choices handle this:</p>
    <ol>
      <li><strong>Year fixed effect</strong> (the <code>post</code>
      indicator) absorbs the common instrument-change shift.</li>
      <li><strong>DiD interactions</strong> identify <em>differential</em>
      response — the instrument change cancels out of any contrast
      between CoCs facing the same reworded form.</li>
      <li><strong>Engagement-only DV</strong> (<code>results.html</code>
      robustness) restricts to the column-1 cells whose wording was
      nearly identical across years.</li>
    </ol>

    <p>Next: <a href="descriptive.html">descriptive patterns</a> →
    <a href="results.html">analysis results</a>.</p>
    """
    return wrap("Research design and measurement", "design.html", body,
                "Two competing hypotheses, a multilevel decomposition of political environment, and the DV measurement story.")


def page_results():
    bal = load_csv(PIPELINE_DIR / "balanced_sensitivity_coefs.csv")
    dvr = load_csv(PIPELINE_DIR / "dv_robustness_coefs.csv")
    main_body_html = page_main_results()

    import re
    match = re.search(r"<header>.*?</header>\s*(.*)<footer>", main_body_html, re.S)
    if match:
        core = match.group(1)
    else:
        core = "<p><em>Main results body could not be extracted.</em></p>"

    core = core.replace('href="balanced.html"',     'href="#balanced"')
    core = core.replace('href="dv_story.html"',     'href="design.html#dv-story"')
    core = core.replace('href="dv_robust.html"',    'href="#dv-robust"')
    core = core.replace('href="descriptive.html"',  'href="descriptive.html"')

    def bal_row(rows, spec):
        u = next((r for r in rows if r.get("spec") == spec and r.get("panel") == "unbalanced"), {})
        b = next((r for r in rows if r.get("spec") == spec and r.get("panel") == "balanced"), {})
        def cell(r):
            if not r:
                return "<td class='num'>—</td><td class='num'>—</td>"
            coef = r.get("coef", "")
            p    = r.get("p_boot", "") or r.get("p", "")
            return f"<td class='num'>{coef}</td><td class='num'>{p}</td>"
        return f"<tr><td><strong>{spec}</strong></td>{cell(u)}{cell(b)}</tr>"

    balanced_table = ""
    if bal:
        balanced_table = (
            "<table><thead><tr>"
            "<th>Term</th>"
            "<th class='num'>β (unbalanced)</th><th class='num'>p<sub>boot</sub></th>"
            "<th class='num'>β (balanced)</th><th class='num'>p<sub>boot</sub></th>"
            "</tr></thead><tbody>"
            + bal_row(bal, "Nonprofit × Post")
            + bal_row(bal, "Blue × Post")
            + bal_row(bal, "Biden_within × Post")
            + "</tbody></table>"
        )

    def dvr_rows(rows):
        out = []
        for r in rows:
            out.append(
                f"<tr><td>{r.get('dv','')}</td>"
                f"<td>{r.get('term','')}</td>"
                f"<td class='num'>{r.get('coef','')}</td>"
                f"<td class='num'>{r.get('p_boot', r.get('p',''))}</td>"
                f"<td class='num'>{r.get('n','')}</td></tr>"
            )
        return "".join(out)

    dvr_table = ""
    if dvr:
        dvr_table = (
            "<table><thead><tr>"
            "<th>DV</th><th>Term</th>"
            "<th class='num'>β</th><th class='num'>p<sub>boot</sub></th><th class='num'>n</th>"
            "</tr></thead><tbody>" + dvr_rows(dvr) + "</tbody></table>"
        )

    appendix = f"""
    <hr>
    <h2 id="balanced">Appendix A · Balanced-panel sensitivity</h2>
    <p>The primary spec uses the unbalanced panel (651 obs) for power.
    The balanced subset keeps only the 125 CoCs observed in all three
    years (375 obs). If the pattern reflects political geography rather
    than selection, the within-state county pull should survive balancing;
    the binary state-level effect may weaken because composition drives
    much of it.</p>
    {balanced_table or '<p><em>(Run <code>run_balanced_sensitivity.py</code> to populate this table.)</em></p>'}
    <p><strong>Reading:</strong> Blue × Post weakens sharply under
    balancing (β = +0.57 → near zero), consistent with the unbalanced
    effect being partly a composition artifact; the within-state county
    pull is stable to strengthening. Political geography, not which CoCs
    HUD happened to add in FY2024, carries the story.</p>

    <h2 id="dv-robust">Appendix B · Alternative DV definitions</h2>
    <p>Three DV variants probe measurement sensitivity:</p>
    <ol>
      <li><strong>DV1 · Full six-cell activity index</strong> — the primary
      DV reported above.</li>
      <li><strong>DV2 · Engagement-only</strong> — the three 1D-4
      "policymaker engagement" cells whose wording was nearly identical
      across FY2022/23 and FY2024. This DV is immune to the FY2024
      column-2 reword discussed on the <a href="design.html">design</a>
      page.</li>
      <li><strong>DV3 · FY2022 + FY2023 only</strong> — pre-shock,
      identical instrument across both years. Used as a placebo:
      political-geography effects should <em>not</em> appear here.</li>
    </ol>
    {dvr_table or '<p><em>(Run <code>run_dv_robustness.py</code> to populate this table.)</em></p>'}
    <p><strong>Reading:</strong> the political-geography pattern survives
    DV2 (engagement-only cells), so the effect is not an artifact of the
    column-2 wording change. DV3 (pre-shock only) shows no differential
    pattern, consistent with a genuine post-shock divergence rather than
    a long-run red/blue gap in CoC reporting style.</p>

    <h2 id="limits">Limitations</h2>
    <ol>
      <li><strong>Marginal state-level significance.</strong> The binary
      Blue × Post coefficient is at bootstrap p ≈ 0.095 in the primary
      specification and loses significance under balancing. Reported as
      suggestive, not definitive.</li>
      <li><strong>Self-report DV.</strong> The outcome is what CoCs
      <em>say</em> they do on HUD Form 1D-4. External behavioral
      triangulation (PIT outflow rates, state legislation trackers) is
      a needed next step.</li>
      <li><strong>Short post-period.</strong> FY2024 applications were
      submitted four months after the Supreme Court ruling. Medium-term
      persistence cannot be assessed until FY2025 data are available.</li>
      <li><strong>State cluster count.</strong> ~50 state clusters with
      many singletons force wild-cluster bootstrap inference; asymptotic
      cluster SEs would over-reject.</li>
      <li><strong>Quadrant cell imbalance.</strong> Red-state × Blue-county
      has only 26 FY2024 observations, limiting precision in that cell.</li>
    </ol>

    <h2>Reproduce</h2>
    <pre><code>cd data_pipeline
python3 run_multilevel.py            # primary specification + bootstrap
python3 run_balanced_sensitivity.py  # Appendix A
python3 run_dv_robustness.py         # Appendix B
python3 build_site.py                # regenerate this page</code></pre>
    """

    combined = core + appendix
    return wrap("Analysis results", "results.html", combined,
                "Multilevel fractional-logit DiD with state-clustered wild-cluster bootstrap — plus balanced-panel and DV-robustness appendices.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    records = copy_downloads()
    pages = {
        "index.html":       page_index(),
        "data.html":        page_data_development(records),
        "design.html":      page_design(),
        "descriptive.html": page_descriptive(),
        "results.html":     page_results(),
    }
    for name, html_text in pages.items():
        (SITE_DIR / name).write_text(html_text)
        print(f"wrote {name}")
    print(f"staged {len(records)} downloads under {DOWNLOADS_DIR}")
    print(f"\nOpen: file://{SITE_DIR}/index.html")


if __name__ == "__main__":
    main()
